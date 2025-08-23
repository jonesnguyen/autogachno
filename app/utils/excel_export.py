from typing import List, Tuple
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import os
from tkinter import messagebox
import logging

from ..config import Config

logger = logging.getLogger(__name__)

def export_excel(data: List[Tuple[str, any, str]], name_dir: str) -> bool:
    """Xuất dữ liệu ra file Excel"""
    today = datetime.now().strftime("%H%M-%d-%m-%Y")
    try:
        export_dir = os.path.join(os.getcwd(), f"{Config.FOLDER_RESULT}\\{name_dir}")
        os.makedirs(export_dir, exist_ok=True)
        
        file_name = f"{today}.xlsx"
        file_path = os.path.join(export_dir, file_name)

        wb = Workbook()
        ws = wb.active
        
        # Định dạng header
        headers = ['STT', 'Số thuê bao', 'Số tiền', 'Ghi chú']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # Ghi dữ liệu
        for idx, (phone, amount, note) in enumerate(data, start=2):
            ws.cell(row=idx, column=1, value=idx - 1)
            ws.cell(row=idx, column=2, value=phone)
            ws.cell(row=idx, column=3, value=amount)
            ws.cell(row=idx, column=4, value=note)
            
        wb.save(file_path)
        
        # Hỏi người dùng có muốn mở file không
        # if messagebox.askyesno(Config.TITLE, f"Dữ liệu được lưu tại: \n{Config.FOLDER_RESULT}/{name_dir}/{today}"):
        #     try:
        #         os.startfile(file_path)
        #     except Exception as e:
        #         logger.warning(f"Không thể mở file Excel: {e}")
                
        return True
        
    except Exception as e:
        logger.error(f"Lỗi xuất Excel: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi xuất Excel: {e}")
        return False
