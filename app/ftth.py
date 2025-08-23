import os
import re
import time
import json 
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import pytz
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import StaleElementReferenceException, NoSuchElementException, TimeoutException
from cryptography.fernet import Fernet
import psycopg2

from bs4 import BeautifulSoup
import sys
import logging
import requests
import threading
from typing import List, Tuple, Optional, Dict, Any
from dataclasses import dataclass

# Cấu hình logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('app.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

@dataclass
class Config:
    """Cấu hình ứng dụng"""
    DRIVER_LINK: str = "https://kpp.bankplus.vn"
    FOLDER_RESULT: str = "ket_qua"
    TITLE: str = "Thông báo"
    CONFIG_FILE: str = "config.json"
    ICON_FILE: str = "viettelpay.ico"
    COPYRIGHT_KEY: bytes = b"h_ThisAAutoToolVjppro-CopyRight-ByCAOAC7690="
    STATUS_COMPLETE: str = "Đã xử lý"
    STATUS_INCOMPLETE: str = "Chưa xử lý"
    
    # Service types
    SERVICES = {
        'payment_internet': 'payment_internet',
        'payment_card': 'deb_cart', 
        'lookup_card': 'lookup_cart',
        'lookup_ftth': 'lookup_ftth',
        'payment_evn': 'deb_evn'
    }
    
    # API Configuration
    API_BASE_URL: str = os.getenv('API_BASE_URL', "http://127.0.0.1:8080")
    NODE_SERVER_URL: str = os.getenv('NODE_SERVER_URL', "http://127.0.0.1:5000")
    API_TIMEOUT: int = 10

# Global variables
stop_flag = False
driver = None
dbfiles = None
times_exits = {}
api_server_thread = None
automation_lock = threading.Lock()
AUTOMATION_MAX_RETRIES = 1  # Mặc định chỉ thử 1 lần cho mỗi mã
DIRECT_DB_MODE = True  # Chỉ xử lý trực tiếp với database, không phụ thuộc Node API
DB_DATABASE_URL = os.getenv('DATABASE_URL', 'postgres://postgres:123456@localhost:5432/autogachno')
DB_MAIN_USER_ID = os.getenv('MAIN_USER_ID', 'admin-local')
DB_MAIN_USER_EMAIL = os.getenv('MAIN_USER_EMAIL', 'Demodiemthu')

# Đăng nhập tự động khi mở trình duyệt
LOGIN_USERNAME = os.getenv('APP_LOGIN_USERNAME', '1000460100_VTP_00073_DB')
LOGIN_PASSWORD = os.getenv('APP_LOGIN_PASSWORD', '686886')

def db_ensure_user(user_id: str, email: str) -> None:
    try:
        with psycopg2.connect(DB_DATABASE_URL) as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT 1 FROM users WHERE id = %s", (user_id,))
                exists = cur.fetchone()
                if exists:
                    return
                cur.execute(
                    """
                    INSERT INTO users (id, email, first_name, last_name, role, status)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """,
                    (user_id, email, 'Admin', 'Local', 'admin', 'active')
                )
                print(f"[DB] Tạo user mặc định {email} ({user_id})")
    except Exception as e:
        print(f"[DB] Lỗi đảm bảo user tồn tại: {e}")


def send_callback_with_retry(order_id: str, code: str, status: str, amount: Any, notes: str, details: Optional[Dict[str, Any]] = None) -> bool:
    """Gửi callback với retry logic tối đa 3 lần."""
    print(f"🔄 [CALLBACK] Bắt đầu gửi callback cho mã {code} - Order: {order_id}")
    print(f"   📊 Status: {status}, Amount: {amount}, Notes: {notes}")
    
    for attempt in range(3):
        try:
            payload = {
                "orderId": order_id,
                "code": code,
                "status": status,
                "amount": str(amount) if isinstance(amount, (int, float)) else None,
                "notes": notes,
            }
            
            if details:
                payload["data"] = {
                    "type": "ftth_details",
                    "details": details,
                }
                print(f"   📋 Data: {details}")
            
            print(f"   📤 Lần thử {attempt + 1}/3: Gửi đến {Config.NODE_SERVER_URL}/api/automation/callback")
            response = requests.post(
                f"{Config.NODE_SERVER_URL}/api/automation/callback",
                json=payload,
                timeout=10,  # Tăng timeout
            )
            
            if response.status_code == 200:
                print(f"   ✅ Callback thành công cho {code} (lần {attempt + 1})")
                logger.info(f"Callback thành công cho {code} (lần {attempt + 1})")
                return True
            else:
                print(f"   ❌ Callback thất bại cho {code} (lần {attempt + 1}): HTTP {response.status_code}")
                logger.warning(f"Callback thất bại cho {code} (lần {attempt + 1}): {response.status_code}")
                
        except Exception as e:
            print(f"   ❌ Callback lỗi cho {code} (lần {attempt + 1}): {e}")
            logger.warning(f"Callback lỗi cho {code} (lần {attempt + 1}): {e}")
        
        if attempt < 2:  # Còn cơ hội retry
            print(f"   ⏳ Chờ 2s trước khi retry...")
            time.sleep(2)  # Delay giữa các lần retry
    
    print(f"   💥 Callback thất bại sau 3 lần thử cho {code}")
    logger.error(f"Callback thất bại sau 3 lần thử cho {code}")
    return False

def update_database_immediately(order_id: str, code: str, status: str, amount: Any, notes: str, details: Optional[Dict[str, Any]] = None) -> bool:
    """Update database ngay lập tức cho từng đơn xử lý xong."""
    print(f"💾 [DATABASE] Cập nhật database cho mã {code} - Order: {order_id}")
    print(f"   📊 Status: {status}, Amount: {amount}")
    print(f"   📝 Notes: {notes}")
    
    if details:
        print(f"   📋 Chi tiết FTTH:")
        for key, value in details.items():
            print(f"      • {key}: {value}")
    
    # Cập nhật trực tiếp DB theo yêu cầu
    try:
        with psycopg2.connect(os.getenv('DATABASE_URL', 'postgres://postgres:123456@localhost:5432/autogachno')) as conn:
            with conn.cursor() as cur:
                import json as pyjson
                result_obj = {
                    'code': code,
                    'status': 'completed' if status == 'success' else status,
                    'amount': str(amount) if amount is not None else None,
                    'notes': notes,
                    'details': details or None,
                }
                result_json = pyjson.dumps(result_obj, ensure_ascii=False)
                # Log điều kiện WHERE dễ đọc
                cur.execute("SELECT input_data, service_type, user_id FROM orders WHERE id = %s", (order_id,))
                row = cur.fetchone()
                if row:
                    input_data, service_type, user_id = row
                    print("   🔎 WHERE ví dụ:")
                    print(f"      WHERE input_data = '{(input_data or '').strip()}'")
                    print(f"        AND service_type = '{service_type}'")
                    print(f"        AND user_id = '{user_id}'")
                    print(f"        AND status IN ('pending','processing')")
                # Update theo id
                # Cập nhật orders (không ràng buộc status để đảm bảo luôn ghi nhận kết quả)
                cur.execute(
                    """
                    UPDATE orders
                    SET status = %s,
                        result_data = %s,
                        updated_at = NOW()
                    WHERE id = %s
                    RETURNING id
                    """,
                    ('completed' if status == 'success' else status, result_json, order_id)
                )
                ok_order = cur.fetchone()
                if ok_order:
                    print(f"   ✅ Updated order {order_id}")
                else:
                    print("   ⚠️ Không có bản ghi phù hợp để update (orders) hoặc đã ở trạng thái cuối cùng")

                # Cập nhật service_transactions cho đúng code (ưu tiên theo order_id + code)
                cur.execute(
                    """
                    UPDATE service_transactions
                    SET status = %s,
                        amount = COALESCE(%s, amount),
                        notes = %s,
                        processing_data = %s,
                        updated_at = NOW()
                    WHERE order_id = %s
                      AND code = %s
                      AND status IN ('pending','processing')
                    RETURNING id
                    """,
                    (
                        'success' if status == 'success' else 'failed',
                        str(amount) if isinstance(amount, (int, float)) else None,
                        notes,
                        result_json,
                        order_id,
                        code,
                    )
                )
                ok_tran = cur.fetchone()
                if ok_tran:
                    print(f"   ✅ Updated transaction for code {code}")
                else:
                    print("   ⚠️ Không có transaction phù hợp để update (order_id+code). Thử cập nhật transaction mới nhất theo code...")
                    # Fallback: cập nhật transaction mới nhất theo code và service_type
                    cur.execute(
                        """
                        SELECT st.id
                        FROM service_transactions st
                        JOIN orders o ON o.id = st.order_id
                        WHERE st.code = %s
                          AND o.service_type = %s
                          AND st.status IN ('pending','processing')
                        ORDER BY st.created_at DESC
                        LIMIT 1
                        """,
                        (code, service_type)
                    )
                    st_row = cur.fetchone()
                    if st_row:
                        st_id = st_row[0]
                        cur.execute(
                            """
                            UPDATE service_transactions
                            SET status = %s,
                                amount = COALESCE(%s, amount),
                                notes = %s,
                                processing_data = %s,
                                updated_at = NOW()
                            WHERE id = %s
                            RETURNING id
                            """,
                            (
                                'success' if status == 'success' else 'failed',
                                str(amount) if isinstance(amount, (int, float)) else None,
                                notes,
                                result_json,
                                st_id,
                            )
                        )
                        ok_tran2 = cur.fetchone()
                        if ok_tran2:
                            print(f"   ✅ Updated latest transaction for code {code} (fallback)")
                            ok_tran = ok_tran2
                        else:
                            print("   ⚠️ Không thể cập nhật transaction (fallback)")
                    else:
                        print("   ℹ️ Không còn transaction pending/processing nào cho code này (fallback)")

                # Dọn trùng: đóng các transaction khác cùng code còn pending/processing (nếu có)
                try:
                    cur.execute(
                        """
                        UPDATE service_transactions st
                        SET status = 'failed',
                            notes = CASE WHEN st.notes IS NULL OR st.notes = '' THEN %s ELSE st.notes || ' | ' || %s END,
                            updated_at = NOW()
                        FROM orders o
                        WHERE st.order_id = o.id
                          AND st.code = %s
                          AND st.order_id <> %s
                          AND o.service_type = %s
                          AND st.status IN ('pending','processing')
                        RETURNING st.id
                        """,
                        ("auto-closed dup", "auto-closed dup", code, order_id, service_type)
                    )
                    closed_dups = cur.fetchall()
                    if closed_dups:
                        print(f"   ✅ Đã đóng {len(closed_dups)} transaction trùng (pending/processing) cho {code}")
                except Exception as _e:
                    print(f"   ⚠️ Lỗi khi đóng transaction trùng: {_e}")

                # Kiểm tra lại còn transaction pending/processing cho mã này không
                try:
                    pendings_after = db_check_pending_orders_for_code(service_type, code, None)
                    if pendings_after:
                        print(f"   🔎 Còn pending/processing sau update cho {code}: {', '.join(pendings_after)}")
                    else:
                        print(f"   ✅ Không còn pending/processing cho {code} sau update")
                except Exception:
                    pass

                return bool(ok_order or ok_tran)
    except Exception as e:
        print(f"   ❌ Lỗi cập nhật DB trực tiếp: {e}")
        return False
    
def db_find_order_id(service_type: str, code: str, user_id: Optional[str] = None) -> Optional[str]:
    """Tìm order_id theo code từ bảng service_transactions đang pending/processing (ưu tiên), khớp service_type của orders."""
    try:
        with psycopg2.connect(DB_DATABASE_URL) as conn:
            with conn.cursor() as cur:
                if user_id:
                    cur.execute(
                        """
                        SELECT st.order_id
                        FROM service_transactions st
                        JOIN orders o ON o.id = st.order_id
                        WHERE st.code = %s
                          AND o.service_type = %s
                          AND o.user_id = %s
                          AND st.status IN ('pending','processing')
                        ORDER BY st.created_at DESC
                        LIMIT 1
                        """,
                        (code, service_type, user_id)
                    )
                else:
                    cur.execute(
                        """
                        SELECT st.order_id
                        FROM service_transactions st
                        JOIN orders o ON o.id = st.order_id
                        WHERE st.code = %s
                          AND o.service_type = %s
                          AND st.status IN ('pending','processing')
                        ORDER BY st.created_at DESC
                        LIMIT 1
                        """,
                        (code, service_type)
                    )
                row = cur.fetchone()
                return row[0] if row else None
    except Exception as e:
        print(f"[DB] Lỗi tìm orderId cho code='{code}': {e}")
        return None

def db_check_pending_orders_for_code(service_type: str, code: str, user_id: Optional[str] = None) -> List[str]:
    """Trả về danh sách order_id còn 'pending'/'processing' cho mã (code) và service_type."""
    try:
        with psycopg2.connect(DB_DATABASE_URL) as conn:
            with conn.cursor() as cur:
                if user_id:
                    cur.execute(
                        """
                        SELECT st.order_id
                        FROM service_transactions st
                        JOIN orders o ON o.id = st.order_id
                        WHERE st.code = %s
                          AND o.service_type = %s
                          AND o.user_id = %s
                          AND st.status IN ('pending','processing')
                        ORDER BY st.created_at DESC
                        """,
                        (code, service_type, user_id)
                    )
                else:
                    cur.execute(
                        """
                        SELECT st.order_id
                        FROM service_transactions st
                        JOIN orders o ON o.id = st.order_id
                        WHERE st.code = %s
                          AND o.service_type = %s
                          AND st.status IN ('pending','processing')
                        ORDER BY st.created_at DESC
                        """,
                        (code, service_type)
                    )
                rows = cur.fetchall()
                return [r[0] for r in rows]
    except Exception as e:
        print(f"[DB] Lỗi kiểm tra pending cho code='{code}': {e}")
        return []

def send_callback_with_retry(order_id: str, code: str, status: str, amount: Any, notes: str, details: Optional[Dict[str, Any]] = None) -> bool:
    """Gửi callback với retry logic tối đa 3 lần."""
    print(f"🔄 [CALLBACK] Bắt đầu gửi callback cho mã {code} - Order: {order_id}")
    print(f"   📊 Status: {status}, Amount: {amount}, Notes: {notes}")
    
    for attempt in range(3):
        try:
            payload = {
                "orderId": order_id,
                "code": code,
                "status": status,
                "amount": str(amount) if isinstance(amount, (int, float)) else None,
                "notes": notes,
            }
            
            if details:
                payload["data"] = {
                    "type": "ftth_details",
                    "details": details,
                }
                print(f"   📋 Data: {details}")
            
            print(f"   📤 Lần thử {attempt + 1}/3: Gửi đến {Config.NODE_SERVER_URL}/api/automation/callback")
            response = requests.post(
                f"{Config.NODE_SERVER_URL}/api/automation/callback",
                json=payload,
                timeout=10,
            )
            
            if response.status_code == 200:
                print(f"   ✅ Callback thành công cho {code} (lần {attempt + 1})")
                logger.info(f"Callback thành công cho {code} (lần {attempt + 1})")
                return True
            else:
                print(f"   ❌ Callback thất bại cho {code} (lần {attempt + 1}): HTTP {response.status_code}")
                logger.warning(f"Callback thất bại cho {code} (lần {attempt + 1}): {response.status_code}")
                
        except Exception as e:
            print(f"   ❌ Callback lỗi cho {code} (lần {attempt + 1}): {e}")
            logger.warning(f"Callback lỗi cho {code} (lần {attempt + 1}): {e}")
        
        if attempt < 2:
            print(f"   ⏳ Chờ 2s trước khi retry...")
            time.sleep(2)
    
    print(f"   💥 Callback thất bại sau 3 lần thử cho {code}")
    logger.error(f"Callback thất bại sau 3 lần thử cho {code}")
    return False

def update_database_immediately(order_id: str, code: str, status: str, amount: Any, notes: str, details: Optional[Dict[str, Any]] = None) -> bool:
    """Update database ngay lập tức cho từng đơn xử lý xong."""
    print(f"💾 [DATABASE] Cập nhật database cho mã {code} - Order: {order_id}")
    print(f"   📊 Status: {status}, Amount: {amount}")
    print(f"   📝 Notes: {notes}")
    
    if details:
        print(f"   📋 Chi tiết FTTH:")
        for key, value in details.items():
            print(f"      • {key}: {value}")
    
    try:
        with psycopg2.connect(os.getenv('DATABASE_URL', 'postgres://postgres:123456@localhost:5432/autogachno')) as conn:
            with conn.cursor() as cur:
                import json as pyjson
                result_obj = {
                    'code': code,
                    'status': 'completed' if status == 'success' else status,
                    'amount': str(amount) if amount is not None else None,
                    'notes': notes,
                    'details': details or None,
                }
                result_json = pyjson.dumps(result_obj, ensure_ascii=False)
                cur.execute("SELECT input_data, service_type, user_id FROM orders WHERE id = %s", (order_id,))
                row = cur.fetchone()
                if row:
                    input_data, service_type, user_id = row
                    print("   🔎 WHERE ví dụ:")
                    print(f"      WHERE input_data = '{(input_data or '').strip()}'")
                    print(f"        AND service_type = '{service_type}'")
                    print(f"        AND user_id = '{user_id}'")
                    print(f"        AND status IN ('pending','processing')")
                cur.execute(
                    """
                    UPDATE orders
                    SET status = %s,
                        result_data = %s,
                        updated_at = NOW()
                    WHERE id = %s AND status IN ('pending','processing')
                    RETURNING id
                    """,
                    ('completed' if status == 'success' else status, result_json, order_id)
                )
                ok = cur.fetchone()
                if ok:
                    print(f"   ✅ Updated order {order_id}")
                    return True
                print("   ⚠️ Không có bản ghi phù hợp để update")
                return False
    except Exception as e:
        print(f"   ❌ Lỗi cập nhật DB trực tiếp: {e}")
        return False

def db_insert_orders_from_lines(service_type: str, user_id: str, lines: List[str]) -> int:
    codes = [ln.strip() for ln in lines if ln and ln.strip()]
    if not codes:
        return 0
    try:
        with psycopg2.connect(DB_DATABASE_URL) as conn:
            with conn.cursor() as cur:
                count = 0
                for code in codes:
                    cur.execute(
                        """
                        INSERT INTO orders (user_id, service_type, status, input_data)
                        VALUES (%s, %s, 'pending', %s)
                        RETURNING id
                        """,
                        (user_id, service_type, code)
                    )
                    cur.fetchone()
                    count += 1
                print(f"[DB] Đã tạo {count} orders cho service '{service_type}'.")
                return count
    except Exception as e:
        print(f"[DB] Lỗi insert orders: {e}")
        return 0

def get_exe_dir():
    exe_path = sys.argv[0]
    exe_dir = os.path.dirname(exe_path)
    return exe_dir

def start_api_server():
    """Bỏ qua mock API server: danh sách đơn lấy từ Database qua Node API."""
    logger.info("Bỏ qua mock_api_server: dùng trực tiếp Node API/DB")
    return False

def ensure_driver_and_login() -> bool:
    """Đảm bảo Chrome driver đã sẵn sàng (bỏ qua đăng nhập)."""
    global driver
    try:
        if driver is None:
            username = LOGIN_USERNAME or "default"
            driver = initialize_browser(username)
            if not driver:
                logger.error("Không thể khởi tạo trình duyệt")
                return False
        return True
    except Exception as e:
        logger.error(f"ensure_driver_and_login error: {e}")
        return False

def process_lookup_ftth_codes(codes: List[str], order_id: Optional[str] = None):
    """Xử lý tra cứu FTTH không cần GUI, điều khiển selenium trực tiếp."""
    print(f"🚀 [AUTOMATION] Bắt đầu xử lý FTTH cho {len(codes)} mã")
    print(f"   📋 Order ID: {order_id or 'Không có'}")
    
    if not ensure_driver_and_login():
        print("   ❌ Không thể khởi tạo driver hoặc đăng nhập")
        logger.error("Không thể khởi tạo driver hoặc đăng nhập")
        return
    
    print("   ✅ Bắt đầu xử lý (bỏ qua đăng nhập)")
    
    try:
        with automation_lock:
            print("   🔒 Đã khóa automation, điều hướng đến trang FTTH...")
            navigate_to_ftth_page_and_select_radio()
            print("   ✅ Đã điều hướng thành công đến trang FTTH")
            
            # Hiển thị tiến trình cho từng order nếu có
            if order_id:
                print("   📋 Danh sách mã sẽ xử lý:")
                for idx, cb in enumerate(codes, 1):
                    print(f"      {idx}. {cb}")

            results = []
            for idx, cbil in enumerate(codes, 1):
                print(f"\n📱 [MÃ {idx}/{len(codes)}] Xử lý mã: {cbil}")
                success = False
                for attempt in range(AUTOMATION_MAX_RETRIES):  # Retry tối đa cấu hình
                    try:
                        cbil = (cbil or "").strip()
                        if not cbil:
                            print(f"   ⚠️  Mã rỗng, bỏ qua")
                            break
                        
                        if attempt > 0:
                            print(f"   🔄 Retry lần {attempt + 1}/{AUTOMATION_MAX_RETRIES} cho mã {cbil}")
                            logger.info(f"Retry lần {attempt + 1} cho mã {cbil}")
                            driver.refresh()
                            time.sleep(2)
                            navigate_to_ftth_page_and_select_radio()
                        else:
                            print(f"   🎯 Lần thử đầu tiên cho mã {cbil}")
                        
                        root.update() if 'root' in globals() else None
                        time.sleep(0.5)
                        
                        print(f"   📝 Điền mã thuê bao: {cbil}")
                        customer = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:contractCode")))
                        customer.clear()
                        customer.send_keys(cbil)
                        
                        print(f"   🔍 Nhấn nút KIỂM TRA...")
                        payment_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:btnPay0")))
                        payment_button.click()
                        time.sleep(1)
                        
                        print(f"   ⏳ Chờ modal loading...")
                        WebDriverWait(driver, 16).until(EC.invisibility_of_element_located((By.ID, "payMoneyForm:j_idt6_modal")))
                        
                        print(f"   📊 Lấy thông tin kết quả...")
                        element41 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:j_idt41")))
                        _is_amount, amount, _pid = amount_by_cbil(cbil, element41, False)
                        details = extract_ftth_details_from_page()
                        
                        print(f"   ✅ Xử lý thành công: Amount = {amount}")
                        if details:
                            print(f"   📋 Chi tiết FTTH:")
                            for key, value in details.items():
                                print(f"      • {key}: {value}")
                        
                        # Thành công, thoát khỏi retry loop
                        success = True
                        results.append({"code": cbil, "amount": amount, "status": "success"})
                        
                        # Update database ngay lập tức cho từng đơn xử lý xong
                        if order_id:
                            print(f"   💾 Bắt đầu update database...")
                            db_success = update_database_immediately(order_id, cbil, "success", amount, "FTTH lookup ok", details)
                            if not db_success:
                                logger.warning(f"Database update thất bại cho {cbil}")
                        else:
                            print(f"⚠️  [WARNING] Không có order_id, bỏ qua database update cho {cbil}")
                        
                        print(f"   🎉 Hoàn thành xử lý mã {cbil}")
                        break
                        
                    except Exception as e:
                        print(f"   ❌ Lần thử {attempt + 1} thất bại: {e}")
                        logger.warning(f"Lần thử {attempt + 1} thất bại cho {cbil}: {e}")
                        if attempt < (AUTOMATION_MAX_RETRIES - 1):  # Còn cơ hội retry
                            print(f"   ⏳ Chờ 1s trước khi retry...")
                            time.sleep(1)  # Delay thông minh
                            continue
                        else:  # Hết retry
                            print(f"   💥 Hết retry, mã {cbil} thất bại hoàn toàn")
                            logger.error(f"FTTH code {cbil} thất bại sau {AUTOMATION_MAX_RETRIES} lần thử: {e}")
                            results.append({"code": cbil, "amount": None, "status": "failed", "message": str(e)})
                            
                            # Update database cho trường hợp thất bại
                            if order_id:
                                print(f"   💾 Update database cho trường hợp thất bại...")
                                db_success = update_database_immediately(order_id, cbil, "failed", None, str(e), None)
                                if not db_success:
                                    logger.warning(f"Database update thất bại cho {cbil}")
                            else:
                                print(f"⚠️  [WARNING] Không có order_id, bỏ qua database update cho {cbil}")
                
                if not success:
                    print(f"   💥 Mã {cbil} không thể xử lý sau 3 lần thử")
                    logger.error(f"Mã {cbil} không thể xử lý sau 3 lần thử")
            
            print(f"\n📊 [KẾT QUẢ] Tổng kết xử lý FTTH:")
            print(f"   ✅ Thành công: {len([r for r in results if r['status'] == 'success'])} mã")
            print(f"   ❌ Thất bại: {len([r for r in results if r['status'] == 'failed'])} mã")
            print(f"   📋 Tổng cộng: {len(results)} mã")
            
            logger.info(f"FTTH processed: {len(results)} items")
    except Exception as e:
        print(f"   💥 Lỗi tổng thể: {e}")
        logger.error(f"process_lookup_ftth_codes error: {e}")


def check_api_health():
    """Kiểm tra Node API (nguồn dữ liệu thật) có hoạt động không"""
    try:
        response = requests.get(f"{Config.NODE_SERVER_URL}/api/health", timeout=5)
        return response.status_code == 200
    except Exception:
        return False

def db_fetch_service_data(service_type: str) -> Optional[Dict[str, Any]]:
    """Đọc thẳng từ Postgres: trả về dữ liệu mẫu theo service_type.
    - Lấy tối đa 20 giao dịch mới nhất từ service_transactions join orders theo service_type
    - Trả về code list theo định dạng của từng dịch vụ và kèm order_id gần nhất + mapping code->orderId
    """
    try:
        with psycopg2.connect(os.getenv('DATABASE_URL', 'postgres://postgres:123456@localhost:5432/autogachno')) as conn:
            with conn.cursor() as cur:
                # Lấy mỗi mã (code) 1 dòng: chọn transaction mới nhất cho từng code
                cur.execute(
                    """
                    SELECT DISTINCT ON (st.code) st.code, st.order_id, st.created_at
                    FROM service_transactions st
                    JOIN orders o ON o.id = st.order_id
                    WHERE o.service_type = %s
                      AND st.status IN ('pending','processing')
                    ORDER BY st.code, st.created_at DESC
                    """,
                    (service_type,)
                )
                rows = cur.fetchall()
                codes = []
                code_order_map = []
                for code_val, order_val, _ in rows:
                    if not code_val:
                        continue
                    codes.append(code_val)
                    if order_val:
                        code_order_map.append({'code': code_val, 'orderId': order_val})
                # Lấy order_id mới nhất theo created_at
                latest_order_id = None
                if rows:
                    latest_row = max(rows, key=lambda r: r[2] or datetime.min)
                    latest_order_id = latest_row[1]

                result: Dict[str, Any] = {}
                if service_type in ("tra_cuu_ftth", "thanh_toan_tv_internet"):
                    result["subscriber_codes"] = codes[:10]
                elif service_type == "gach_dien_evn":
                    result["bill_codes"] = codes[:10]
                elif service_type in ("nap_tien_da_mang", "nap_tien_viettel", "tra_cuu_no_tra_sau"):
                    result["phone_numbers"] = codes[:10]
                else:
                    # fallback chung
                    result["codes"] = codes[:10]

                result["order_id"] = latest_order_id
                result["code_order_map"] = code_order_map
                return result
    except Exception as e:
        logger.error(f"Lỗi đọc DB cho {service_type}: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi kết nối DB: {e}")
        return None

def fetch_api_data(service_type: str) -> Optional[Dict]:
    """Đã chuyển sang đọc trực tiếp DB (bỏ API)."""
    return db_fetch_service_data(service_type)

def populate_text_widget(text_widget, data_list):
    """Đổ dữ liệu vào Text widget"""
    try:
        text_widget.config(state="normal")
        text_widget.delete("1.0", "end")
        if data_list:
            text_widget.insert("1.0", "\n".join(data_list))
        text_widget.config(state="normal")
    except Exception as e:
        logger.error(f"Lỗi đổ dữ liệu vào text widget: {e}")

def populate_entry_widget(entry_widget, value):
    """Đổ dữ liệu vào Entry widget"""
    try:
        entry_widget.delete(0, "end")
        if value:
            entry_widget.insert(0, str(value))
    except Exception as e:
        logger.error(f"Lỗi đổ dữ liệu vào entry widget: {e}")

def populate_combobox_widget(combobox_widget, value):
    """Đổ dữ liệu vào Combobox widget"""
    try:
        if value and value in combobox_widget['values']:
            combobox_widget.set(value)
    except Exception as e:
        logger.error(f"Lỗi đổ dữ liệu vào combobox widget: {e}")

def get_data_ftth(text_widget, order_entry: Optional[ttk.Entry] = None):
    """Lấy dữ liệu API cho Tra cứu FTTH và set Order ID nếu có"""
    try:
        data = fetch_api_data("tra_cuu_ftth")
        if data and "subscriber_codes" in data:
            codes = data["subscriber_codes"]
            populate_text_widget(text_widget, codes)
            order_id = data.get("order_id")
            code_map = data.get("code_order_map") or []
            info_msg = f"Đã tải {len(codes)} mã thuê bao FTTH"
            if order_id:
                print(f"[INFO] Order ID từ DB (gần nhất): {order_id}", flush=True)
                logger.info(f"Order ID từ DB (FTTH, gần nhất): {order_id}")
                info_msg += f"\nOrder ID (gần nhất): {order_id}"
            # In mapping chi tiết mã -> orderId nếu có
            if code_map:
                print("[INFO] Mapping mã -> Order ID:", flush=True)
                for item in code_map:
                    try:
                        print(f"  {item.get('code')}: {item.get('orderId')}", flush=True)
                    except Exception:
                        pass
            #messagebox.showinfo(Config.TITLE, info_msg)
        else:
            logger.warning("Không có dữ liệu từ DB")
            #messagebox.showwarning(Config.TITLE, "Không có dữ liệu từ DB")
    except Exception as e:
        logger.error(f"Lỗi lấy dữ liệu FTTH: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi lấy dữ liệu FTTH: {e}")

def export_excel(data: List[Tuple[str, Any, str]], name_dir: str) -> bool:
    """Xuất dữ liệu ra file Excel"""
    today = datetime.now().strftime("%H%M-%d-%m-%Y")
    try:
        export_dir = os.path.join(os.getcwd(), f"{Config.FOLDER_RESULT}\\{name_dir}")
        os.makedirs(export_dir, exist_ok=True)
        
        file_name = f"{today}.xlsx"
        file_path = os.path.join(export_dir, file_name)

        wb = Workbook()
        ws = wb.active
        
        # Định dạng header
        headers = ['STT', 'Số thuê bao', 'Số tiền', 'Ghi chú']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # Ghi dữ liệu
        for idx, (phone, amount, note) in enumerate(data, start=2):
            ws.cell(row=idx, column=1, value=idx - 1)
            ws.cell(row=idx, column=2, value=phone)
            ws.cell(row=idx, column=3, value=amount)
            ws.cell(row=idx, column=4, value=note)
            
        wb.save(file_path)
        
        # Hỏi người dùng có muốn mở file không
        # if messagebox.askyesno(Config.TITLE, f"Dữ liệu được lưu tại: \n{Config.FOLDER_RESULT}/{name_dir}/{today}"):
        #     try:
        #         os.startfile(file_path)
        #     except Exception as e:
        #         logger.warning(f"Không thể mở file Excel: {e}")
                
        return True
        
    except Exception as e:
        logger.error(f"Lỗi xuất Excel: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi xuất Excel: {e}")
        return False

def valid_data(data: List[Any]) -> bool:
    """Kiểm tra dữ liệu đầu vào"""
    try:
        if not check_username():
            return False
        for item in data:
            # Nếu là danh sách (ví dụ danh sách mã), yêu cầu có ít nhất 1 phần tử không rỗng
            if isinstance(item, (list, tuple)):
                has_nonempty = any((isinstance(x, str) and x.strip()) for x in item)
                if not has_nonempty:
                    #messagebox.showwarning(Config.TITLE, "Vui lòng nhập đầy đủ thông tin")
                    return False
            else:
                # Xử lý chuỗi/tham số đơn lẻ
                text = str(item) if item is not None else ""
                if not text.strip():
                    #messagebox.showwarning(Config.TITLE, "Vui lòng nhập đầy đủ thông tin")
                    return False
        return True
    except Exception as e:
        logger.error(f"Lỗi kiểm tra dữ liệu: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi kiểm tra dữ liệu: {e}")
        return False

def delete_ctmed(cmted: tk.Text):
    """Xóa nội dung text widget"""
    cmted.config(state="normal")
    cmted.delete("1.0", "end")
    cmted.config(state="disabled")

def insert_ctmed(cmted: tk.Text, cbil: str):
    """Thêm text vào widget"""
    cmted.config(state="normal")
    cmted.insert("1.0", f"{cbil}\n")
    cmted.config(state="disabled")

def stop_tool():
    """Dừng chương trình"""
    global stop_flag
    stop_flag = True
    #messagebox.showinfo(Config.TITLE, "Đã dừng chương trình")

def update_stop_flag():
    """Reset stop flag"""
    global stop_flag
    stop_flag = False

def toggle_input_amount(select, label, combobox):
    selected_value = select.get()
    if selected_value == "Gạch nợ trả sau":
        combobox.pack_forget()
        label.pack_forget()
    else:
        combobox.pack(side="right")
        label.pack(side="right")
    root.update()

def handle_choose_select(choose: str) -> int:
    """Xử lý lựa chọn loại thanh toán"""
    try:
        choose = choose.strip()
        if choose == "Nạp trả trước":
            return 1
        else:
            return 2
    except Exception as e:
        logger.error(f"Lỗi xử lý loại thanh toán: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi xử lý loại thanh toán: {e}")
        return 1

def get_chrome_driver(username: str = "default") -> Optional[webdriver.Chrome]:
    """Tạo Chrome driver"""
    try:
        profile_dir = os.path.join(os.getcwd(), "chrome_profile", username)
        os.makedirs(profile_dir, exist_ok=True)
        
        options = Options()
        options.add_argument(f"--user-data-dir={profile_dir}")
        options.add_argument("--start-maximized")
        options.add_argument("--disable-notifications")
        options.add_argument("--disable-popup-blocking")
        options.add_experimental_option("excludeSwitches", ["enable-logging"])
        
        # Thêm các tùy chọn bảo mật
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        
        driver = webdriver.Chrome(options=options)
        return driver
    except Exception as e:
        logger.error(f"Lỗi khởi tạo Chrome Driver: {e}")
        return None

def check_username() -> bool:
    """Kiểm tra username đăng nhập dựa trên thông tin hiển thị trên trang."""
    try:
        # Bỏ qua kiểm tra đăng nhập (yêu cầu)
        if driver is None:
            return True

        expected_username = (dbfiles or {}).get("username", "").strip()
        if not expected_username:
            return True

        attempts = 3
        for _ in range(attempts):
            try:
                dl_info = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "dl-info-detail"))
                )
                # Tìm tất cả các div info trong dl-info-detail
                info_divs = dl_info.find_elements(By.CLASS_NAME, "info")
                if len(info_divs) >= 3:
                    # Lấy mã đại lý từ div đầu tiên
                    agency_span = info_divs[0].find_element(By.TAG_NAME, "span")
                    agency_code = (agency_span.text or "").strip()
                    
                    # Lấy số điện thoại từ div thứ 3
                    phone_span = info_divs[2].find_element(By.TAG_NAME, "span")
                    phone_number = (phone_span.text or "").strip()
                    
                    # Kiểm tra username có khớp với mã đại lý hoặc số điện thoại không
                    if (expected_username == agency_code or 
                        expected_username == phone_number or
                        expected_username in agency_code or
                        expected_username in phone_number):
                        return True
                    else:
                        #messagebox.showerror(Config.TITLE, f"Vui lòng sử dụng đúng tài khoản đã đăng ký. Mong đợi: {expected_username}, Tìm thấy: {agency_code}")
                        return False
                return True
            except StaleElementReferenceException:
                time.sleep(0.5)
                continue
            except Exception as e:
                logger.warning(f"Lỗi parse thông tin đăng nhập: {e}")
                time.sleep(0.5)
                continue
        # Nếu vẫn không lấy được sau retry
        #messagebox.showerror(Config.TITLE, "Không tìm thấy thông tin tài khoản trên Viettel Pay Pro")
        return False
    except Exception as e:
        logger.error(f"Lỗi kiểm tra username: {e}")
        #messagebox.showerror(Config.TITLE, "Không tìm thấy thông tin tài khoản trên Viettel Pay Pro")
        return False

def get_number_uses() -> Tuple[int, Dict[str, int]]:
    """Lấy số lần sử dụng các dịch vụ"""
    try:
        services = {
            Config.SERVICES['payment_internet']: 99999,
            Config.SERVICES['payment_card']: 99999,
            Config.SERVICES['lookup_card']: 99999,
            Config.SERVICES['lookup_ftth']: 99999,
            Config.SERVICES['payment_evn']: 99999
        }
        return 0, services
    except Exception as e:
        logger.error(f"Lỗi lấy số lần sử dụng: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi lấy số lần sử dụng: {e}")
        return 0, {}

def handle_choose_amount(am: str) -> str:
    """Xử lý lựa chọn số tiền"""
    try:
        amount_map = {
            "10.000đ": "0", "20.000đ": "1", "30.000đ": "2", "50.000đ": "3",
            "100.000đ": "4", "200.000đ": "5", "300.000đ": "6", "500.000đ": "7"
        }
        return amount_map.get(am, "0")
    except Exception as e:
        logger.error(f"Lỗi xử lý số tiền: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi xử lý số tiền: {e}")
        return "0"

def amount_by_cbil(cbil: str, element, lookup: bool = False) -> Tuple[bool, Any, Optional[str]]:
    """Lấy số tiền theo mã thuê bao"""
    try:
        amount = "Không tìm thấy mã thuê bao"
        payment_id = None
        
        html_content = element.get_attribute('outerHTML')
        soup = BeautifulSoup(html_content, 'html.parser')
        pay_content_groups = soup.find_all("div", class_="row pay-content mb-3")
        
        for group in pay_content_groups:
            p_tags = group.find_all("p")
            is_found = any(cbil in p_tag.text for p_tag in p_tags)
            
            if lookup and is_found:
                button_tag = group.find("button", {"id": re.compile(r'payMoneyForm:btnView\d*')})
                if button_tag:
                    payment_id = button_tag['id']
            
            if is_found:
                for p_tag in p_tags:
                    if "VND" in p_tag.text:
                        str_price = p_tag.text.split("VND")[0].strip()
                        amount = int(str_price.replace(",", ""))
                        if amount >= 5000:
                            return True, amount, payment_id
                        else:
                            return False, amount, payment_id
                            
        return False, amount, payment_id
        
    except Exception as e:
        logger.error(f"Lỗi lấy số tiền: {e}")
        return False, "Lỗi thanh toán", None

def navigate_to_ftth_page_and_select_radio():
    """Đi tới trang FTTH và chọn radio 'Số thuê bao'"""
    try:
        # Bỏ qua kiểm tra đăng nhập theo yêu cầu
        target_url = "https://kpp.bankplus.vn/pages/newInternetTelevisionViettel.jsf?serviceCode=000003&serviceType=INTERNET"
        driver.get(target_url)
        # Chờ input mã thuê bao xuất hiện
        time.sleep(5)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:contractCode")))
        # Chọn radio Số thuê bao (id payMoneyForm:console:3)
        try:
            radio_input = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, "payMoneyForm:console:3")))
            radio_box = radio_input.find_element(By.XPATH, "../../div[contains(@class,'ui-radiobutton-box')]")
            radio_box.click()
        except Exception:
            # fallback click vào label
            try:
                label = driver.find_element(By.XPATH, "//label[@for='payMoneyForm:console:3']")
                label.click()
            except Exception:
                pass
    except Exception as e:
        logger.warning(f"Không thể điều hướng FTTH hoặc chọn radio: {e}")
        raise  # Re-raise để caller biết có lỗi

def extract_ftth_details_from_page() -> Dict[str, Any]:
    """Trích thông tin chi tiết FTTH từ trang hiện tại sau khi nhấn KIỂM TRA."""
    details: Dict[str, Any] = {}
    try:
        html_content = driver.page_source
        soup = BeautifulSoup(html_content, 'html.parser')
        label_to_key = {
            'Mã hợp đồng:': 'contract_code',
            'Chủ hợp đồng:': 'contract_owner',
            'Số thuê bao đại diện:': 'representative_subscriber',
            'Dịch vụ:': 'service',
            'Số điện thoại liên hệ:': 'contact_phone',
            'Nợ cước:': 'debt_amount',
        }
        # Duyệt các hàng hiển thị label-value
        for row in soup.find_all('div', class_='row'):
            cols = row.find_all('div', class_='col-6')
            if len(cols) != 2:
                continue
            label_tag = cols[0].find('label')
            value_p = cols[1].find('p')
            if not label_tag or not value_p:
                continue
            label_text = label_tag.get_text(strip=True)
            value_text = value_p.get_text(strip=True)
            key = label_to_key.get(label_text)
            if not key:
                continue
            if key == 'debt_amount':
                try:
                    # Lấy số từ chuỗi "170,000 VND"
                    num_str = re.findall(r"[\d\.,]+", value_text)
                    if num_str:
                        details[key] = int(num_str[0].replace('.', '').replace(',', ''))
                    else:
                        details[key] = None
                except Exception:
                    details[key] = None
            else:
                details[key] = value_text
    except Exception as e:
        logger.warning(f"Lỗi trích chi tiết FTTH: {e}")
    return details

def get_error_alert_text() -> Optional[str]:
    """Trả về nội dung thông báo lỗi (role="alert") nếu có."""
    try:
        elements = driver.find_elements(By.CSS_SELECTOR, "li[role='alert'] span.ui-messages-error-summary")
        for el in elements:
            text_val = (el.text or "").strip()
            if text_val:
                return text_val
    except Exception:
        pass
    return None

def lookup_ftth(tkinp_ctm, tkinp_ctmed, tkinp_order: Optional[ttk.Entry] = None):
    
    try:
        delete_ctmed(tkinp_ctmed)
        update_stop_flag()
        navigate_to_ftth_page_and_select_radio()
        cbils = tkinp_ctm.get("1.0", "end-1c").splitlines()
        if not valid_data([cbils]):
            return False
        # Hiển thị mapping mã -> orderId trước khi xử lý
        print("Order ID:")
        code_to_order: Dict[str, Optional[str]] = {}
        for raw in cbils:
            c = (raw or "").strip()
            if not c:
                continue
            oid = db_find_order_id('tra_cuu_ftth', c, None)
            code_to_order[c] = oid
            print(f"  {c}: {oid if oid else 'Không tìm thấy'}")

        data = []
        for cbil in cbils:
            root.update()
            time.sleep(1)
            cbil = cbil.strip()
            if not stop_flag and cbil.strip() != "":
                success = False
                for attempt in range(3):  # Retry tối đa 3 lần
                    try:
                        # Log tiến trình theo orderId
                        print(f"   🔧 Đang xử lý {cbil} | Order ID: {code_to_order.get(cbil) if code_to_order.get(cbil) else 'Không tìm thấy'}")
                        if attempt > 0:
                            logger.info(f"Retry lần {attempt + 1} cho mã {cbil}")
                            driver.refresh()
                            time.sleep(2)
                            navigate_to_ftth_page_and_select_radio()
                        
                        # Điền vào form
                        customer = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:contractCode")))
                        customer.clear()
                        customer.send_keys(cbil)
                        # Nhấn nút thanh toán
                        payment_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:btnPay0")))
                        payment_button.click()
                        time.sleep(1)
                        WebDriverWait(driver, 16).until(EC.invisibility_of_element_located((By.ID, "payMoneyForm:j_idt6_modal")))
                        # Kiểm tra thông báo lỗi (nếu có) và dừng retry khi chứa từ "không"
                        alert_text = get_error_alert_text()
                        if alert_text or ("không" in alert_text.lower()):
                            note_err = alert_text.strip()
                            data.append([cbil, 0, Config.STATUS_INCOMPLETE])
                            insert_ctmed(tkinp_ctmed, f"{cbil} - Lỗi | {note_err}")
                            if order_id_val:
                                try:
                                    _ = update_database_immediately(order_id_val, cbil, "failed", None, note_err, None)
                                except Exception as _e:
                                    logger.warning(f"DB update lỗi (failed) cho {cbil}: {_e}")
                            break
                        element41 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:j_idt41")))
                        is_amount, amount, payment_id = amount_by_cbil(cbil, element41, False)
                        details = extract_ftth_details_from_page()
                        note_text = f"HD:{details.get('contract_code','')} | Chu:{details.get('contract_owner','')} | SDT:{details.get('contact_phone','')} | No:{details.get('debt_amount','')}"
                        
                        # Thành công, thoát khỏi retry loop
                        success = True
                        data.append([cbil, amount, note_text])
                        insert_ctmed(tkinp_ctmed, f"{cbil} - {amount} | {note_text}")
                        
                        # Update trực tiếp DB: dùng mapping đã tra ở trên
                        order_id_val = code_to_order.get(cbil) or db_find_order_id('tra_cuu_ftth', cbil.strip(), None)
                        if not order_id_val:
                            print(f"   ⚠️ Không tìm thấy orderId cho code='{cbil}'. Bỏ qua update.")
                            # Hiển thị danh sách order_id pending/processing (nếu có) để dễ kiểm tra
                            pendings = db_check_pending_orders_for_code('tra_cuu_ftth', cbil.strip(), None)
                            if pendings:
                                print(f"   🔎 Pending/processing orderIds for {cbil}: {', '.join(pendings)}")
                        else:
                            try:
                                db_success = update_database_immediately(order_id_val, cbil, "success", amount, note_text, details)
                                if not db_success:
                                    logger.warning(f"Database update thất bại cho {cbil}")
                            except Exception as _e:
                                logger.warning(f"DB update lỗi (success) cho {cbil}: {_e}")
                        
                        tkinp_ctm.delete("1.0", "1.end+1c")
                        break
                        
                    except Exception as e:
                        logger.warning(f"Lần thử {attempt + 1} thất bại cho {cbil}: {e}")
                        if attempt < (AUTOMATION_MAX_RETRIES - 1):  # Còn cơ hội retry
                            time.sleep(1)  # Delay thông minh
                            continue
                        else:  # Hết retry
                            logger.error(f"Mã {cbil} thất bại sau {AUTOMATION_MAX_RETRIES} lần thử: {e}")
                            data.append([cbil, 0, Config.STATUS_INCOMPLETE])
                            insert_ctmed(tkinp_ctmed, f"{cbil} - Lỗi")
                            tkinp_ctm.delete("1.0", "1.end+1c")
                            break
                
                if not success:
                    logger.error(f"Mã {cbil} không thể xử lý sau 3 lần thử")
        time.sleep(2)
        if len(data) > 0:
            name_dir = "Tra cứu FTTH"
            export_excel(data, name_dir)
    except Exception as e:
        logger.error(f"Lỗi tra cứu FTTH: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi tra cứu FTTH: {e}")

def form_lookup_ftth():
    cus_frm = tk.Frame(root)
    cus_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    ctm_frm = tk.Frame(cus_frm)
    ctm_frm.pack(expand=True, side="left")
    ctmed_frm = tk.Frame(cus_frm)
    ctmed_frm.pack(expand=True, side="right")
    btn_frm = tk.Frame(root)
    btn_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    tklbl_ctm = tk.Label(ctm_frm, text="Số thuê bao (mỗi dòng tạo 1 order)")
    tklbl_ctm.pack(side="top")
    tkinp_ctm = tk.Text(ctm_frm, height=16, width=24)
    tkinp_ctm.pack(side="left", pady=8)
    tklbl_ctm = tk.Label(ctmed_frm, text="Đã xử lý")
    tklbl_ctm.pack(side="top")
    tkinp_ctmed = tk.Text(ctmed_frm, height=16, width=32, bg="#ccc", state="disabled")
    tkinp_ctmed.pack(side="left", pady=8)
    # Loại bỏ ô Order ID: orderId sẽ mặc định dùng chính cbil; nút tạo orders từ DB
    style = ttk.Style()
    style.configure("Red.TButton", foreground="red")
    style.configure("Blue.TButton", foreground="blue")
    style.configure("Green.TButton", foreground="green")
    tkbtn_get_data = ttk.Button(btn_frm, text="Get dữ liệu", command=lambda: get_data_ftth(tkinp_ctm, None))
    tkbtn_get_data.pack(side='left', padx=5, pady=5)
    tkbtn_get_data.configure(style="Green.TButton")
    # Bỏ nút tạo orders: chỉ get dữ liệu và xử lý cập nhật cho DB sẵn có
    tkbtn_payment = ttk.Button(btn_frm, text="Bắt đầu", command=lambda: lookup_ftth(tkinp_ctm, tkinp_ctmed, None))
    tkbtn_payment.pack(side='left', padx=5, pady=5)
    tkbtn_payment.configure(style="Blue.TButton") 
    tkbtn_destroy = ttk.Button(btn_frm, text="Dừng lại", command=stop_tool)
    tkbtn_destroy.pack(side='right', padx=5, pady=5)
    tkbtn_destroy.configure(style="Red.TButton") 

root = tk.Tk()
root.title("HPK Tool - Viettel Pay Automation")
root.geometry("500x550") 
root.option_add("*Font", "Arial 10")
try:
    root.iconbitmap(Config.ICON_FILE)
except Exception as e:
    logger.warning(f"Không thể tải icon: {e}")
    pass

def clear_widgets(main_frm):
    for widget in root.winfo_children():
        if widget is not main_frm:
            widget.destroy()

def read_config():
    """Không dùng đọc cấu hình nữa."""
    return ""



def set_file_config(files):
    try:
        with open(Config.CONFIG_FILE, "r", encoding='utf-8') as file:
            data = json.load(file)
            data["files"] = files
        with open(Config.CONFIG_FILE, "w", encoding='utf-8') as file:
            json.dump(data, file, indent=4, ensure_ascii=False)
    except Exception as e:
        logger.error(f"Lỗi lưu cấu hình: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi lưu cấu hình: {e}")

def handle_key_active(key):
    try:
        key = (key.get()).strip()
        # TODO: Implement key validation through web API instead of local database
        #messagebox.showwarning(Config.TITLE, "Tính năng kích hoạt key đang được phát triển")
    except Exception as e:
        logger.error(f"Lỗi kích hoạt key: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi kích hoạt: {e}")

def check_exits_key(key):
    return True

def show_activate_form():
    # Không dùng kích hoạt key nữa
    pass

def initialize_browser(username="default"):
    global driver
    driver = get_chrome_driver(username)
    if driver:
        driver.get(Config.DRIVER_LINK)
        return driver
    return None

def cleanup():
    global driver
    if driver:
        try:
            driver.quit()
        except Exception as e:
            logger.warning(f"Lỗi khi đóng driver: {e}")
        finally:
            driver = None

def is_logged_in(driver):
    # Bỏ qua kiểm tra trạng thái đăng nhập: luôn cho phép chạy
    return True

def login_process():
    """Đăng nhập tự động khi mở trình duyệt bằng tài khoản cấu hình."""
    try:
        inp_usr = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, "loginForm:userName")))
        inp_usr.clear()
        inp_usr.send_keys(LOGIN_USERNAME)
        inp_pwd = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, "loginForm:password")))
        inp_pwd.clear()
        inp_pwd.send_keys(LOGIN_PASSWORD)
        print("[LOGIN] Đã điền thông tin đăng nhập")
    except Exception as e:
        logger.warning(f"Lỗi đăng nhập: {e}")
        pass

def show_services_form():
    try:
        main_frm = tk.Frame(root)
        main_frm.pack(expand=True, side="top", padx=6, pady=6, fill="both")
        tklbl_choose = tk.Label(main_frm, text="Loại thanh toán:")
        tklbl_choose.pack(side="left")
        tkcbb_choose = ttk.Combobox(main_frm, values=[
            "Tra cứu FTTH",
            "Gạch điện EVN", 
            "Nạp tiền đa mạng",
            "Nạp tiền mạng Viettel",
            "Thanh toán TV - Internet",
            "Tra cứu nợ thuê bao trả sau"
        ], width="32", state="readonly")
        tkcbb_choose.pack(side="left", padx=6, expand=True, fill="x")
        tkcbb_choose.set("Tra cứu FTTH")
        def handle_choose_services(event, choose, main_frm):
            service = choose.get()
            clear_widgets(main_frm)
            if service == "Tra cứu FTTH":
                form_lookup_ftth()
            elif service == "Gạch điện EVN":
                form_debt_electric() 
            elif service == "Nạp tiền đa mạng":
                form_payment_phone()
            elif service == "Nạp tiền mạng Viettel":
                form_payment_viettel()
            elif service == "Thanh toán TV - Internet":
                form_payment_internet()
            elif service == "Tra cứu nợ thuê bao trả sau":
                form_lookup_card()
        tkcbb_choose.bind("<<ComboboxSelected>>", lambda event: handle_choose_services(event, tkcbb_choose, main_frm))
        handle_choose_services(None, tkcbb_choose, main_frm)
    except Exception as e:
        logger.error(f"Lỗi hiển thị form dịch vụ: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi hiển thị form dịch vụ: {e}")

def main():
    global driver
    # Khởi tạo trình duyệt và đăng nhập tự động bằng tài khoản mặc định
    try:
        driver = initialize_browser(LOGIN_USERNAME or "default")
        try:
            login_process()
        except Exception:
            pass
    except Exception as e:
        logger.error(f"Lỗi khởi tạo trình duyệt: {e}")
    # Hiển thị form dịch vụ luôn, không cần đọc cấu hình
    show_services_form()

if __name__ == "__main__":
    try:
        main()
        root.protocol("WM_DELETE_WINDOW", lambda: [cleanup(), root.destroy()])
        root.mainloop()
    except Exception as e:
        logger.error(f"Lỗi chính: {e}")
        #messagebox.showerror("Lỗi", f"Lỗi khởi động ứng dụng: {e}")
    finally:
        cleanup()

# (Removed duplicate definitions below to avoid NameError/overrides)
