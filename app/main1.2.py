import os
import re
import time
import json 
import tkinter as tk
from tkinter import ttk, #messagebox
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import pytz
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import StaleElementReferenceException, NoSuchElementException, TimeoutException
from cryptography.fernet import Fernet
import psycopg2

from bs4 import BeautifulSoup
import sys
import logging
import requests
import threading
from typing import List, Tuple, Optional, Dict, Any
from dataclasses import dataclass

# Cấu hình logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('app.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

@dataclass
class Config:
    """Cấu hình ứng dụng"""
    DRIVER_LINK: str = "https://kpp.bankplus.vn"
    FOLDER_RESULT: str = "ket_qua"
    TITLE: str = "Thông báo"
    CONFIG_FILE: str = "config.json"
    ICON_FILE: str = "viettelpay.ico"
    COPYRIGHT_KEY: bytes = b"h_ThisAAutoToolVjppro-CopyRight-ByCAOAC7690="
    STATUS_COMPLETE: str = "Đã xử lý"
    STATUS_INCOMPLETE: str = "Chưa xử lý"
    
    # Service types
    SERVICES = {
        'payment_internet': 'payment_internet',
        'payment_card': 'deb_cart', 
        'lookup_card': 'lookup_cart',
        'lookup_ftth': 'lookup_ftth',
        'payment_evn': 'deb_evn'
    }
    
    # API Configuration
    API_BASE_URL: str = os.getenv('API_BASE_URL', "http://127.0.0.1:8080")
    NODE_SERVER_URL: str = os.getenv('NODE_SERVER_URL', "http://127.0.0.1:5000")
    API_TIMEOUT: int = 10
    
    # PIN Configuration
    DEFAULT_PIN: str = os.getenv('APP_DEFAULT_PIN', '9999')
    
    @classmethod
    def load_from_config(cls):
        """Load cấu hình từ file config.json"""
        try:
            config_path = os.path.join(os.path.dirname(__file__), 'config.json')
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    config_data = json.load(f)
                    if 'default_pin' in config_data:
                        cls.DEFAULT_PIN = config_data['default_pin']
                        print(f"[CONFIG] Đã load PIN từ config: {cls.DEFAULT_PIN}")
            else:
                print(f"[CONFIG] Không tìm thấy file config.json tại {config_path}")
        except Exception as e:
            print(f"[CONFIG] Lỗi load config: {e}")

# Global variables
stop_flag = False
driver = None
dbfiles = None
times_exits = {}
api_server_thread = None
automation_lock = threading.Lock()
AUTOMATION_MAX_RETRIES = 1  # Mặc định chỉ thử 1 lần cho mỗi mã
DIRECT_DB_MODE = True  # Chỉ xử lý trực tiếp với database, không phụ thuộc Node API
DB_DATABASE_URL = os.getenv('DATABASE_URL', 'postgres://postgres:123456@localhost:5432/autogachno')
DB_MAIN_USER_ID = os.getenv('MAIN_USER_ID', 'admin-local')
DB_MAIN_USER_EMAIL = os.getenv('MAIN_USER_EMAIL', 'Demodiemthu')

# Đăng nhập tự động khi mở trình duyệt
LOGIN_USERNAME = os.getenv('APP_LOGIN_USERNAME', '1000460100_VTP_00073_DB')
LOGIN_PASSWORD = os.getenv('APP_LOGIN_PASSWORD', '686886')

# Load cấu hình từ file config.json
Config.load_from_config()


def read_config():
    """Không dùng đọc cấu hình nữa."""
    return ""


def set_file_config(files):
    try:
        with open(Config.CONFIG_FILE, "r", encoding='utf-8') as file:
            data = json.load(file)
            data["files"] = files
        with open(Config.CONFIG_FILE, "w", encoding='utf-8') as file:
            json.dump(data, file, indent=4, ensure_ascii=False)
    except Exception as e:
        logger.error(f"Lỗi lưu cấu hình: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi lưu cấu hình: {e}")

def handle_key_active(key):
    try:
        key = (key.get()).strip()
        # TODO: Implement key validation through web API instead of local database
        #messagebox.showwarning(Config.TITLE, "Tính năng kích hoạt key đang được phát triển")
    except Exception as e:
        logger.error(f"Lỗi kích hoạt key: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi kích hoạt: {e}")

def check_exits_key(key):
    return True

def show_activate_form():
    # Không dùng kích hoạt key nữa
    pass

def db_ensure_user(user_id: str, email: str) -> None:
    try:
        with psycopg2.connect(DB_DATABASE_URL) as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT 1 FROM users WHERE id = %s", (user_id,))
                exists = cur.fetchone()
                if exists:
                    return
                cur.execute(
                    """
                    INSERT INTO users (id, email, first_name, last_name, role, status)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """,
                    (user_id, email, 'Admin', 'Local', 'admin', 'active')
                )
                print(f"[DB] Tạo user mặc định {email} ({user_id})")
    except Exception as e:
        print(f"[DB] Lỗi đảm bảo user tồn tại: {e}")

def update_database_immediately(order_id: str, code: str, status: str, amount: Any, notes: str, details: Optional[Dict[str, Any]] = None) -> bool:
    """Update database ngay lập tức cho từng đơn xử lý xong."""
    print(f"💾 [DATABASE] Cập nhật database cho mã {code} - Order: {order_id}")
    print(f"   📊 Status: {status}, Amount: {amount}")
    print(f"   📝 Notes: {notes}")
    
    if details:
        print(f"   📋 Chi tiết FTTH:")
        for key, value in details.items():
            print(f"      • {key}: {value}")
    
    # Cập nhật trực tiếp DB theo yêu cầu
    try:
        with psycopg2.connect(os.getenv('DATABASE_URL', 'postgres://postgres:123456@localhost:5432/autogachno')) as conn:
            with conn.cursor() as cur:
                import json as pyjson
                result_obj = {
                    'code': code,
                    'status': 'completed' if status == 'success' else status,
                    'amount': str(amount) if amount is not None else None,
                    'notes': notes,
                    'details': details or None,
                }
                result_json = pyjson.dumps(result_obj, ensure_ascii=False)
                # Log điều kiện WHERE dễ đọc
                cur.execute("SELECT input_data, service_type, user_id FROM orders WHERE id = %s", (order_id,))
                row = cur.fetchone()
                if row:
                    input_data, service_type, user_id = row
                    print("   🔎 WHERE ví dụ:")
                    print(f"      WHERE input_data = '{(input_data or '').strip()}'")
                    print(f"        AND service_type = '{service_type}'")
                    print(f"        AND user_id = '{user_id}'")
                    print(f"        AND status IN ('pending','processing')")
                # Update theo id
                # Cập nhật orders (không ràng buộc status để đảm bảo luôn ghi nhận kết quả)
                cur.execute(
                    """
                    UPDATE orders
                    SET status = %s,
                        result_data = %s,
                        updated_at = NOW()
                    WHERE id = %s
                    RETURNING id
                    """,
                    ('completed' if status == 'success' else status, result_json, order_id)
                )
                ok_order = cur.fetchone()
                if ok_order:
                    print(f"   ✅ Updated order {order_id}")
                else:
                    print("   ⚠️ Không có bản ghi phù hợp để update (orders) hoặc đã ở trạng thái cuối cùng")

                # Cập nhật service_transactions cho đúng code (ưu tiên theo order_id + code)
                cur.execute(
                    """
                    UPDATE service_transactions
                    SET status = %s,
                        amount = COALESCE(%s, amount),
                        notes = %s,
                        processing_data = %s,
                        updated_at = NOW()
                    WHERE order_id = %s
                      AND code = %s
                      AND status IN ('pending','processing')
                    RETURNING id
                    """,
                    (
                        'success' if status == 'success' else 'failed',
                        str(amount) if isinstance(amount, (int, float)) else None,
                        notes,
                        result_json,
                        order_id,
                        code,
                    )
                )
                ok_tran = cur.fetchone()
                if ok_tran:
                    print(f"   ✅ Updated transaction for code {code}")
                else:
                    print("   ⚠️ Không có transaction phù hợp để update (order_id+code). Thử cập nhật transaction mới nhất theo code...")
                    # Fallback: cập nhật transaction mới nhất theo code và service_type
                    cur.execute(
                        """
                        SELECT st.id
                        FROM service_transactions st
                        JOIN orders o ON o.id = st.order_id
                        WHERE st.code = %s
                          AND o.service_type = %s
                          AND st.status IN ('pending','processing')
                        ORDER BY st.created_at DESC
                        LIMIT 1
                        """,
                        (code, service_type)
                    )
                    st_row = cur.fetchone()
                    if st_row:
                        st_id = st_row[0]
                        cur.execute(
                            """
                            UPDATE service_transactions
                            SET status = %s,
                                amount = COALESCE(%s, amount),
                                notes = %s,
                                processing_data = %s,
                                updated_at = NOW()
                            WHERE id = %s
                            RETURNING id
                            """,
                            (
                                'success' if status == 'success' else 'failed',
                                str(amount) if isinstance(amount, (int, float)) else None,
                                notes,
                                result_json,
                                st_id,
                            )
                        )
                        ok_tran2 = cur.fetchone()
                        if ok_tran2:
                            print(f"   ✅ Updated latest transaction for code {code} (fallback)")
                            ok_tran = ok_tran2
                        else:
                            print("   ⚠️ Không thể cập nhật transaction (fallback)")
                    else:
                        print("   ℹ️ Không còn transaction pending/processing nào cho code này (fallback)")

                # Dọn trùng: đóng các transaction khác cùng code còn pending/processing (nếu có)
                try:
                    cur.execute(
                        """
                        UPDATE service_transactions st
                        SET status = 'failed',
                            notes = CASE WHEN st.notes IS NULL OR st.notes = '' THEN %s ELSE st.notes || ' | ' || %s END,
                            updated_at = NOW()
                        FROM orders o
                        WHERE st.order_id = o.id
                          AND st.code = %s
                          AND st.order_id <> %s
                          AND o.service_type = %s
                          AND st.status IN ('pending','processing')
                        RETURNING st.id
                        """,
                        ("auto-closed dup", "auto-closed dup", code, order_id, service_type)
                    )
                    closed_dups = cur.fetchall()
                    if closed_dups:
                        print(f"   ✅ Đã đóng {len(closed_dups)} transaction trùng (pending/processing) cho {code}")
                except Exception as _e:
                    print(f"   ⚠️ Lỗi khi đóng transaction trùng: {_e}")

                # Kiểm tra lại còn transaction pending/processing cho mã này không
                try:
                    pendings_after = db_check_pending_orders_for_code(service_type, code, None)
                    if pendings_after:
                        print(f"   🔎 Còn pending/processing sau update cho {code}: {', '.join(pendings_after)}")
                    else:
                        print(f"   ✅ Không còn pending/processing cho {code} sau update")
                except Exception:
                    pass

                return bool(ok_order or ok_tran)
    except Exception as e:
        print(f"   ❌ Lỗi cập nhật DB trực tiếp: {e}")
        return False
    
def db_find_order_id(service_type: str, code: str, user_id: Optional[str] = None) -> Optional[str]:
    """Tìm order_id theo code từ bảng service_transactions đang pending/processing (ưu tiên), khớp service_type của orders."""
    try:
        with psycopg2.connect(DB_DATABASE_URL) as conn:
            with conn.cursor() as cur:
                if user_id:
                    cur.execute(
                        """
                        SELECT st.order_id
                        FROM service_transactions st
                        JOIN orders o ON o.id = st.order_id
                        WHERE st.code = %s
                          AND o.service_type = %s
                          AND o.user_id = %s
                          AND st.status IN ('pending','processing')
                        ORDER BY st.created_at DESC
                        LIMIT 1
                        """,
                        (code, service_type, user_id)
                    )
                else:
                    cur.execute(
                        """
                        SELECT st.order_id
                        FROM service_transactions st
                        JOIN orders o ON o.id = st.order_id
                        WHERE st.code = %s
                          AND o.service_type = %s
                          AND st.status IN ('pending','processing')
                        ORDER BY st.created_at DESC
                        LIMIT 1
                        """,
                        (code, service_type)
                    )
                row = cur.fetchone()
                return row[0] if row else None
    except Exception as e:
        print(f"[DB] Lỗi tìm orderId cho code='{code}': {e}")
        return None

def db_check_pending_orders_for_code(service_type: str, code: str, user_id: Optional[str] = None) -> List[str]:
    """Trả về danh sách order_id còn 'pending'/'processing' cho mã (code) và service_type."""
    try:
        with psycopg2.connect(DB_DATABASE_URL) as conn:
            with conn.cursor() as cur:
                if user_id:
                    cur.execute(
                        """
                        SELECT st.order_id
                        FROM service_transactions st
                        JOIN orders o ON o.id = st.order_id
                        WHERE st.code = %s
                          AND o.service_type = %s
                          AND o.user_id = %s
                          AND st.status IN ('pending','processing')
                        ORDER BY st.created_at DESC
                        """,
                        (code, service_type, user_id)
                    )
                else:
                    cur.execute(
                        """
                        SELECT st.order_id
                        FROM service_transactions st
                        JOIN orders o ON o.id = st.order_id
                        WHERE st.code = %s
                          AND o.service_type = %s
                          AND st.status IN ('pending','processing')
                        ORDER BY st.created_at DESC
                        """,
                        (code, service_type)
                    )
                rows = cur.fetchall()
                return [r[0] for r in rows]
    except Exception as e:
        print(f"[DB] Lỗi kiểm tra pending cho code='{code}': {e}")
        return []

def db_insert_orders_from_lines(service_type: str, user_id: str, lines: List[str]) -> int:
    codes = [ln.strip() for ln in lines if ln and ln.strip()]
    if not codes:
        return 0
    try:
        with psycopg2.connect(DB_DATABASE_URL) as conn:
            with conn.cursor() as cur:
                count = 0
                for code in codes:
                    cur.execute(
                        """
                        INSERT INTO orders (user_id, service_type, status, input_data)
                        VALUES (%s, %s, 'pending', %s)
                        RETURNING id
                        """,
                        (user_id, service_type, code)
                    )
                    cur.fetchone()
                    count += 1
                print(f"[DB] Đã tạo {count} orders cho service '{service_type}'.")
                return count
    except Exception as e:
        print(f"[DB] Lỗi insert orders: {e}")
        return 0


def send_callback_with_retry(order_id: str, code: str, status: str, amount: Any, notes: str, details: Optional[Dict[str, Any]] = None) -> bool:
    """Gửi callback với retry logic tối đa 3 lần."""
    print(f"🔄 [CALLBACK] Bắt đầu gửi callback cho mã {code} - Order: {order_id}")
    print(f"   📊 Status: {status}, Amount: {amount}, Notes: {notes}")
    
    for attempt in range(3):
        try:
            payload = {
                "orderId": order_id,
                "code": code,
                "status": status,
                "amount": str(amount) if isinstance(amount, (int, float)) else None,
                "notes": notes,
            }
            
            if details:
                payload["data"] = {
                    "type": "ftth_details",
                    "details": details,
                }
                print(f"   📋 Data: {details}")
            
            print(f"   📤 Lần thử {attempt + 1}/3: Gửi đến {Config.NODE_SERVER_URL}/api/automation/callback")
            response = requests.post(
                f"{Config.NODE_SERVER_URL}/api/automation/callback",
                json=payload,
                timeout=10,  # Tăng timeout
            )
            
            if response.status_code == 200:
                print(f"   ✅ Callback thành công cho {code} (lần {attempt + 1})")
                logger.info(f"Callback thành công cho {code} (lần {attempt + 1})")
                return True
            else:
                print(f"   ❌ Callback thất bại cho {code} (lần {attempt + 1}): HTTP {response.status_code}")
                logger.warning(f"Callback thất bại cho {code} (lần {attempt + 1}): {response.status_code}")
                
        except Exception as e:
            print(f"   ❌ Callback lỗi cho {code} (lần {attempt + 1}): {e}")
            logger.warning(f"Callback lỗi cho {code} (lần {attempt + 1}): {e}")
        
        if attempt < 2:  # Còn cơ hội retry
            print(f"   ⏳ Chờ 2s trước khi retry...")
            time.sleep(2)  # Delay giữa các lần retry
    
    print(f"   💥 Callback thất bại sau 3 lần thử cho {code}")
    logger.error(f"Callback thất bại sau 3 lần thử cho {code}")
    return False


def get_exe_dir():
    exe_path = sys.argv[0]
    exe_dir = os.path.dirname(exe_path)
    return exe_dir

def start_api_server():
    """Bỏ qua mock API server: danh sách đơn lấy từ Database qua Node API."""
    logger.info("Bỏ qua mock_api_server: dùng trực tiếp Node API/DB")
    return False

def ensure_driver_and_login() -> bool:
    """Đảm bảo Chrome driver đã sẵn sàng (bỏ qua đăng nhập)."""
    global driver
    try:
        if driver is None:
            username = LOGIN_USERNAME or "default"
            driver = initialize_browser(username)
            if not driver:
                logger.error("Không thể khởi tạo trình duyệt")
                return False
        return True
    except Exception as e:
        logger.error(f"ensure_driver_and_login error: {e}")
        return False

def process_lookup_ftth_codes(codes: List[str], order_id: Optional[str] = None):
    """Xử lý tra cứu FTTH không cần GUI, điều khiển selenium trực tiếp."""
    print(f"🚀 [AUTOMATION] Bắt đầu xử lý FTTH cho {len(codes)} mã")
    print(f"   📋 Order ID: {order_id or 'Không có'}")
    
    if not ensure_driver_and_login():
        print("   ❌ Không thể khởi tạo driver hoặc đăng nhập")
        logger.error("Không thể khởi tạo driver hoặc đăng nhập")
        return
    
    print("   ✅ Bắt đầu xử lý (bỏ qua đăng nhập)")
    
    try:
        with automation_lock:
            print("   🔒 Đã khóa automation, điều hướng đến trang FTTH...")
            navigate_to_ftth_page_and_select_radio()
            print("   ✅ Đã điều hướng thành công đến trang FTTH")
            
            # Hiển thị tiến trình cho từng order nếu có
            if order_id:
                print("   📋 Danh sách mã sẽ xử lý:")
                for idx, cb in enumerate(codes, 1):
                    print(f"      {idx}. {cb}")
            
            results = []
            for idx, cbil in enumerate(codes, 1):
                print(f"\n📱 [MÃ {idx}/{len(codes)}] Xử lý mã: {cbil}")
                success = False
                for attempt in range(AUTOMATION_MAX_RETRIES):  # Retry tối đa cấu hình
                    try:
                        cbil = (cbil or "").strip()
                        if not cbil:
                            print(f"   ⚠️  Mã rỗng, bỏ qua")
                            break
                        
                        if attempt > 0:
                            print(f"   🔄 Retry lần {attempt + 1}/{AUTOMATION_MAX_RETRIES} cho mã {cbil}")
                            logger.info(f"Retry lần {attempt + 1} cho mã {cbil}")
                            driver.refresh()
                            time.sleep(2)
                            navigate_to_ftth_page_and_select_radio()
                        else:
                            print(f"   🎯 Lần thử đầu tiên cho mã {cbil}")
                        
                        root.update() if 'root' in globals() else None
                        time.sleep(0.5)
                        
                        print(f"   📝 Điền mã thuê bao: {cbil}")
                        customer = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:contractCode")))
                        customer.clear()
                        customer.send_keys(cbil)
                        
                        print(f"   🔍 Nhấn nút KIỂM TRA...")
                        payment_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:btnPay0")))
                        payment_button.click()
                        time.sleep(1)
                        
                        print(f"   ⏳ Chờ modal loading...")
                        WebDriverWait(driver, 16).until(EC.invisibility_of_element_located((By.ID, "payMoneyForm:j_idt6_modal")))
                        
                        print(f"   📊 Lấy thông tin kết quả...")
                        element41 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:j_idt41")))
                        _is_amount, amount, _pid = amount_by_cbil(cbil, element41, False)
                        details = extract_ftth_details_from_page()
                        
                        print(f"   ✅ Xử lý thành công: Amount = {amount}")
                        if details:
                            print(f"   📋 Chi tiết FTTH:")
                            for key, value in details.items():
                                print(f"      • {key}: {value}")
                        
                        # Thành công, thoát khỏi retry loop
                        success = True
                        results.append({"code": cbil, "amount": amount, "status": "success"})
                        
                        # Update database ngay lập tức cho từng đơn xử lý xong
                        if order_id:
                            print(f"   💾 Bắt đầu update database...")
                            db_success = update_database_immediately(order_id, cbil, "success", amount, "FTTH lookup ok", details)
                            if not db_success:
                                logger.warning(f"Database update thất bại cho {cbil}")
                        else:
                            print(f"⚠️  [WARNING] Không có order_id, bỏ qua database update cho {cbil}")
                        
                        print(f"   🎉 Hoàn thành xử lý mã {cbil}")
                        break
                        
                    except Exception as e:
                        print(f"   ❌ Lần thử {attempt + 1} thất bại: {e}")
                        logger.warning(f"Lần thử {attempt + 1} thất bại cho {cbil}: {e}")
                        if attempt < (AUTOMATION_MAX_RETRIES - 1):  # Còn cơ hội retry
                            print(f"   ⏳ Chờ 1s trước khi retry...")
                            time.sleep(1)  # Delay thông minh
                            continue
                        else:  # Hết retry
                            print(f"   💥 Hết retry, mã {cbil} thất bại hoàn toàn")
                            logger.error(f"FTTH code {cbil} thất bại sau {AUTOMATION_MAX_RETRIES} lần thử: {e}")
                            results.append({"code": cbil, "amount": None, "status": "failed", "message": str(e)})
                            
                            # Update database cho trường hợp thất bại
                            if order_id:
                                print(f"   💾 Update database cho trường hợp thất bại...")
                                db_success = update_database_immediately(order_id, cbil, "failed", None, str(e), None)
                                if not db_success:
                                    logger.warning(f"Database update thất bại cho {cbil}")
                            else:
                                print(f"⚠️  [WARNING] Không có order_id, bỏ qua database update cho {cbil}")
                
                if not success:
                    print(f"   💥 Mã {cbil} không thể xử lý sau 3 lần thử")
                    logger.error(f"Mã {cbil} không thể xử lý sau 3 lần thử")
            
            print(f"\n📊 [KẾT QUẢ] Tổng kết xử lý FTTH:")
            print(f"   ✅ Thành công: {len([r for r in results if r['status'] == 'success'])} mã")
            print(f"   ❌ Thất bại: {len([r for r in results if r['status'] == 'failed'])} mã")
            print(f"   📋 Tổng cộng: {len(results)} mã")
            
            logger.info(f"FTTH processed: {len(results)} items")
    except Exception as e:
        print(f"   💥 Lỗi tổng thể: {e}")
        logger.error(f"process_lookup_ftth_codes error: {e}")

def get_data_ftth(text_widget, order_entry: Optional[ttk.Entry] = None):
    """Lấy dữ liệu API cho Tra cứu FTTH và set Order ID nếu có"""
    try:
        data = fetch_api_data("tra_cuu_ftth")
        if data and "subscriber_codes" in data:
            codes = data["subscriber_codes"]
            populate_text_widget(text_widget, codes)
            
            # Tự động điền mã PIN từ config (nếu có pin_widget)
            if 'pin_widget' in locals() and pin_widget:
                print(f"[DEBUG] Tự động điền mã PIN: {Config.DEFAULT_PIN}")
                populate_entry_widget(pin_widget, Config.DEFAULT_PIN)
            order_id = data.get("order_id")
            code_map = data.get("code_order_map") or []
            info_msg = f"Đã tải {len(codes)} mã thuê bao FTTH"
            if order_id:
                print(f"[INFO] Order ID từ DB (gần nhất): {order_id}", flush=True)
                logger.info(f"Order ID từ DB (FTTH, gần nhất): {order_id}")
                info_msg += f"\nOrder ID (gần nhất): {order_id}"
            # In mapping chi tiết mã -> orderId nếu có
            if code_map:
                print("[INFO] Mapping mã -> Order ID:", flush=True)
                for item in code_map:
                    try:
                        print(f"  {item.get('code')}: {item.get('orderId')}", flush=True)
                    except Exception:
                        pass
            #messagebox.showinfo(Config.TITLE, info_msg)
        else:
            #messagebox.showwarning(Config.TITLE, "Không có dữ liệu từ DB")
    except Exception as e:
        logger.error(f"Lỗi lấy dữ liệu FTTH: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi lấy dữ liệu FTTH: {e}")


def navigate_to_ftth_page_and_select_radio():
    """Đi tới trang FTTH và chọn radio 'Số thuê bao'"""
    try:
        # Bỏ qua kiểm tra đăng nhập theo yêu cầu
        target_url = "https://kpp.bankplus.vn/pages/newInternetTelevisionViettel.jsf?serviceCode=000003&serviceType=INTERNET"
        driver.get(target_url)
        # Chờ input mã thuê bao xuất hiện
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:contractCode")))
        # Chọn radio Số thuê bao (id payMoneyForm:console:3)
        try:
            radio_input = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, "payMoneyForm:console:3")))
            radio_box = radio_input.find_element(By.XPATH, "../../div[contains(@class,'ui-radiobutton-box')]")
            radio_box.click()
        except Exception:
            # fallback click vào label
            try:
                label = driver.find_element(By.XPATH, "//label[@for='payMoneyForm:console:3']")
                label.click()
            except Exception:
                pass
    except Exception as e:
        logger.warning(f"Không thể điều hướng FTTH hoặc chọn radio: {e}")
        raise  # Re-raise để caller biết có lỗi

def extract_ftth_details_from_page() -> Dict[str, Any]:
    """Trích thông tin chi tiết FTTH từ trang hiện tại sau khi nhấn KIỂM TRA."""
    details: Dict[str, Any] = {}
    try:
        html_content = driver.page_source
        soup = BeautifulSoup(html_content, 'html.parser')
        label_to_key = {
            'Mã hợp đồng:': 'contract_code',
            'Chủ hợp đồng:': 'contract_owner',
            'Số thuê bao đại diện:': 'representative_subscriber',
            'Dịch vụ:': 'service',
            'Số điện thoại liên hệ:': 'contact_phone',
            'Nợ cước:': 'debt_amount',
        }
        # Duyệt các hàng hiển thị label-value
        for row in soup.find_all('div', class_='row'):
            cols = row.find_all('div', class_='col-6')
            if len(cols) != 2:
                continue
            label_tag = cols[0].find('label')
            value_p = cols[1].find('p')
            if not label_tag or not value_p:
                continue
            label_text = label_tag.get_text(strip=True)
            value_text = value_p.get_text(strip=True)
            key = label_to_key.get(label_text)
            if not key:
                continue
            if key == 'debt_amount':
                try:
                    # Lấy số từ chuỗi "170,000 VND"
                    num_str = re.findall(r"[\d\.,]+", value_text)
                    if num_str:
                        details[key] = int(num_str[0].replace('.', '').replace(',', ''))
                    else:
                        details[key] = None
                except Exception:
                    details[key] = None
            else:
                details[key] = value_text
    except Exception as e:
        logger.warning(f"Lỗi trích chi tiết FTTH: {e}")
    return details


def lookup_ftth(tkinp_ctm, tkinp_ctmed, tkinp_order: Optional[ttk.Entry] = None):
    
    try:
        delete_ctmed(tkinp_ctmed)
        update_stop_flag()
        navigate_to_ftth_page_and_select_radio()
        cbils = tkinp_ctm.get("1.0", "end-1c").splitlines()
        if not valid_data([cbils]):
            return False
        # Hiển thị mapping mã -> orderId trước khi xử lý
        print("Order ID:")
        code_to_order: Dict[str, Optional[str]] = {}
        for raw in cbils:
            c = (raw or "").strip()
            if not c:
                continue
            oid = db_find_order_id('tra_cuu_ftth', c, None)
            code_to_order[c] = oid
            print(f"  {c}: {oid if oid else 'Không tìm thấy'}")

        data = []
        for cbil in cbils:
            root.update()
            time.sleep(1)
            cbil = cbil.strip()
            if not stop_flag and cbil.strip() != "":
                success = False
                for attempt in range(3):  # Retry tối đa 3 lần
                    try:
                        # Log tiến trình theo orderId
                        print(f"   🔧 Đang xử lý {cbil} | Order ID: {code_to_order.get(cbil) if code_to_order.get(cbil) else 'Không tìm thấy'}")
                        if attempt > 0:
                            logger.info(f"Retry lần {attempt + 1} cho mã {cbil}")
                            driver.refresh()
                            time.sleep(2)
                            navigate_to_ftth_page_and_select_radio()
                        
                        # Điền vào form
                        customer = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:contractCode")))
                        customer.clear()
                        customer.send_keys(cbil)
                        # Nhấn nút thanh toán
                        payment_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:btnPay0")))
                        payment_button.click()
                        time.sleep(1)
                        WebDriverWait(driver, 16).until(EC.invisibility_of_element_located((By.ID, "payMoneyForm:j_idt6_modal")))
                        # Kiểm tra thông báo lỗi (nếu có) và dừng retry khi chứa từ "không"
                        alert_text = get_error_alert_text()
                        if alert_text and ("không" in alert_text.lower()):
                            note_err = alert_text.strip()
                            data.append([cbil, 0, Config.STATUS_INCOMPLETE])
                            insert_ctmed(tkinp_ctmed, f"{cbil} - Lỗi | {note_err}")
                            order_id_val = code_to_order.get(cbil) or db_find_order_id('tra_cuu_ftth', cbil.strip(), None)
                            if order_id_val:
                                try:
                                    _ = update_database_immediately(order_id_val, cbil, "failed", None, note_err, None)
                                except Exception as _e:
                                    logger.warning(f"DB update lỗi (failed) cho {cbil}: {_e}")
                            break
                        element41 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:j_idt41")))
                        is_amount, amount, payment_id = amount_by_cbil(cbil, element41, False)
                        details = extract_ftth_details_from_page()
                        note_text = f"HD:{details.get('contract_code','')} | Chu:{details.get('contract_owner','')} | SDT:{details.get('contact_phone','')} | No:{details.get('debt_amount','')}"
                        
                        # Thành công, thoát khỏi retry loop
                        success = True
                        data.append([cbil, amount, note_text])
                        insert_ctmed(tkinp_ctmed, f"{cbil} - {amount} | {note_text}")
                        
                        # Update trực tiếp DB: dùng mapping đã tra ở trên
                        order_id_val = code_to_order.get(cbil) or db_find_order_id('tra_cuu_ftth', cbil.strip(), None)
                        if not order_id_val:
                            print(f"   ⚠️ Không tìm thấy orderId cho code='{cbil}'. Bỏ qua update.")
                            # Hiển thị danh sách order_id pending/processing (nếu có) để dễ kiểm tra
                            pendings = db_check_pending_orders_for_code('tra_cuu_ftth', cbil.strip(), None)
                            if pendings:
                                print(f"   🔎 Pending/processing orderIds for {cbil}: {', '.join(pendings)}")
                        else:
                            try:
                                db_success = update_database_immediately(order_id_val, cbil, "success", amount, note_text, details)
                                if not db_success:
                                    logger.warning(f"Database update thất bại cho {cbil}")
                            except Exception as _e:
                                logger.warning(f"DB update lỗi (success) cho {cbil}: {_e}")
                        
                        tkinp_ctm.delete("1.0", "1.end+1c")
                        break
                        
                    except Exception as e:
                        logger.warning(f"Lần thử {attempt + 1} thất bại cho {cbil}: {e}")
                        if attempt < (AUTOMATION_MAX_RETRIES - 1):  # Còn cơ hội retry
                            time.sleep(1)  # Delay thông minh
                            continue
                        else:  # Hết retry
                            logger.error(f"Mã {cbil} thất bại sau {AUTOMATION_MAX_RETRIES} lần thử: {e}")
                            data.append([cbil, 0, Config.STATUS_INCOMPLETE])
                            insert_ctmed(tkinp_ctmed, f"{cbil} - Lỗi")
                            tkinp_ctm.delete("1.0", "1.end+1c")
                            break
                
                if not success:
                    logger.error(f"Mã {cbil} không thể xử lý sau 3 lần thử")
        time.sleep(2)
        if len(data) > 0:
            name_dir = "Tra cứu FTTH"
            export_excel(data, name_dir)
    except Exception as e:
        logger.error(f"Lỗi tra cứu FTTH: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi tra cứu FTTH: {e}")

def form_lookup_ftth():
    cus_frm = tk.Frame(root)
    cus_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    ctm_frm = tk.Frame(cus_frm)
    ctm_frm.pack(expand=True, side="left")
    ctmed_frm = tk.Frame(cus_frm)
    ctmed_frm.pack(expand=True, side="right")
    btn_frm = tk.Frame(root)
    btn_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    tklbl_ctm = tk.Label(ctm_frm, text="Số thuê bao (mỗi dòng tạo 1 order)")
    tklbl_ctm.pack(side="top")
    tkinp_ctm = tk.Text(ctm_frm, height=16, width=24)
    tkinp_ctm.pack(side="left", pady=8)
    tklbl_ctm = tk.Label(ctmed_frm, text="Đã xử lý")
    tklbl_ctm.pack(side="top")
    tkinp_ctmed = tk.Text(ctmed_frm, height=16, width=32, bg="#ccc", state="disabled")
    tkinp_ctmed.pack(side="left", pady=8)
    # Loại bỏ ô Order ID: orderId sẽ mặc định dùng chính cbil; nút tạo orders từ DB
    style = ttk.Style()
    style.configure("Red.TButton", foreground="red")
    style.configure("Blue.TButton", foreground="blue")
    style.configure("Green.TButton", foreground="green")
    tkbtn_get_data = ttk.Button(btn_frm, text="Get dữ liệu", command=lambda: get_data_ftth(tkinp_ctm, None))
    tkbtn_get_data.pack(side='left', padx=5, pady=5)
    tkbtn_get_data.configure(style="Green.TButton")
    # Bỏ nút tạo orders: chỉ get dữ liệu và xử lý cập nhật cho DB sẵn có
    tkbtn_payment = ttk.Button(btn_frm, text="Bắt đầu", command=lambda: lookup_ftth(tkinp_ctm, tkinp_ctmed, None))
    tkbtn_payment.pack(side='left', padx=5, pady=5)
    tkbtn_payment.configure(style="Blue.TButton") 
    tkbtn_destroy = ttk.Button(btn_frm, text="Dừng lại", command=stop_tool)
    tkbtn_destroy.pack(side='right', padx=5, pady=5)
    tkbtn_destroy.configure(style="Red.TButton") 


def process_evn_payment_codes(codes: List[str], order_id: Optional[str] = None):
    """Xử lý thanh toán điện EVN không cần GUI, điều khiển selenium trực tiếp."""
    print(f"🚀 [AUTOMATION] Bắt đầu xử lý EVN cho {len(codes)} mã")
    print(f"   📋 Order ID: {order_id or 'Không có'}")
    
    if not ensure_driver_and_login():
        print("   ❌ Không thể khởi tạo driver hoặc đăng nhập")
        logger.error("Không thể khởi tạo driver hoặc đăng nhập")
        return
    
    try:
        with automation_lock:
            print("   🔒 Đã khóa automation, điều hướng đến trang EVN...")
            navigate_to_evn_page()
            print("   ✅ Đã điều hướng thành công đến trang EVN")
            
            results = []
            for idx, cbil in enumerate(codes, 1):
                print(f"\n📱 [MÃ {idx}/{len(codes)}] Xử lý mã: {cbil}")
                success = False
                for attempt in range(AUTOMATION_MAX_RETRIES):  # Retry tối đa theo cấu hình
                    try:
                        cbil = (cbil or "").strip()
                        if not cbil:
                            print(f"   ⚠️  Mã rỗng, bỏ qua")
                            break
                        
                        if attempt > 0:
                            print(f"   🔄 Retry lần {attempt + 1}/3 cho mã {cbil}")
                            logger.info(f"Retry lần {attempt + 1} cho mã {cbil}")
                            driver.refresh()
                            time.sleep(2)
                            navigate_to_evn_page()
                        else:
                            print(f"   🎯 Lần thử đầu tiên cho mã {cbil}")
                        
                        root.update() if 'root' in globals() else None
                        time.sleep(0.5)
                        
                        print(f"   📝 Điền mã hóa đơn: {cbil}")
                        customer = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "collectElectricBillForm:j_idt29")))
                        customer.clear()
                        customer.send_keys(cbil)
                        
                        print(f"   🔍 Nhấn nút KIỂM TRA...")
                        payment_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "collectElectricBillForm:j_idt31")))
                        payment_button.click()
                        time.sleep(1)
                        
                        print(f"   ⏳ Chờ modal loading...")
                        WebDriverWait(driver, 16).until(EC.invisibility_of_element_located((By.ID, "payMoneyForm:j_idt6_modal")))
                        
                        print(f"   📊 Lấy thông tin kết quả...")
                        element41 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:j_idt41")))
                        _is_amount, amount, _pid = amount_by_cbil(cbil, element41, False)
                        
                        print(f"   ✅ Xử lý thành công: Amount = {amount}")
                        
                        # Thành công, thoát khỏi retry loop
                        success = True
                        results.append({"code": cbil, "amount": amount, "status": "success"})
                        
                        # Update database ngay lập tức cho từng đơn xử lý xong
                        if order_id:
                            print(f"   💾 Bắt đầu update database...")
                            db_success = update_database_immediately(order_id, cbil, "success", amount, "EVN payment ok", None)
                            if not db_success:
                                logger.warning(f"Database update thất bại cho {cbil}")
                        else:
                            print(f"⚠️  [WARNING] Không có order_id, bỏ qua database update cho {cbil}")
                        
                        print(f"   🎉 Hoàn thành xử lý mã {cbil}")
                        break
                        
                    except Exception as e:
                        print(f"   ❌ Lần thử {attempt + 1} thất bại: {e}")
                        logger.warning(f"Lần thử {attempt + 1} thất bại cho {cbil}: {e}")
                        if attempt < 2:  # Còn cơ hội retry
                            print(f"   ⏳ Chờ 1s trước khi retry...")
                            time.sleep(1)  # Delay thông minh
                            continue
                        else:  # Hết retry
                            print(f"   💥 Hết retry, mã {cbil} thất bại hoàn toàn")
                            logger.error(f"EVN code {cbil} thất bại sau 3 lần thử: {e}")
                            results.append({"code": cbil, "amount": None, "status": "failed", "message": str(e)})
                            
                            # Update database cho trường hợp thất bại
                            if order_id:
                                print(f"   💾 Update database cho trường hợp thất bại...")
                                db_success = update_database_immediately(order_id, cbil, "failed", None, str(e), None)
                                if not db_success:
                                    logger.warning(f"Database update thất bại cho {cbil}")
                            else:
                                print(f"⚠️  [WARNING] Không có order_id, bỏ qua database update cho {cbil}")
                
                if not success:
                    print(f"   💥 Mã {cbil} không thể xử lý sau 3 lần thử")
                    logger.error(f"Mã {cbil} không thể xử lý sau 3 lần thử")
            
            print(f"\n📊 [KẾT QUẢ] Tổng kết xử lý EVN:")
            print(f"   ✅ Thành công: {len([r for r in results if r['status'] == 'success'])} mã")
            print(f"   ❌ Thất bại: {len([r for r in results if r['status'] == 'failed'])} mã")
            print(f"   📋 Tổng cộng: {len(results)} mã")
            
            logger.info(f"EVN processed: {len(results)} items")
    except Exception as e:
        print(f"   💥 Lỗi tổng thể: {e}")
        logger.error(f"process_evn_payment_codes error: {e}")


def get_data_evn(text_widget, phone_widget, pin_widget):
    """Lấy dữ liệu API cho Gạch điện EVN"""
    try:
        data = fetch_api_data("gach_dien_evn")
        if data:
            if "bill_codes" in data:
                populate_text_widget(text_widget, data["bill_codes"])
            if "receiver_phone" in data:
                populate_entry_widget(phone_widget, data["receiver_phone"])
            
            # Tự động điền mã PIN từ config
            print(f"[DEBUG] Tự động điền mã PIN: {Config.DEFAULT_PIN}")
            populate_entry_widget(pin_widget, Config.DEFAULT_PIN)
            
            count = len(data.get("bill_codes", []))
            order_id = data.get("order_id")
            info_msg = f"Đã tải {count} mã hóa đơn điện EVN"
            if order_id:
                print(f"[INFO] Order ID từ DB: {order_id}", flush=True)
                logger.info(f"Order ID từ DB (EVN): {order_id}")
                info_msg += f"\nOrder ID: {order_id}"
            #messagebox.showinfo(Config.TITLE, info_msg)
        else:
            #messagebox.showwarning(Config.TITLE, "Không có dữ liệu từ DB")
    except Exception as e:
        logger.error(f"Lỗi lấy dữ liệu EVN: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi lấy dữ liệu EVN: {e}")


def navigate_to_evn_page():
    """Điều hướng đến trang thanh toán điện EVN."""
    try:
        
        target_url = "https://kpp.bankplus.vn/pages/collectElectricBill.jsf?serviceCode=EVN"
        driver.get(target_url)
        # Chờ input mã thuê bao xuất hiện
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "collectElectricBillForm:j_idt29")))
        time.sleep(2)
    except Exception as e:
        logger.warning(f"Không thể điều hướng EVN: {e}")
        raise

def debt_electric(tkinp_ctm, tkinp_ctmed, tkinp_phone, tkinp_pin):
    try:
        delete_ctmed(tkinp_ctmed)
        update_stop_flag()
        pin = tkinp_pin.get()
        phone = tkinp_phone.get()
        cbils = tkinp_ctm.get("1.0", "end-1c").splitlines()
        # EVN: chỉ 1 form nhập mã hóa đơn. Không cần số điện thoại và mã pin.
        if not valid_data([cbils]):
            return False
        data = []
        for cbil in cbils:
            cbil = cbil.strip()
            root.update()
            time.sleep(1)
            if not stop_flag and cbil.strip() != "":
                try:
                    # Điền mã hóa đơn
                    customer = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:billCodeId")))
                    customer.clear()
                    customer.send_keys(cbil)
                    time.sleep(0.5)

                    # Nhấn Pay/Check
                    payment = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:btnPay")))
                    payment.click()
                    time.sleep(1)

                    # Chờ modal ẩn
                    WebDriverWait(driver, 16).until(EC.invisibility_of_element_located((By.ID, "payMoneyForm:j_idt6_modal")))

                    # Bắt lỗi EVN nếu có
                    alert_text = get_error_alert_text()
                    if alert_text and ("không" in alert_text.lower() or "đã xảy ra lỗi" in alert_text.lower()):
                        note_err = alert_text.strip()
                        data.append([cbil, 0, Config.STATUS_INCOMPLETE])
                        insert_ctmed(tkinp_ctmed, f"{cbil} - Lỗi | {note_err}")
                        # Cập nhật DB failed cho EVN nếu có orderId
                        order_id_val = db_find_order_id('gach_dien_evn', cbil.strip(), None)
                        if order_id_val:
                            try:
                                _ = update_database_immediately(order_id_val, cbil, "failed", None, note_err, None)
                            except Exception as _e:
                                logger.warning(f"DB update lỗi (EVN failed) cho {cbil}: {_e}")
                        continue

                    # Lấy amount
                    lblamount = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:j_idt49")))
                    try:
                        text_of_amount = lblamount.text
                        amount_str = text_of_amount.replace('VND', '').replace('.', '')
                        amount = int(amount_str)
                    except:
                        amount = lblamount.text

                    # Xác nhận
                    time.sleep(0.5)
                    confirm = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:yesIdEVN")))
                    confirm.click()

                    # Thành công
                    data.append([cbil, amount, Config.STATUS_COMPLETE])
                    insert_ctmed(tkinp_ctmed, f"{cbil} - {amount}")
                    # Cập nhật DB success
                    order_id_val = db_find_order_id('gach_dien_evn', cbil.strip(), None)
                    if order_id_val:
                        try:
                            _ = update_database_immediately(order_id_val, cbil, "success", amount, "EVN payment ok", None)
                        except Exception as _e:
                            logger.warning(f"DB update lỗi (EVN success) cho {cbil}: {_e}")
                    continue
                except Exception as e:
                    data.append([cbil, 0, Config.STATUS_INCOMPLETE])
                    insert_ctmed(tkinp_ctmed, f"{cbil} - Lỗi")
                    continue
        time.sleep(2)
        if len(data) > 0:
            name_dir = "Thanh toán điện EVN"
            export_excel(data, name_dir)
    except Exception as e:
        logger.error(f"Lỗi thanh toán điện EVN: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi thanh toán điện EVN: {e}")

def form_debt_electric():
    cus_frm = tk.Frame(root)
    cus_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    ctm_frm = tk.Frame(cus_frm)
    ctm_frm.pack(expand=True, side="left")
    ctmed_frm = tk.Frame(cus_frm)
    ctmed_frm.pack(expand=True, side="right")
    # Loại bỏ ô nhập SĐT và PIN theo yêu cầu
    btn_frm = tk.Frame(root)
    btn_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    tklbl_ctm = tk.Label(ctm_frm, text="Mã thuê bao")
    tklbl_ctm.pack(side="top")
    tkinp_ctm = tk.Text(ctm_frm, height=18, width=24)
    tkinp_ctm.pack(side="left", pady=8)
    tklbl_ctm = tk.Label(ctmed_frm, text="Đã xử lý")
    tklbl_ctm.pack(side="top")
    tkinp_ctmed = tk.Text(ctmed_frm, width=32, height=18, bg="#ccc", state="disabled")
    tkinp_ctmed.pack(side="left", pady=8)
    # Placeholder rỗng để giữ hàm gọi hiện tại (không dùng)
    tkinp_phone = ttk.Entry(root)
    tkinp_pin = ttk.Entry(root)
    style = ttk.Style()
    style.configure("Red.TButton", foreground="red")
    style.configure("Blue.TButton", foreground="blue")
    style.configure("Green.TButton", foreground="green")
    tkbtn_get_data = ttk.Button(btn_frm, text="Get dữ liệu", command=lambda: get_data_evn(tkinp_ctm, tkinp_phone, tkinp_pin))
    tkbtn_get_data.pack(side='left', padx=5, pady=5)
    tkbtn_get_data.configure(style="Green.TButton")
    tkbtn_payment = ttk.Button(btn_frm, text="Bắt đầu", command=lambda: debt_electric(tkinp_ctm, tkinp_ctmed, tkinp_phone, tkinp_pin))
    tkbtn_payment.pack(side='left', padx=5, pady=5)
    tkbtn_payment.configure(style="Blue.TButton") 
    tkbtn_destroy = ttk.Button(btn_frm, text="Dừng lại", command=stop_tool)
    tkbtn_destroy.pack(side='right', padx=5, pady=5)
    tkbtn_destroy.configure(style="Red.TButton") 

def process_topup_multinetwork_codes(codes: List[str], order_id: Optional[str] = None):
    """Xử lý nạp tiền đa mạng - hỗ trợ cả nạp trả trước và gạch nợ trả sau."""
    print(f"🚀 [AUTOMATION] Bắt đầu xử lý Topup đa mạng cho {len(codes)} mã")
    print(f"   📋 Order ID: {order_id or 'Không có'}")
    
    if not ensure_driver_and_login():
        print("   ❌ Không thể khởi tạo driver hoặc đăng nhập")
        logger.error("Không thể khởi tạo driver hoặc đăng nhập")
        return
    
    try:
        with automation_lock:
            print("   🔒 Đã khóa automation, điều hướng đến trang Topup đa mạng...")
            navigate_to_topup_multinetwork_page()
            print("   ✅ Đã điều hướng thành công đến trang Topup đa mạng")
            
            results = []
            for idx, cbil in enumerate(codes, 1):
                print(f"\n📱 [MÃ {idx}/{len(codes)}] Xử lý mã: {cbil}")
                
                # Hiển thị tiến trình tương tự FTTH
                print(f"   🔄 Đang xử lý {cbil} | Order ID: {order_id or 'Không có'}")
                print(f"   📍 Loại dịch vụ: {'Nạp trả trước' if '|' in cbil else 'Gạch nợ trả sau'}")
                
                # Phân tích dữ liệu để xác định loại dịch vụ
                is_prepaid = '|' in cbil  # Nạp trả trước: có dấu | (sđt|số tiền)
                if is_prepaid:
                    # Nạp trả trước: sđt|số tiền
                    parts = cbil.split('|')
                    if len(parts) != 2:
                        print(f"   ❌ Sai định dạng: {cbil} (cần: sđt|số tiền)")
                        results.append({"code": cbil, "amount": None, "status": "failed", "message": "Sai định dạng"})
                        continue
                    
                    phone_number = parts[0].strip()
                    amount_str = parts[1].strip()
                    try:
                        amount = int(amount_str)
                        valid_amounts = [10000, 20000, 30000, 50000, 100000, 200000, 300000, 500000]
                        if amount not in valid_amounts:
                            print(f"   ❌ Số tiền không hợp lệ: {amount} (chỉ cho phép: {valid_amounts})")
                            results.append({"code": cbil, "amount": None, "status": "failed", "message": f"Số tiền {amount} không hợp lệ"})
                            continue
                    except ValueError:
                        print(f"   ❌ Số tiền không hợp lệ: {amount_str}")
                        results.append({"code": cbil, "amount": None, "status": "failed", "message": "Số tiền không hợp lệ"})
                        continue
                    
                    print(f"   🎯 Nạp trả trước: {phone_number} | Số tiền: {amount:,}đ")
                    process_code = phone_number
                else:
                    # Gạch nợ trả sau: chỉ số điện thoại
                    phone_number = cbil.strip()
                    print(f"   🎯 Gạch nợ trả sau: {phone_number}")
                    process_code = phone_number
                
                success = False
                for attempt in range(3):  # Retry tối đa 3 lần
                    try:
                        if attempt > 0:
                            print(f"   🔄 Retry lần {attempt + 1}/3 cho mã {cbil}")
                            logger.info(f"Retry lần {attempt + 1} cho mã {cbil}")
                            driver.refresh()
                            time.sleep(2)
                            navigate_to_topup_multinetwork_page()
                        else:
                            print(f"   🎯 Lần thử đầu tiên cho mã {cbil}")
                        
                        root.update() if 'root' in globals() else None
                        time.sleep(0.5)
                        
                        print(f"   📝 Điền số điện thoại: {process_code}")
                        print(f"   🔄 Tiến trình: {cbil} - Bước 1/4: Điền số điện thoại")
                        phone_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:phoneNumber")))
                        phone_input.clear()
                        phone_input.send_keys(process_code)
                        
                        # Nếu là nạp trả trước, nhập số tiền
                        if is_prepaid:  # Nạp trả trước
                            print(f"   🔄 Tiến trình: {cbil} - Bước 2/4: Điền số tiền")
                            try:
                                print(f"   💰 Điền số tiền: {amount:,}đ")
                                amount_input = WebDriverWait(driver, 10).until(
                                    EC.presence_of_element_located((By.ID, "payMoneyForm:amount"))
                                )
                                amount_input.clear()
                                amount_input.send_keys(str(amount))
                                time.sleep(1)
                            except:
                                # Nếu không tìm thấy input số tiền, thử tìm element khác
                                amount_input = WebDriverWait(driver, 10).until(
                                    EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='number'], input[name*='amount'], .amount-input"))
                                )
                                amount_input.clear()
                                amount_input.send_keys(str(amount))
                                time.sleep(1)
                        
                        # Tự động điền mã PIN từ config
                        print(f"   🔄 Tiến trình: {cbil} - Bước 3/4: Điền mã PIN")
                        try:
                            print(f"   🔐 Điền mã PIN: {Config.DEFAULT_PIN}")
                            pin_input = WebDriverWait(driver, 10).until(
                                EC.presence_of_element_located((By.ID, "payMoneyForm:pin"))
                            )
                            pin_input.clear()
                            pin_input.send_keys(Config.DEFAULT_PIN)
                            time.sleep(1)
                        except:
                            # Nếu không tìm thấy input PIN theo ID, thử tìm element khác
                            try:
                                pin_input = WebDriverWait(driver, 10).until(
                                    EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='password'], input[name*='pin'], .pin-input, input[placeholder*='PIN'], input[placeholder*='pin']"))
                                )
                                pin_input.clear()
                                pin_input.send_keys(Config.DEFAULT_PIN)
                                time.sleep(1)
                                print(f"   🔐 Điền mã PIN thành công (fallback): {Config.DEFAULT_PIN}")
                            except Exception as pin_error:
                                print(f"   ⚠️ Không thể tìm thấy input PIN: {pin_error}")
                        
                        print(f"   🔄 Tiến trình: {cbil} - Bước 4/4: Xử lý giao dịch")
                        print(f"   🔍 Nhấn nút TIẾP TỤC...")
                        continue_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "payMoneyForm:btnContinue")))
                        continue_button.click()
                        time.sleep(1)
                        
                        print(f"   ⏳ Chờ modal loading...")
                        WebDriverWait(driver, 16).until(EC.invisibility_of_element_located((By.ID, "payMoneyForm:j_idt6_modal")))
                        
                        # Kiểm tra thông báo lỗi
                        error_text = get_error_alert_text()
                        if error_text:
                            print(f"   ❌ Có thông báo lỗi: {error_text}")
                            results.append({"code": cbil, "amount": None, "status": "failed", "message": error_text})
                            
                            # Update database cho trường hợp thất bại
                            if order_id:
                                print(f"   💾 Update database cho trường hợp thất bại...")
                                # Lưu thông tin loại dịch vụ vào notes
                                if is_prepaid:
                                    notes = f"Multi-network: Nạp trả trước - {cbil} | Lỗi: {error_text}"
                                else:
                                    notes = f"Multi-network: Gạch nợ trả sau - {cbil} | Lỗi: {error_text}"
                                
                                db_success = update_database_immediately(order_id, process_code, "failed", None, notes, None)
                                if not db_success:
                                    logger.warning(f"Database update thất bại cho {process_code}")
                            break
                        
                        # Lấy thông tin kết quả từ trang
                        print(f"   📊 Lấy thông tin kết quả...")
                        try:
                            # Tìm element chứa thông tin kết quả
                            result_element = WebDriverWait(driver, 10).until(
                                EC.presence_of_element_located((By.CSS_SELECTOR, ".result-info, .payment-result, .success-message, [class*='result'], [class*='success']"))
                            )
                            result_text = result_element.text.strip()
                            print(f"   📋 Kết quả: {result_text}")
                            
                            # Phân tích kết quả để tạo notes chi tiết
                            if "thành công" in result_text.lower() or "success" in result_text.lower():
                                result_status = "success"
                                if is_prepaid:
                                    notes = f"Multi-network: Nạp trả trước - {cbil} | Số tiền: {amount:,}đ | Kết quả: {result_text}"
                                else:
                                    notes = f"Multi-network: Gạch nợ trả sau - {cbil} | Kết quả: {result_text}"
                            else:
                                result_status = "failed"
                                if is_prepaid:
                                    notes = f"Multi-network: Nạp trả trước - {cbil} | Số tiền: {amount:,}đ | Kết quả: {result_text}"
                                else:
                                    notes = f"Multi-network: Gạch nợ trả sau - {cbil} | Kết quả: {result_text}"
                                
                        except Exception as result_error:
                            print(f"   ⚠️ Không thể lấy thông tin kết quả: {result_error}")
                            result_status = "success"
                            if is_prepaid:
                                notes = f"Multi-network: Nạp trả trước - {cbil} | Số tiền: {amount:,}đ"
                            else:
                                notes = f"Multi-network: Gạch nợ trả sau - {cbil}"
                        
                        print(f"   ✅ Xử lý thành công cho {'nạp trả trước' if is_prepaid else 'gạch nợ trả sau'} {process_code}")
                        
                        # Hiển thị kết quả chi tiết tương tự FTTH
                        if 'result_text' in locals():
                            print(f"   📋 Kết quả chi tiết:")
                            print(f"      • Mã: {cbil}")
                            print(f"      • Loại dịch vụ: {'Nạp trả trước' if is_prepaid else 'Gạch nợ trả sau'}")
                            if is_prepaid:
                                print(f"      • Số tiền: {amount:,}đ")
                            print(f"      • Kết quả: {result_text}")
                            print(f"      • Trạng thái: {result_status}")
                        
                        # Thành công, thoát khỏi retry loop
                        success = True
                        results.append({"code": cbil, "amount": amount if is_prepaid else None, "status": result_status, "message": result_text if 'result_text' in locals() else "Thành công"})
                        
                        # Update database ngay lập tức cho từng đơn xử lý xong
                        if order_id:
                            print(f"   💾 Bắt đầu update database...")
                            amount_for_db = amount if is_prepaid else None
                            db_success = update_database_immediately(order_id, process_code, result_status, amount_for_db, notes, None)
                            if not db_success:
                                logger.warning(f"Database update thất bại cho {process_code}")
                        else:
                            print(f"⚠️  [WARNING] Không có order_id, bỏ qua database update cho {process_code}")
                        
                        print(f"   🎉 Hoàn thành xử lý mã {cbil}")
                        break
                        
                    except Exception as e:
                        print(f"   ❌ Lần thử {attempt + 1} thất bại: {e}")
                        logger.warning(f"Lần thử {attempt + 1} thất bại cho {cbil}: {e}")
                        if attempt < 2:  # Còn cơ hội retry
                            print(f"   ⏳ Chờ 1s trước khi retry...")
                            time.sleep(1)  # Delay thông minh
                            continue
                        else:  # Hết retry
                            print(f"   💥 Hết retry, mã {cbil} thất bại hoàn toàn")
                            logger.error(f"Topup đa mạng code {cbil} thất bại sau 3 lần thử: {e}")
                            results.append({"code": cbil, "amount": None, "status": "failed", "message": str(e)})
                            
                            # Update database cho trường hợp thất bại
                            if order_id:
                                print(f"   💾 Update database cho trường hợp thất bại...")
                                # Lưu thông tin loại dịch vụ vào notes
                                if is_prepaid:
                                    notes = f"Multi-network: Nạp trả trước - {cbil} | Lỗi: {str(e)}"
                                else:
                                    notes = f"Multi-network: Gạch nợ trả sau - {cbil} | Lỗi: {str(e)}"
                                
                                db_success = update_database_immediately(order_id, process_code, "failed", None, notes, None)
                                if not db_success:
                                    logger.warning(f"Database update thất bại cho {process_code}")
                            else:
                                print(f"⚠️  [WARNING] Không có order_id, bỏ qua database update cho {process_code}")
                
                if not success:
                    print(f"   💥 Mã {cbil} không thể xử lý sau 3 lần thử")
                    logger.error(f"Mã {cbil} không thể xử lý sau 3 lần thử")
            
            print(f"\n📊 [KẾT QUẢ] Tổng kết xử lý Topup đa mạng:")
            print(f"   ✅ Thành công: {len([r for r in results if r['status'] == 'success'])} mã")
            print(f"   ❌ Thất bại: {len([r for r in results if r['status'] == 'failed'])} mã")
            print(f"   📋 Tổng cộng: {len(results)} mã")
            
            # Hiển thị chi tiết từng mã tương tự FTTH
            print(f"\n📋 [CHI TIẾT] Kết quả xử lý từng mã:")
            for result in results:
                status_icon = "✅" if result['status'] == 'success' else "❌"
                amount_info = f" | Số tiền: {result['amount']:,}đ" if result.get('amount') else ""
                message_info = f" | {result.get('message', '')}" if result.get('message') else ""
                print(f"   {status_icon} {result['code']}{amount_info}{message_info}")
            
            logger.info(f"Topup multinetwork processed: {len(results)} items")
    except Exception as e:
        print(f"   💥 Lỗi tổng thể: {e}")
        logger.error(f"process_topup_multinetwork_codes error: {e}")


def get_data_multi_network(text_widget, pin_widget, form_widget, amount_widget, payment_type: str = None):
    """Lấy dữ liệu API cho Nạp tiền đa mạng"""
    try:
        print(f"[DEBUG] get_data_multi_network được gọi với payment_type: {payment_type}")
        
        # Lấy dữ liệu theo loại dịch vụ đã chọn
        if payment_type:
            print(f"[DEBUG] Gọi db_fetch_service_data với payment_type: {payment_type}")
            data = db_fetch_service_data("nap_tien_da_mang", payment_type)
        else:
            print(f"[DEBUG] Gọi fetch_api_data (không có payment_type)")
            data = fetch_api_data("nap_tien_da_mang")
            
        if data:
            if "phone_numbers" in data:
                populate_text_widget(text_widget, data["phone_numbers"])
            
            # Tự động điền mã PIN từ config
            print(f"[DEBUG] Tự động điền mã PIN: {Config.DEFAULT_PIN}")
            populate_entry_widget(pin_widget, Config.DEFAULT_PIN)
            
            if "payment_type" in data:
                populate_combobox_widget(form_widget, data["payment_type"])
            if "amount" in data:
                populate_combobox_widget(amount_widget, data["amount"])
            
            count = len(data.get("phone_numbers", []))
            order_id = data.get("order_id")
            service_type_text = "Nạp trả trước (sđt|số tiền)" if payment_type == "prepaid" else "Gạch nợ trả sau (chỉ số điện thoại)" if payment_type == "postpaid" else "Đa mạng"
            info_msg = f"Đã tải {count} số điện thoại {service_type_text}"
            if order_id:
                print(f"[INFO] Order ID từ DB: {order_id}", flush=True)
                logger.info(f"Order ID từ DB (Đa mạng - {service_type_text}): {order_id}")
                info_msg += f"\nOrder ID: {order_id}"
            #messagebox.showinfo(Config.TITLE, info_msg)
        else:
            #messagebox.showwarning(Config.TITLE, "Không có dữ liệu từ DB")
    except Exception as e:
        logger.error(f"Lỗi lấy dữ liệu đa mạng: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi lấy dữ liệu đa mạng: {e}")


def navigate_to_topup_multinetwork_page():
    """Điều hướng đến trang nạp tiền đa mạng."""
    try:
        
        target_url = "https://kpp.bankplus.vn/pages/chargecard.jsf"
        driver.get(target_url)
        # Chờ input số điện thoại xuất hiện
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:phoneNumberId")))
        time.sleep(2)
    except Exception as e:
        logger.warning(f"Không thể điều hướng Topup đa mạng: {e}")
        raise

def toggle_input_amount(select, label, combobox):
    selected_value = select.get()
    if selected_value == "Gạch nợ trả sau":
        combobox.pack_forget()
        label.pack_forget()
    else:
        combobox.pack(side="right")
        label.pack(side="right")
    root.update()

def handle_choose_select(choose: str) -> int:
    """Xử lý lựa chọn loại thanh toán"""
    try:
        choose = choose.strip()
        if choose == "Nạp trả trước":
            return 1
        else:
            return 2
    except Exception as e:
        logger.error(f"Lỗi xử lý loại thanh toán: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi xử lý loại thanh toán: {e}")
        return 1


def payment_phone(tkinp_ctm, tkinp_ctmed, tkinp_pin, tkcbb_form, tkcbb_amount):
    try:
        delete_ctmed(tkinp_ctmed)
        update_stop_flag()
        cbils = tkinp_ctm.get("1.0", "end-1c").splitlines()
        pin = tkinp_pin.get()
        cbb_type = tkcbb_form.get()
        type_sub = handle_choose_select(cbb_type)
        if type_sub == 1:
            amount = tkcbb_amount.get()
            isnext = valid_data([cbils, pin, amount])
            if isnext:
                rsl_amount = handle_choose_amount(amount)
        else:
            isnext = valid_data([cbils, pin])
        if not isnext:
            return False
        data = []
        for cbil in cbils:
            cbil = cbil.strip()
            root.update()
            time.sleep(1)
            if not stop_flag and cbil.strip() != "":
                driver.refresh()
                time.sleep(2)
                navigate_to_topup_multinetwork_page()
                try:
                    phonenum = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:phoneNumberId")))
                    phonenum.clear()
                    phonenum.send_keys(cbil)
                    phonenum.send_keys(Keys.TAB)
                    time.sleep(1)
                except:
                    time.sleep(2)
                    phonenum = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:phoneNumberId")))
                    phonenum.clear()
                    phonenum.send_keys(cbil)
                    phonenum.send_keys(Keys.TAB)
                time.sleep(0.5)
                if type_sub == 1:
                    try:
                        try:
                            cfm_modalTT = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, "indexForm:dlgConfirmTT_modal")))
                            driver.execute_script("arguments[0].style.zIndex = '-99';", cfm_modalTT)
                            time.sleep(1)
                        except:
                            pass
                        spl_lbl = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:supplier")))
                        spl_lbl.click()
                        spl_0 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:supplier_0")))
                        spl_0.click()
                        cfm_pay = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:yesTTId")))
                        cfm_pay.click()
                    except:
                        pass
                    script = f"""
                    var element = document.querySelector('input[id="indexForm:subAmountId:{rsl_amount}"]').closest('div');
                    if (!element.classList.contains('ui-state-active')) {{
                        element.click();
                    }}
                    """
                    driver.execute_script(script)
                else:
                    try:
                        try:
                            cfm_modalTT = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.ID, "indexForm:dlgConfirmTT_modal")))
                            driver.execute_script("arguments[0].style.zIndex = '-99';", cfm_modalTT)
                        except:
                            pass
                        try:
                            time.sleep(0.5)
                            spl_lbl = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:supplier")))
                            spl_lbl.click()
                            time.sleep(0.5)
                            spl_1 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:supplier_1")))
                            spl_1.click()
                            time.sleep(0.5)
                            cfm_pay = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:yesTTId")))
                            cfm_pay.click()
                            time.sleep(0.5)
                        except:
                            pass
                        try:
                            btn_check = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:btnCheck")))
                            btn_check.click()
                        except:
                            pass
                        lbl_debt = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, "indexForm:debtId_input")))
                        debt_str = lbl_debt.get_attribute('value')
                        debt = int(debt_str.replace(".", "").replace(",", ""))
                        if debt >= 5000:
                            inp_amount = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:transAmountId_input")))
                            inp_amount.clear()
                            inp_amount.send_keys(debt)
                        else:
                            data.append([cbil, debt, Config.STATUS_COMPLETE])
                            tkinp_ctm.delete("1.0", "1.end+1c")
                            insert_ctmed(tkinp_ctmed, f"{cbil} - {debt}")
                            continue
                    except:
                        data.append([cbil, 0, Config.STATUS_COMPLETE])
                        tkinp_ctm.delete("1.0", "1.end+1c")
                        insert_ctmed(tkinp_ctmed, f"{cbil} - không nợ cước")
                        continue
                try:
                    pin_id = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:pinId")))
                    pin_id.clear()
                    pin_id.send_keys(pin)
                    btn_pay = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:btnPay")))
                    btn_pay.click()
                    try:
                        cfm_modal = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, "indexForm:dlgConfirm_modal")))
                        driver.execute_script("arguments[0].style.zIndex = '-99';", cfm_modal)
                    except:
                        pass
                    time.sleep(0.5)
                    btn_confirm = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:yesIdCard")))
                    btn_confirm.click()
                    if type_sub == 1:
                        data.append([cbil, amount, Config.STATUS_COMPLETE])
                        tkinp_ctm.delete("1.0", "1.end+1c")
                        insert_ctmed(tkinp_ctmed, f"{cbil} - {amount}")
                    else:
                        data.append([cbil, debt, Config.STATUS_COMPLETE])
                        tkinp_ctm.delete("1.0", "1.end+1c")
                        insert_ctmed(tkinp_ctmed, f"{cbil} - {debt}")
                except:
                    data.append([cbil, 0, Config.STATUS_INCOMPLETE])
                    tkinp_ctm.delete("1.0", "1.end+1c")
                    insert_ctmed(tkinp_ctmed, f"{cbil} - {debt}")
        time.sleep(2)
        if len(data) > 0:
            name_dir = "Nạp tiền đa mạng"
            export_excel(data, name_dir)
    except Exception as e:
        logger.error(f"Lỗi thanh toán điện thoại: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi thanh toán điện thoại: {e}")

def form_payment_phone():
    cus_frm = tk.Frame(root)
    cus_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    ctm_frm = tk.Frame(cus_frm)
    ctm_frm.pack(expand=True, side="left")
    ctmed_frm = tk.Frame(cus_frm)
    ctmed_frm.pack(expand=True, side="right")
    form_frm = tk.Frame(root)
    form_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    pin_frm = tk.Frame(root)
    pin_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    btn_frm = tk.Frame(root)
    btn_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    tklbl_ctm = tk.Label(ctm_frm, text="Số điện thoại")
    tklbl_ctm.pack(side="top")
    tkinp_ctm = tk.Text(ctm_frm, height=12, width=24)
    tkinp_ctm.pack(side="left", pady=8)
    tklbl_ctm = tk.Label(ctmed_frm, text="Đã xử lý")
    tklbl_ctm.pack(side="top")
    tkinp_ctmed = tk.Text(ctmed_frm, width=32, height=12, bg="#ccc", state="disabled")
    tkinp_ctmed.pack(side="left", pady=8)
    tklbl_form = tk.Label(form_frm, text="Hình thức:")
    tklbl_form.pack(side="left")
    tkcbb_form = ttk.Combobox(form_frm, values=["Nạp trả trước", "Gạch nợ trả sau"], width="14", state="readonly")
    tkcbb_form.pack(side="left")
    tkcbb_form.set("Nạp trả trước")
    tkcbb_form.bind("<<ComboboxSelected>>", lambda event: toggle_input_amount(tkcbb_form, tklbl_amount, tkcbb_amount))
    tkcbb_amount = ttk.Combobox(form_frm, values=["10.000đ", "20.000đ", "30.000đ", "50.000đ", "100.000đ", "200.000đ", "300.000đ", "500.000đ"], width="10", state="readonly")
    tkcbb_amount.pack(side="right")
    tklbl_amount = tk.Label(form_frm, text="Số tiền nạp:")
    tklbl_amount.pack(side="right")
    tklbl_pin = tk.Label(pin_frm, text="Mã pin:")
    tklbl_pin.pack(side="left")
    tkinp_pin = ttk.Entry(pin_frm, width=12)
    tkinp_pin.pack(side="left", padx=4)
    style = ttk.Style()
    style.configure("Red.TButton", foreground="red")
    style.configure("Blue.TButton", foreground="blue")
    style.configure("Green.TButton", foreground="green")
    def get_data_with_payment_type():
        selected = tkcbb_form.get()
        print(f"[DEBUG] Combobox được chọn: '{selected}'")
        
        if selected == "Nạp trả trước":
            payment_type = "prepaid"
        elif selected == "Gạch nợ trả sau":
            payment_type = "postpaid"
        else:
            payment_type = None
            
        print(f"[DEBUG] Payment type được map: {payment_type}")
        get_data_multi_network(tkinp_ctm, tkinp_pin, tkcbb_form, tkcbb_amount, payment_type)
    
    tkbtn_get_data = ttk.Button(btn_frm, text="Get dữ liệu", command=get_data_with_payment_type)
    tkbtn_get_data.pack(side='left', padx=5, pady=5)
    tkbtn_get_data.configure(style="Green.TButton")
    tkbtn_payment = ttk.Button(btn_frm, text="Bắt đầu", command=lambda: payment_phone(tkinp_ctm, tkinp_ctmed, tkinp_pin, tkcbb_form, tkcbb_amount))
    tkbtn_payment.pack(side='left', padx=5, pady=5)
    tkbtn_payment.configure(style="Blue.TButton") 
    tkbtn_destroy = ttk.Button(btn_frm, text="Dừng lại", command=stop_tool)
    tkbtn_destroy.pack(side='right', padx=5, pady=5)
    tkbtn_destroy.configure(style="Red.TButton") 


def process_topup_viettel_codes(codes: List[str], order_id: Optional[str] = None):
    """Xử lý nạp tiền Viettel không cần GUI, điều khiển selenium trực tiếp."""
    print(f"🚀 [AUTOMATION] Bắt đầu xử lý Topup Viettel cho {len(codes)} mã")
    print(f"   📋 Order ID: {order_id or 'Không có'}")
    
    if not ensure_driver_and_login():
        print("   ❌ Không thể khởi tạo driver hoặc đăng nhập")
        logger.error("Không thể khởi tạo driver hoặc đăng nhập")
        return
    try:
        with automation_lock:
            print("   🔒 Đã khóa automation, điều hướng đến trang Topup Viettel...")
            navigate_to_topup_viettel_page()
            print("   ✅ Đã điều hướng thành công đến trang Topup Viettel")
            
            results = []
            for idx, cbil in enumerate(codes, 1):
                print(f"\n📱 [MÃ {idx}/{len(codes)}] Xử lý mã: {cbil}")
                success = False
                for attempt in range(3):  # Retry tối đa 3 lần
                    try:
                        cbil = (cbil or "").strip()
                        if not cbil:
                            print(f"   ⚠️  Mã rỗng, bỏ qua")
                            break
                        
                        if attempt > 0:
                            print(f"   🔄 Retry lần {attempt + 1}/3 cho mã {cbil}")
                            logger.info(f"Retry lần {attempt + 1} cho mã {cbil}")
                            driver.refresh()
                            time.sleep(2)
                            navigate_to_topup_viettel_page()
                        else:
                            print(f"   🎯 Lần thử đầu tiên cho mã {cbil}")
                        
                        root.update() if 'root' in globals() else None
                        time.sleep(0.5)
                        
                        print(f"   📝 Điền số điện thoại: {cbil}")
                        phone_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:phoneNumber")))
                        phone_input.clear()
                        phone_input.send_keys(cbil)
                        
                        print(f"   🔍 Nhấn nút TIẾP TỤC...")
                        continue_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "payMoneyForm:btnContinue")))
                        continue_button.click()
                        time.sleep(1)
                        
                        print(f"   ⏳ Chờ modal loading...")
                        WebDriverWait(driver, 16).until(EC.invisibility_of_element_located((By.ID, "payMoneyForm:j_idt6_modal")))
                        
                        print(f"   ✅ Xử lý thành công cho số điện thoại {cbil}")
                        
                        # Thành công, thoát khỏi retry loop
                        success = True
                        results.append({"code": cbil, "amount": None, "status": "success"})
                        
                        # Update database ngay lập tức cho từng đơn xử lý xong
                        if order_id:
                            print(f"   💾 Bắt đầu update database...")
                            db_success = update_database_immediately(order_id, cbil, "success", None, "Topup Viettel ok", None)
                            if not db_success:
                                logger.warning(f"Database update thất bại cho {cbil}")
                        else:
                            print(f"⚠️  [WARNING] Không có order_id, bỏ qua database update cho {cbil}")
                        
                        print(f"   🎉 Hoàn thành xử lý mã {cbil}")
                        break
                        
                    except Exception as e:
                        print(f"   ❌ Lần thử {attempt + 1} thất bại: {e}")
                        logger.warning(f"Lần thử {attempt + 1} thất bại cho {cbil}: {e}")
                        if attempt < 2:  # Còn cơ hội retry
                            print(f"   ⏳ Chờ 1s trước khi retry...")
                            time.sleep(1)  # Delay thông minh
                            continue
                        else:  # Hết retry
                            print(f"   💥 Hết retry, mã {cbil} thất bại hoàn toàn")
                            logger.error(f"Topup Viettel code {cbil} thất bại sau 3 lần thử: {e}")
                            results.append({"code": cbil, "amount": None, "status": "failed", "message": str(e)})
                            
                            # Update database cho trường hợp thất bại
                            if order_id:
                                print(f"   💾 Update database cho trường hợp thất bại...")
                                db_success = update_database_immediately(order_id, cbil, "failed", None, str(e), None)
                                if not db_success:
                                    logger.warning(f"Database update thất bại cho {cbil}")
                            else:
                                print(f"⚠️  [WARNING] Không có order_id, bỏ qua database update cho {cbil}")
                
                if not success:
                    print(f"   💥 Mã {cbil} không thể xử lý sau 3 lần thử")
                    logger.error(f"Mã {cbil} không thể xử lý sau 3 lần thử")
            
            print(f"\n📊 [KẾT QUẢ] Tổng kết xử lý Topup Viettel:")
            print(f"   ✅ Thành công: {len([r for r in results if r['status'] == 'success'])} mã")
            print(f"   ❌ Thất bại: {len([r for r in results if r['status'] == 'failed'])} mã")
            print(f"   📋 Tổng cộng: {len(results)} mã")
            
            logger.info(f"Topup Viettel processed: {len(results)} items")
    except Exception as e:
        print(f"   💥 Lỗi tổng thể: {e}")
        logger.error(f"process_topup_viettel_codes error: {e}")


def get_data_viettel(text_widget, pin_widget, amount_widget):
    """Lấy dữ liệu API cho Nạp tiền mạng Viettel"""
    try:
        data = fetch_api_data("nap_tien_viettel")
        if data:
            if "phone_numbers" in data:
                populate_text_widget(text_widget, data["phone_numbers"])
            
            # Tự động điền mã PIN từ config
            print(f"[DEBUG] Tự động điền mã PIN: {Config.DEFAULT_PIN}")
            populate_entry_widget(pin_widget, Config.DEFAULT_PIN)
            
            if "amount" in data:
                populate_entry_widget(amount_widget, data["amount"])
            
            count = len(data.get("phone_numbers", []))
            order_id = data.get("order_id")
            info_msg = f"Đã tải {count} số điện thoại Viettel"
            if order_id:
                print(f"[INFO] Order ID từ DB: {order_id}", flush=True)
                logger.info(f"Order ID từ DB (Viettel): {order_id}")
                info_msg += f"\nOrder ID: {order_id}"
            #messagebox.showinfo(Config.TITLE, info_msg)
        else:
            #messagebox.showwarning(Config.TITLE, "Không có dữ liệu từ DB")
    except Exception as e:
        logger.error(f"Lỗi lấy dữ liệu Viettel: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi lấy dữ liệu Viettel: {e}")

def navigate_to_topup_viettel_page():
    """Điều hướng đến trang nạp tiền Viettel."""
    try:
        
        target_url = "https://kpp.bankplus.vn/pages/chargecard.jsf"
        driver.get(target_url)
        # Chờ input số điện thoại xuất hiện
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:phoneNumber")))
        time.sleep(2)
    except Exception as e:
        logger.warning(f"Không thể điều hướng Topup Viettel: {e}")
        raise


def payment_viettel(tkinp_ctm, tkinp_ctmed, tkinp_pin, tkinp_amount):
    try:
        delete_ctmed(tkinp_ctmed)
        update_stop_flag()
        pin = tkinp_pin.get()
        amount = tkinp_amount.get()
        cbils = tkinp_ctm.get("1.0", "end-1c").splitlines()
        if not valid_data([cbils, pin, amount]):
            return False
        data = []
        for cbil in cbils:
            root.update()
            time.sleep(0.5)
            cbil = cbil.strip()
            if not stop_flag and cbil.strip() != "":
                time.sleep(0.5)
                try:
                    phonenum = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:phoneNumberId")))
                    phonenum.clear()
                    phonenum.send_keys(cbil)
                    phonenum.send_keys(Keys.TAB)
                    time.sleep(0.5)
                except:
                    time.sleep(2)
                    phonenum = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:phoneNumberId")))
                    phonenum.clear()
                    phonenum.send_keys(cbil)
                    phonenum.send_keys(Keys.TAB)
                try:
                    cfm_modalTT = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.ID, "indexForm:dlgConfirmTT_modal")))
                    driver.execute_script("arguments[0].style.zIndex = '-99';", cfm_modalTT)
                    time.sleep(0.5)
                    spl_lbl = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.ID, "indexForm:supplier")))
                    spl_lbl.click()
                    time.sleep(0.5)
                    spl_0 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:supplier_0")))
                    spl_0.click()
                    time.sleep(0.5)
                    cfm_pay = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:yesTTId")))
                    cfm_pay.click()
                    time.sleep(0.5)
                except:
                    pass
                try:
                    inp_amount = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:transAmountId_input")))
                    inp_amount.clear()
                    inp_amount.send_keys(amount)
                    time.sleep(0.5)
                    pin_id = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:pinId")))
                    pin_id.clear()
                    pin_id.send_keys(pin)
                    time.sleep(0.5)
                    btn_pay = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:btnPay")))
                    btn_pay.click()
                    time.sleep(0.5)
                    try:
                        cfm_modal = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, "indexForm:dlgConfirm_modal")))
                        driver.execute_script("arguments[0].style.zIndex = '-99';", cfm_modal)
                    except:
                        pass
                    btn_confirm = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:yesIdCard")))
                    btn_confirm.click()
                    data.append([cbil, amount, Config.STATUS_COMPLETE])
                    insert_ctmed(tkinp_ctmed, f"{cbil} - {amount}")
                    tkinp_ctm.delete("1.0", "1.end+1c")
                except:
                    data.append([cbil, 0, Config.STATUS_INCOMPLETE])
                    insert_ctmed(tkinp_ctmed, f"{cbil} - null")
                    tkinp_ctm.delete("1.0", "1.end+1c")
                    continue
        time.sleep(2)
        if len(data) > 0:
            name_dir = "Nạp tiền mạng Viettel"
            export_excel(data, name_dir)
    except Exception as e:
        logger.error(f"Lỗi thanh toán Viettel: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi thanh toán Viettel: {e}")

def form_payment_viettel():
    cus_frm = tk.Frame(root)
    cus_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    ctm_frm = tk.Frame(cus_frm)
    ctm_frm.pack(expand=True, side="left")
    ctmed_frm = tk.Frame(cus_frm)
    ctmed_frm.pack(expand=True, side="right")
    form_frm = tk.Frame(root)
    form_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    pin_frm = tk.Frame(root)
    pin_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    button_frm = tk.Frame(root)
    button_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    tklbl_ctm = tk.Label(ctm_frm, text="Số điện thoại")
    tklbl_ctm.pack(side="top")
    tkinp_ctm = tk.Text(ctm_frm, height=12, width=24)
    tkinp_ctm.pack(side="left", pady=8)
    tklbl_ctm = tk.Label(ctmed_frm, text="Đã xử lý")
    tklbl_ctm.pack(side="top")
    tkinp_ctmed = tk.Text(ctmed_frm, width=32, height=12, bg="#ccc", state="disabled")
    tkinp_ctmed.pack(side="left", pady=8)
    tkinp_amount = tk.Entry(pin_frm, width=12)
    tkinp_amount.pack(side="right", padx=4)
    tklbl_amount = tk.Label(pin_frm, text="Số tiền nạp:")
    tklbl_amount.pack(side="right")
    tklbl_pin = tk.Label(pin_frm, text="Mã pin:")
    tklbl_pin.pack(side="left")
    tkinp_pin = ttk.Entry(pin_frm, width=12)
    tkinp_pin.pack(side="left", padx=4)
    style = ttk.Style()
    style.configure("Red.TButton", foreground="red")
    style.configure("Blue.TButton", foreground="blue")
    style.configure("Green.TButton", foreground="green")
    tkbtn_get_data = ttk.Button(button_frm, text="Get dữ liệu", command=lambda: get_data_viettel(tkinp_ctm, tkinp_pin, tkinp_amount))
    tkbtn_get_data.pack(side='left', padx=5, pady=5)
    tkbtn_get_data.configure(style="Green.TButton")
    tkbtn_payment = ttk.Button(button_frm, text="Bắt đầu", command=lambda: payment_viettel(tkinp_ctm, tkinp_ctmed, tkinp_pin, tkinp_amount))
    tkbtn_payment.pack(side='left', padx=5, pady=5)
    tkbtn_payment.configure(style="Blue.TButton") 
    tkbtn_destroy = ttk.Button(button_frm, text="Dừng lại", command=stop_tool)
    tkbtn_destroy.pack(side='right', padx=5, pady=5)
    tkbtn_destroy.configure(style="Red.TButton") 


def process_tv_internet_codes(codes: List[str], order_id: Optional[str] = None):
    """Xử lý thanh toán TV-Internet không cần GUI, điều khiển selenium trực tiếp."""
    print(f"🚀 [AUTOMATION] Bắt đầu xử lý TV-Internet cho {len(codes)} mã")
    print(f"   📋 Order ID: {order_id or 'Không có'}")
    
    if not ensure_driver_and_login():
        print("   ❌ Không thể khởi tạo driver hoặc đăng nhập")
        logger.error("Không thể khởi tạo driver hoặc đăng nhập")
        return
    
    try:
        with automation_lock:
            print("   🔒 Đã khóa automation, điều hướng đến trang TV-Internet...")
            navigate_to_tv_internet_page()
            print("   ✅ Đã điều hướng thành công đến trang TV-Internet")
            
            results = []
            for idx, cbil in enumerate(codes, 1):
                print(f"\n📱 [MÃ {idx}/{len(codes)}] Xử lý mã: {cbil}")
                success = False
                for attempt in range(3):  # Retry tối đa 3 lần
                    try:
                        cbil = (cbil or "").strip()
                        if not cbil:
                            print(f"   ⚠️  Mã rỗng, bỏ qua")
                            break
                        
                        if attempt > 0:
                            print(f"   🔄 Retry lần {attempt + 1}/3 cho mã {cbil}")
                            logger.info(f"Retry lần {attempt + 1} cho mã {cbil}")
                            driver.refresh()
                            time.sleep(2)
                            navigate_to_tv_internet_page()
                        else:
                            print(f"   🎯 Lần thử đầu tiên cho mã {cbil}")
                        
                        root.update() if 'root' in globals() else None
                        time.sleep(0.5)
                        
                        print(f"   📝 Điền mã thuê bao: {cbil}")
                        customer = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:contractCode")))
                        customer.clear()
                        customer.send_keys(cbil)
                        
                        print(f"   🔍 Nhấn nút KIỂM TRA...")
                        payment_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:btnPay0")))
                        payment_button.click()
                        time.sleep(1)
                        
                        print(f"   ⏳ Chờ modal loading...")
                        WebDriverWait(driver, 16).until(EC.invisibility_of_element_located((By.ID, "payMoneyForm:j_idt6_modal")))
                        
                        print(f"   📊 Lấy thông tin kết quả...")
                        element41 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:j_idt41")))
                        _is_amount, amount, _pid = amount_by_cbil(cbil, element41, False)
                        
                        print(f"   ✅ Xử lý thành công: Amount = {amount}")
                        
                        # Thành công, thoát khỏi retry loop
                        success = True
                        results.append({"code": cbil, "amount": amount, "status": "success"})
                        
                        # Update database ngay lập tức cho từng đơn xử lý xong
                        if order_id:
                            print(f"   💾 Bắt đầu update database...")
                            db_success = update_database_immediately(order_id, cbil, "success", amount, "TV-Internet payment ok", None)
                            if not db_success:
                                logger.warning(f"Database update thất bại cho {cbil}")
                        else:
                            print(f"⚠️  [WARNING] Không có order_id, bỏ qua database update cho {cbil}")
                        
                        print(f"   🎉 Hoàn thành xử lý mã {cbil}")
                        break
                        
                    except Exception as e:
                        print(f"   ❌ Lần thử {attempt + 1} thất bại: {e}")
                        logger.warning(f"Lần thử {attempt + 1} thất bại cho {cbil}: {e}")
                        if attempt < 2:  # Còn cơ hội retry
                            print(f"   ⏳ Chờ 1s trước khi retry...")
                            time.sleep(1)  # Delay thông minh
                            continue
                        else:  # Hết retry
                            print(f"   💥 Hết retry, mã {cbil} thất bại hoàn toàn")
                            logger.error(f"TV-Internet code {cbil} thất bại sau 3 lần thử: {e}")
                            results.append({"code": cbil, "amount": None, "status": "failed", "message": str(e)})
                            
                            # Update database cho trường hợp thất bại
                            if order_id:
                                print(f"   💾 Update database cho trường hợp thất bại...")
                                db_success = update_database_immediately(order_id, cbil, "failed", None, str(e), None)
                                if not db_success:
                                    logger.warning(f"Database update thất bại cho {cbil}")
                            else:
                                print(f"⚠️  [WARNING] Không có order_id, bỏ qua database update cho {cbil}")
                
                if not success:
                    print(f"   💥 Mã {cbil} không thể xử lý sau 3 lần thử")
                    logger.error(f"Mã {cbil} không thể xử lý sau 3 lần thử")
            
            print(f"\n📊 [KẾT QUẢ] Tổng kết xử lý TV-Internet:")
            print(f"   ✅ Thành công: {len([r for r in results if r['status'] == 'success'])} mã")
            print(f"   ❌ Thất bại: {len([r for r in results if r['status'] == 'failed'])} mã")
            print(f"   📋 Tổng cộng: {len(results)} mã")
            
            logger.info(f"TV-Internet processed: {len(results)} items")
    except Exception as e:
        print(f"   💥 Lỗi tổng thể: {e}")
        logger.error(f"process_tv_internet_codes error: {e}")


def get_data_tv_internet(text_widget, pin_widget):
    """Lấy dữ liệu API cho Thanh toán TV - Internet"""
    try:
        data = fetch_api_data("thanh_toan_tv_internet")
        if data:
            if "subscriber_codes" in data:
                populate_text_widget(text_widget, data["subscriber_codes"])
            
            # Tự động điền mã PIN từ config
            print(f"[DEBUG] Tự động điền mã PIN: {Config.DEFAULT_PIN}")
            populate_entry_widget(pin_widget, Config.DEFAULT_PIN)
            
            count = len(data.get("subscriber_codes", []))
            order_id = data.get("order_id")
            info_msg = f"Đã tải {count} mã thuê bao TV-Internet"
            if order_id:
                print(f"[INFO] Order ID từ DB: {order_id}", flush=True)
                logger.info(f"Order ID từ DB (TV-Internet): {order_id}")
                info_msg += f"\nOrder ID: {order_id}"
            #messagebox.showinfo(Config.TITLE, info_msg)
        else:
            #messagebox.showwarning(Config.TITLE, "Không có dữ liệu từ DB")
    except Exception as e:
        logger.error(f"Lỗi lấy dữ liệu TV-Internet: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi lấy dữ liệu TV-Internet: {e}")


def navigate_to_tv_internet_page():
    """Điều hướng đến trang thanh toán TV-Internet."""
    try:
        target_url = "https://kpp.bankplus.vn/pages/newInternetTelevisionViettel.jsf?serviceCode=000003&serviceType=INTERNET"
        driver.get(target_url)
        # Chờ input mã thuê bao xuất hiện
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:contractCode")))
        time.sleep(2)
    except Exception as e:
        logger.warning(f"Không thể điều hướng TV-Internet: {e}")
        raise


def payment_internet(tkinp_ctm, tkinp_ctmed, tkinp_pin):
    try:
        delete_ctmed(tkinp_ctmed)
        update_stop_flag()
        pin = tkinp_pin.get()
        cbils = tkinp_ctm.get("1.0", "end-1c").splitlines()
        if not valid_data([cbils, pin]):
            return False
        data = []
        for cbil in cbils:
            root.update()
            time.sleep(0.5)
            cbil = cbil.strip()
            if not stop_flag and cbil.strip() != "":
                try:
                    customer = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:contractCode")))
                    customer.clear()
                    customer.send_keys(cbil)
                    time.sleep(0.5)
                    payment_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:btnPay0")))
                    payment_button.click()
                    time.sleep(1)
                    WebDriverWait(driver, 16).until(EC.invisibility_of_element_located((By.ID, "payMoneyForm:j_idt6_modal")))
                    element41 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:j_idt41")))
                    is_amount, amount, payment_id = amount_by_cbil(cbil, element41, True)
                    if not is_amount:
                        data.append([cbil, amount, Config.STATUS_COMPLETE])
                        insert_ctmed(tkinp_ctmed, f"{cbil} - {amount}")
                        tkinp_ctm.delete("1.0", "1.end+1c")
                        continue
                    else:
                        payment_btn1 = WebDriverWait(driver, 16).until(EC.presence_of_element_located((By.ID, payment_id)))
                        payment_btn1.click()
                        pin_id = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:pinId")))
                        pin_id.clear()
                        pin_id.send_keys(pin)
                        pay_btn = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:btnPay")))
                        pay_btn.click()
                        try:
                            cfm_modal = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.ID, "payMoneyForm:dlgConfirm_modal")))
                            driver.execute_script("arguments[0].style.zIndex = '-99';", cfm_modal)
                        except:
                            pass
                        confirm_btn = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:yesId0")))
                        confirm_btn.click()
                        data.append([cbil, amount, Config.STATUS_COMPLETE])
                        insert_ctmed(tkinp_ctmed, f"{cbil} - {amount}")
                        tkinp_ctm.delete("1.0", "1.end+1c")
                except Exception as e:
                    data.append([cbil, 0, Config.STATUS_INCOMPLETE])
                    insert_ctmed(tkinp_ctmed, f"{cbil} - Lỗi")
                    tkinp_ctm.delete("1.0", "1.end+1c")
        time.sleep(2)
        if len(data) > 0:
            name_dir = "Thanh toán TV - Internet"
            export_excel(data, name_dir)
    except Exception as e:
        logger.error(f"Lỗi thanh toán internet: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi thanh toán internet: {e}")

def form_payment_internet():
    cus_frm = tk.Frame(root)
    cus_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    ctm_frm = tk.Frame(cus_frm)
    ctm_frm.pack(expand=True, side="left")
    ctmed_frm = tk.Frame(cus_frm)
    ctmed_frm.pack(expand=True, side="right")
    phnum_frm = tk.Frame(root)
    phnum_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    btn_frm = tk.Frame(root)
    btn_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    tklbl_ctm = tk.Label(ctm_frm, text="Mã thuê bao")
    tklbl_ctm.pack(side="top")
    tkinp_ctm = tk.Text(ctm_frm, height=12, width=24)
    tkinp_ctm.pack(side="left", pady=8)
    tklbl_ctm = tk.Label(ctmed_frm, text="Đã xử lý")
    tklbl_ctm.pack(side="top")
    tkinp_ctmed = tk.Text(ctmed_frm, width=32, height=12, bg="#ccc", state="disabled")
    tkinp_ctmed.pack(side="left", pady=8)
    tklbl_pin = tk.Label(phnum_frm, text="Mã pin:")
    tklbl_pin.pack(side="left")
    tkinp_pin = ttk.Entry(phnum_frm, width=22)
    tkinp_pin.pack(side="left", padx=4)
    style = ttk.Style()
    style.configure("Red.TButton", foreground="red")
    style.configure("Blue.TButton", foreground="blue")
    style.configure("Green.TButton", foreground="green")
    tkbtn_get_data = ttk.Button(btn_frm, text="Get dữ liệu", command=lambda: get_data_tv_internet(tkinp_ctm, tkinp_pin))
    tkbtn_get_data.pack(side='left', padx=5, pady=5)
    tkbtn_get_data.configure(style="Green.TButton")
    tkbtn_payment = ttk.Button(btn_frm, text="Bắt đầu", command=lambda: payment_internet(tkinp_ctm, tkinp_ctmed, tkinp_pin))
    tkbtn_payment.pack(side='left', padx=5, pady=5)
    tkbtn_payment.configure(style="Blue.TButton") 
    tkbtn_destroy = ttk.Button(btn_frm, text="Dừng lại", command=stop_tool)
    tkbtn_destroy.pack(side='right', padx=5, pady=5)
    tkbtn_destroy.configure(style="Red.TButton") 

def process_postpaid_lookup_codes(codes: List[str], order_id: Optional[str] = None):
    """Xử lý tra cứu trả sau không cần GUI, điều khiển selenium trực tiếp."""
    print(f"🚀 [AUTOMATION] Bắt đầu xử lý Postpaid cho {len(codes)} mã")
    print(f"   📋 Order ID: {order_id or 'Không có'}")
    
    if not ensure_driver_and_login():
        print("   ❌ Không thể khởi tạo driver hoặc đăng nhập")
        logger.error("Không thể khởi tạo driver hoặc đăng nhập")
        return
    
    try:
        with automation_lock:
            print("   🔒 Đã khóa automation, điều hướng đến trang Postpaid...")
            navigate_to_postpaid_lookup_page()
            print("   ✅ Đã điều hướng thành công đến trang Postpaid")
            
            results = []
            for idx, cbil in enumerate(codes, 1):
                print(f"\n📱 [MÃ {idx}/{len(codes)}] Xử lý mã: {cbil}")
                success = False
                for attempt in range(3):  # Retry tối đa 3 lần
                    try:
                        cbil = (cbil or "").strip()
                        if not cbil:
                            print(f"   ⚠️  Mã rỗng, bỏ qua")
                            break
                        
                        if attempt > 0:
                            print(f"   🔄 Retry lần {attempt + 1}/3 cho mã {cbil}")
                            logger.info(f"Retry lần {attempt + 1} cho mã {cbil}")
                            driver.refresh()
                            time.sleep(2)
                            navigate_to_postpaid_lookup_page()
                        else:
                            print(f"   🎯 Lần thử đầu tiên cho mã {cbil}")
                        
                        root.update() if 'root' in globals() else None
                        time.sleep(0.5)
                        
                        print(f"   📝 Điền mã thuê bao: {cbil}")
                        customer = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:contractCode")))
                        customer.clear()
                        customer.send_keys(cbil)
                        
                        print(f"   🔍 Nhấn nút KIỂM TRA...")
                        payment_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:btnPay0")))
                        payment_button.click()
                        time.sleep(1)
                        
                        print(f"   ⏳ Chờ modal loading...")
                        WebDriverWait(driver, 16).until(EC.invisibility_of_element_located((By.ID, "payMoneyForm:j_idt6_modal")))
                        
                        print(f"   📊 Lấy thông tin kết quả...")
                        element41 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:j_idt41")))
                        _is_amount, amount, _pid = amount_by_cbil(cbil, element41, False)
                        
                        print(f"   ✅ Xử lý thành công: Amount = {amount}")
                        
                        # Thành công, thoát khỏi retry loop
                        success = True
                        results.append({"code": cbil, "amount": amount, "status": "success"})
                        
                        # Update database ngay lập tức cho từng đơn xử lý xong
                        if order_id:
                            print(f"   💾 Bắt đầu update database...")
                            db_success = update_database_immediately(order_id, cbil, "success", amount, "Postpaid lookup ok", None)
                            if not db_success:
                                logger.warning(f"Database update thất bại cho {cbil}")
                        else:
                            print(f"⚠️  [WARNING] Không có order_id, bỏ qua database update cho {cbil}")
                        
                        print(f"   🎉 Hoàn thành xử lý mã {cbil}")
                        break
                        
                    except Exception as e:
                        print(f"   ❌ Lần thử {attempt + 1} thất bại: {e}")
                        logger.warning(f"Lần thử {attempt + 1} thất bại cho {cbil}: {e}")
                        if attempt < 2:  # Còn cơ hội retry
                            print(f"   ⏳ Chờ 1s trước khi retry...")
                            time.sleep(1)  # Delay thông minh
                            continue
                        else:  # Hết retry
                            print(f"   💥 Hết retry, mã {cbil} thất bại hoàn toàn")
                            logger.error(f"Postpaid code {cbil} thất bại sau 3 lần thử: {e}")
                            results.append({"code": cbil, "amount": None, "status": "failed", "message": str(e)})
                            
                            # Update database cho trường hợp thất bại
                            if order_id:
                                print(f"   💾 Update database cho trường hợp thất bại...")
                                db_success = update_database_immediately(order_id, cbil, "failed", None, str(e), None)
                                if not db_success:
                                    logger.warning(f"Database update thất bại cho {cbil}")
                            else:
                                print(f"⚠️  [WARNING] Không có order_id, bỏ qua database update cho {cbil}")
                
                if not success:
                    print(f"   💥 Mã {cbil} không thể xử lý sau 3 lần thử")
                    logger.error(f"Mã {cbil} không thể xử lý sau 3 lần thử")
            
            print(f"\n📊 [KẾT QUẢ] Tổng kết xử lý Postpaid:")
            print(f"   ✅ Thành công: {len([r for r in results if r['status'] == 'success'])} mã")
            print(f"   ❌ Thất bại: {len([r for r in results if r['status'] == 'failed'])} mã")
            print(f"   📋 Tổng cộng: {len(results)} mã")
            
            logger.info(f"Postpaid processed: {len(results)} items")
    except Exception as e:
        print(f"   💥 Lỗi tổng thể: {e}")
        logger.error(f"process_postpaid_lookup_codes error: {e}")


def get_data_postpaid(text_widget):
    """Lấy dữ liệu API cho Tra cứu nợ thuê bao trả sau"""
    try:
        data = fetch_api_data("tra_cuu_no_tra_sau")
        if data and "phone_numbers" in data:
            populate_text_widget(text_widget, data["phone_numbers"])
            order_id = data.get("order_id")
            info_msg = f"Đã tải {len(data['phone_numbers'])} số điện thoại trả sau"
            if order_id:
                print(f"[INFO] Order ID từ DB: {order_id}", flush=True)
                logger.info(f"Order ID từ DB (Trả sau): {order_id}")
                info_msg += f"\nOrder ID: {order_id}"
            #messagebox.showinfo(Config.TITLE, info_msg)
        else:
            #messagebox.showwarning(Config.TITLE, "Không có dữ liệu từ API")
    except Exception as e:
        logger.error(f"Lỗi lấy dữ liệu trả sau: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi lấy dữ liệu trả sau: {e}")


def navigate_to_postpaid_lookup_page():
    """Điều hướng đến trang tra cứu trả sau."""
    try:
        target_url = "https://kpp.bankplus.vn/pages/chargecard.jsf"
        driver.get(target_url)
        # Chờ input mã thuê bao xuất hiện
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:phoneNumberId")))
        time.sleep(2)
    except Exception as e:
        logger.warning(f"Không thể điều hướng Postpaid: {e}")
        raise


def lookup_card(tkinp_ctm, tkinp_ctmed):
    try:
        delete_ctmed(tkinp_ctmed)
        update_stop_flag()
        cbils = tkinp_ctm.get("1.0", "end-1c").splitlines()
        if not valid_data([cbils]):
            return False
        data = []
        for cbil in cbils:
            cbil = cbil.strip()
            root.update()
            time.sleep(1)
            if not stop_flag and cbil.strip() != "":
                # Điều hướng đến trang tra cứu trả sau trước khi xử lý
                navigate_to_postpaid_lookup_page()
                time.sleep(2)
                try:
                    phonenum = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:phoneNumberId")))
                    phonenum.clear()
                    phonenum.send_keys(cbil)
                    phonenum.send_keys(Keys.TAB)
                    time.sleep(1)
                except:
                    time.sleep(2)
                    phonenum = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:phoneNumberId")))
                    phonenum.clear()
                    phonenum.send_keys(cbil)
                    phonenum.send_keys(Keys.TAB)
                try:
                    try:
                        cfm_modalTT = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.ID, "indexForm:dlgConfirmTT_modal")))
                        driver.execute_script("arguments[0].style.zIndex = '-99';", cfm_modalTT)
                    except:
                        pass
                    try:
                        time.sleep(0.5)
                        spl_lbl = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:supplier")))
                        spl_lbl.click()
                        time.sleep(0.5)
                        spl_1 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:supplier_1")))
                        spl_1.click()
                        time.sleep(0.5)
                        cfm_pay = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:yesTTId")))
                        cfm_pay.click()
                        time.sleep(0.5)
                    except:
                        pass
                    try:
                        btn_check = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:btnCheck")))
                        btn_check.click()
                    except:
                        pass

                    # Chờ modal/tiến trình, rồi kiểm tra thông báo info "không còn nợ cước"
                    time.sleep(1)
                    info_text = get_info_alert_text()
                    if info_text and ("không còn nợ cước" in info_text.lower()):
                        note_text = info_text.strip()
                        data.append([cbil, 0, Config.STATUS_COMPLETE])
                        insert_ctmed(tkinp_ctmed, f"{cbil} - 0 | {note_text}")
                        # Cập nhật DB success với amount=0 và notes = "code - 0 | alert"
                        order_id_val = db_find_order_id('tra_cuu_no_tra_sau', cbil.strip(), None)
                        if order_id_val:
                            try:
                                notes_db = f"{cbil} - 0 | {note_text}"
                                _ = update_database_immediately(order_id_val, cbil, "success", 0, notes_db, None)
                            except Exception as _e:
                                logger.warning(f"DB update lỗi (Postpaid no debt) cho {cbil}: {_e}")
                        continue

                    # Lấy giá trị nợ cước nếu có
                    lbl_debt = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, "indexForm:debtId_input")))
                    debt_str = lbl_debt.get_attribute('value')
                    debt = int(debt_str.replace(".", "").replace(",", ""))
                    data.append([cbil, debt, Config.STATUS_COMPLETE])
                    insert_ctmed(tkinp_ctmed, f"{cbil} - {debt}")
                    # Cập nhật DB success với amount=debt
                    order_id_val = db_find_order_id('tra_cuu_no_tra_sau', cbil.strip(), None)
                    if order_id_val:
                        try:
                            _ = update_database_immediately(order_id_val, cbil, "success", debt, "Postpaid lookup ok", None)
                        except Exception as _e:
                            logger.warning(f"DB update lỗi (Postpaid debt) cho {cbil}: {_e}")
                    continue
                except:
                    # Nếu có thông báo info/alert khác thì lưu notes giống hiển thị
                    note_text = get_info_alert_text() or get_error_alert_text() or ""
                    data.append([cbil, "Không tìm thấy nợ cước", Config.STATUS_INCOMPLETE])
                    display_line = f"{cbil} - null{(' | ' + note_text) if note_text else ''}"
                    insert_ctmed(tkinp_ctmed, display_line)
                    # Cập nhật DB failed với notes giống hiển thị
                    order_id_val = db_find_order_id('tra_cuu_no_tra_sau', cbil.strip(), None)
                    if order_id_val:
                        try:
                            _ = update_database_immediately(order_id_val, cbil, "failed", None, display_line, None)
                        except Exception as _e:
                            logger.warning(f"DB update lỗi (Postpaid failed) cho {cbil}: {_e}")
                    continue
        time.sleep(2)
        if len(data) > 0:
            name_dir = "Tra cứu nợ trả sau"
            export_excel(data, name_dir)
    except Exception as e:
        logger.error(f"Lỗi tra cứu nợ trả sau: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi tra cứu nợ trả sau: {e}")

def form_lookup_card():
    cus_frm = tk.Frame(root)
    cus_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    ctm_frm = tk.Frame(cus_frm)
    ctm_frm.pack(expand=True, side="left")
    ctmed_frm = tk.Frame(cus_frm)
    ctmed_frm.pack(expand=True, side="right")
    button_frm = tk.Frame(root)
    button_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    tklbl_ctm = tk.Label(ctm_frm, text="SĐT tra cứu")
    tklbl_ctm.pack(side="top")
    tkinp_ctm = tk.Text(ctm_frm, height=16, width=24)
    tkinp_ctm.pack(side="left", pady=8)
    tklbl_ctm = tk.Label(ctmed_frm, text="Đã xử lý")
    tklbl_ctm.pack(side="top")
    tkinp_ctmed = tk.Text(ctmed_frm, height=16, width=32, bg="#ccc", state="disabled")
    tkinp_ctmed.pack(side="left", pady=8)
    style = ttk.Style()
    style.configure("Red.TButton", foreground="red")
    style.configure("Blue.TButton", foreground="blue")
    style.configure("Green.TButton", foreground="green")
    tkbtn_get_data = ttk.Button(button_frm, text="Get dữ liệu", command=lambda: get_data_postpaid(tkinp_ctm))
    tkbtn_get_data.pack(side='left', padx=5, pady=5)
    tkbtn_get_data.configure(style="Green.TButton")
    tkbtn_payment = ttk.Button(button_frm, text="Bắt đầu", command=lambda: lookup_card(tkinp_ctm, tkinp_ctmed))
    tkbtn_payment.pack(side='left', padx=5, pady=5)
    tkbtn_payment.configure(style="Blue.TButton") 
    tkbtn_destroy = ttk.Button(button_frm, text="Dừng lại", command=stop_tool)
    tkbtn_destroy.pack(side='right', padx=5, pady=5)
    tkbtn_destroy.configure(style="Red.TButton") 


def check_api_health():
    """Kiểm tra Node API (nguồn dữ liệu thật) có hoạt động không"""
    try:
        response = requests.get(f"{Config.NODE_SERVER_URL}/api/health", timeout=5)
        return response.status_code == 200
    except Exception:
        return False

def db_fetch_service_data(service_type: str, payment_type: str = None) -> Optional[Dict[str, Any]]:
    """Đọc thẳng từ Postgres: trả về dữ liệu mẫu theo service_type.
    - Lấy tối đa 20 giao dịch mới nhất từ service_transactions join orders theo service_type
    - Trả về code list theo định dạng của từng dịch vụ và kèm order_id gần nhất + mapping code->orderId
    - Đối với nap_tien_da_mang: có thể lọc theo payment_type (prepaid/postpaid)
    """
    try:
        with psycopg2.connect(os.getenv('DATABASE_URL', 'postgres://postgres:123456@localhost:5432/autogachno')) as conn:
            with conn.cursor() as cur:
                # Lấy mỗi mã (code) 1 dòng: chọn transaction mới nhất cho từng code
                if service_type == "nap_tien_da_mang" and payment_type:
                    # Lọc theo loại dịch vụ cho nạp tiền đa mạng
                    if payment_type == "prepaid":
                        # Nạp trả trước: chỉ lấy format sđt|số tiền (có dấu |)
                        print(f"[DEBUG] Lọc dữ liệu Nạp trả trước (có dấu |)")
                        cur.execute(
                            """
                            SELECT DISTINCT ON (st.code) st.code, st.order_id, st.created_at
                            FROM service_transactions st
                            JOIN orders o ON o.id = st.order_id
                            WHERE o.service_type = %s
                              AND st.status IN ('pending','processing')
                              AND st.code LIKE '%%|%%'
                            ORDER BY st.code, st.created_at DESC
                            """,
                            (service_type,)
                        )
                    elif payment_type == "postpaid":
                        # Gạch nợ trả sau: chỉ lấy số điện thoại (không có dấu |)
                        print(f"[DEBUG] Lọc dữ liệu Gạch nợ trả sau (không có dấu |)")
                        cur.execute(
                            """
                            SELECT DISTINCT ON (st.code) st.code, st.order_id, st.created_at
                            FROM service_transactions st
                            JOIN orders o ON o.id = st.order_id
                            WHERE o.service_type = %s
                              AND st.status IN ('pending','processing')
                              AND st.code NOT LIKE '%%|%%'
                            ORDER BY st.code, st.created_at DESC
                            """,
                            (service_type,)
                        )
                    else:
                        # Không lọc: lấy tất cả
                        print(f"[DEBUG] Lấy tất cả dữ liệu đa mạng")
                        cur.execute(
                            """
                            SELECT DISTINCT ON (st.code) st.code, st.order_id, st.created_at
                            FROM service_transactions st
                            JOIN orders o ON o.id = st.order_id
                            WHERE o.service_type = %s
                              AND st.status IN ('pending','processing')
                            ORDER BY st.code, st.created_at DESC
                            """,
                            (service_type,)
                        )
                else:
                    # Các dịch vụ khác: lấy tất cả
                    cur.execute(
                        """
                        SELECT DISTINCT ON (st.code) st.code, st.order_id, st.created_at
                        FROM service_transactions st
                        JOIN orders o ON o.id = st.order_id
                        WHERE o.service_type = %s
                          AND st.status IN ('pending','processing')
                        ORDER BY st.code, st.created_at DESC
                        """,
                        (service_type,)
                    )
                
                rows = cur.fetchall()
                print(f"[DEBUG] Số dòng dữ liệu trả về: {len(rows)}")
                codes = []
                code_order_map = []
                for code_val, order_val, _ in rows:
                    if not code_val:
                        continue
                    codes.append(code_val)
                    if order_val:
                        code_order_map.append({'code': code_val, 'orderId': order_val})
                
                print(f"[DEBUG] Dữ liệu codes: {codes}")
                print(f"[DEBUG] Payment type: {payment_type}")
                # Lấy order_id mới nhất theo created_at
                latest_order_id = None
                if rows:
                    latest_row = max(rows, key=lambda r: r[2] or datetime.min)
                    latest_order_id = latest_row[1]

                result: Dict[str, Any] = {}
                if service_type in ("tra_cuu_ftth", "thanh_toan_tv_internet"):
                    result["subscriber_codes"] = codes[:10]
                elif service_type == "gach_dien_evn":
                    result["bill_codes"] = codes[:10]
                elif service_type in ("nap_tien_da_mang", "nap_tien_viettel", "tra_cuu_no_tra_sau"):
                    result["phone_numbers"] = codes[:10]
                else:
                    # fallback chung
                    result["codes"] = codes[:10]

                result["order_id"] = latest_order_id
                result["code_order_map"] = code_order_map
                return result
    except Exception as e:
        logger.error(f"Lỗi đọc DB cho {service_type}: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi kết nối DB: {e}")
        return None

def fetch_api_data(service_type: str) -> Optional[Dict]:
    """Đã chuyển sang đọc trực tiếp DB (bỏ API)."""
    return db_fetch_service_data(service_type)

def populate_text_widget(text_widget, data_list):
    """Đổ dữ liệu vào Text widget"""
    try:
        text_widget.config(state="normal")
        text_widget.delete("1.0", "end")
        if data_list:
            text_widget.insert("1.0", "\n".join(data_list))
        text_widget.config(state="normal")
    except Exception as e:
        logger.error(f"Lỗi đổ dữ liệu vào text widget: {e}")

def populate_entry_widget(entry_widget, value):
    """Đổ dữ liệu vào Entry widget"""
    try:
        entry_widget.delete(0, "end")
        if value:
            entry_widget.insert(0, str(value))
    except Exception as e:
        logger.error(f"Lỗi đổ dữ liệu vào entry widget: {e}")

def populate_combobox_widget(combobox_widget, value):
    """Đổ dữ liệu vào Combobox widget"""
    try:
        if value and value in combobox_widget['values']:
            combobox_widget.set(value)
    except Exception as e:
        logger.error(f"Lỗi đổ dữ liệu vào combobox widget: {e}")


def export_excel(data: List[Tuple[str, Any, str]], name_dir: str) -> bool:
    """Xuất dữ liệu ra file Excel"""
    today = datetime.now().strftime("%H%M-%d-%m-%Y")
    try:
        export_dir = os.path.join(os.getcwd(), f"{Config.FOLDER_RESULT}\\{name_dir}")
        os.makedirs(export_dir, exist_ok=True)
        
        file_name = f"{today}.xlsx"
        file_path = os.path.join(export_dir, file_name)

        wb = Workbook()
        ws = wb.active
        
        # Định dạng header
        headers = ['STT', 'Số thuê bao', 'Số tiền', 'Ghi chú']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # Ghi dữ liệu
        for idx, (phone, amount, note) in enumerate(data, start=2):
            ws.cell(row=idx, column=1, value=idx - 1)
            ws.cell(row=idx, column=2, value=phone)
            ws.cell(row=idx, column=3, value=amount)
            ws.cell(row=idx, column=4, value=note)
            
        wb.save(file_path)
        
        # Hỏi người dùng có muốn mở file không
        if messagebox.askyesno(Config.TITLE, f"Dữ liệu được lưu tại: \n{Config.FOLDER_RESULT}/{name_dir}/{today}"):
            try:
                os.startfile(file_path)
            except Exception as e:
                logger.warning(f"Không thể mở file Excel: {e}")
                
        return True
        
    except Exception as e:
        logger.error(f"Lỗi xuất Excel: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi xuất Excel: {e}")
        return False

def valid_data(data: List[Any]) -> bool:
    """Kiểm tra dữ liệu đầu vào"""
    try:
        if not check_username():
            return False
        for item in data:
            # Nếu là danh sách (ví dụ danh sách mã), yêu cầu có ít nhất 1 phần tử không rỗng
            if isinstance(item, (list, tuple)):
                has_nonempty = any((isinstance(x, str) and x.strip()) for x in item)
                if not has_nonempty:
                    #messagebox.showwarning(Config.TITLE, "Vui lòng nhập đầy đủ thông tin")
                    return False
            else:
                # Xử lý chuỗi/tham số đơn lẻ
                text = str(item) if item is not None else ""
                if not text.strip():
                    #messagebox.showwarning(Config.TITLE, "Vui lòng nhập đầy đủ thông tin")
                    return False
        return True
    except Exception as e:
        logger.error(f"Lỗi kiểm tra dữ liệu: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi kiểm tra dữ liệu: {e}")
        return False

def delete_ctmed(cmted: tk.Text):
    """Xóa nội dung text widget"""
    cmted.config(state="normal")
    cmted.delete("1.0", "end")
    cmted.config(state="disabled")

def insert_ctmed(cmted: tk.Text, cbil: str):
    """Thêm text vào widget"""
    cmted.config(state="normal")
    cmted.insert("1.0", f"{cbil}\n")
    cmted.config(state="disabled")

def stop_tool():
    """Dừng chương trình"""
    global stop_flag
    stop_flag = True
    #messagebox.showinfo(Config.TITLE, "Đã dừng chương trình")

def update_stop_flag():
    """Reset stop flag"""
    global stop_flag
    stop_flag = False

def get_chrome_driver(username: str = "default") -> Optional[webdriver.Chrome]:
    """Tạo Chrome driver"""
    try:
        profile_dir = os.path.join(os.getcwd(), "chrome_profile", username)
        os.makedirs(profile_dir, exist_ok=True)
        
        options = Options()
        options.add_argument(f"--user-data-dir={profile_dir}")
        options.add_argument("--start-maximized")
        options.add_argument("--disable-notifications")
        options.add_argument("--disable-popup-blocking")
        options.add_experimental_option("excludeSwitches", ["enable-logging"])
        
        # Thêm các tùy chọn bảo mật
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        
        driver = webdriver.Chrome(options=options)
        return driver
    except Exception as e:
        logger.error(f"Lỗi khởi tạo Chrome Driver: {e}")
        return None

def check_username() -> bool:
    """Kiểm tra username đăng nhập dựa trên thông tin hiển thị trên trang."""
    try:
        # Bỏ qua kiểm tra đăng nhập (yêu cầu)
        if driver is None:
            return True

        expected_username = (dbfiles or {}).get("username", "").strip()
        if not expected_username:
            return True

        attempts = 3
        for _ in range(attempts):
            try:
                dl_info = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "dl-info-detail"))
                )
                # Tìm tất cả các div info trong dl-info-detail
                info_divs = dl_info.find_elements(By.CLASS_NAME, "info")
                if len(info_divs) >= 3:
                    # Lấy mã đại lý từ div đầu tiên
                    agency_span = info_divs[0].find_element(By.TAG_NAME, "span")
                    agency_code = (agency_span.text or "").strip()
                    
                    # Lấy số điện thoại từ div thứ 3
                    phone_span = info_divs[2].find_element(By.TAG_NAME, "span")
                    phone_number = (phone_span.text or "").strip()
                    
                    # Kiểm tra username có khớp với mã đại lý hoặc số điện thoại không
                    if (expected_username == agency_code or 
                        expected_username == phone_number or
                        expected_username in agency_code or
                        expected_username in phone_number):
                        return True
                    else:
                        #messagebox.showerror(Config.TITLE, f"Vui lòng sử dụng đúng tài khoản đã đăng ký. Mong đợi: {expected_username}, Tìm thấy: {agency_code}")
                        return False
                return True
            except StaleElementReferenceException:
                time.sleep(0.5)
                continue
            except Exception as e:
                logger.warning(f"Lỗi parse thông tin đăng nhập: {e}")
                time.sleep(0.5)
                continue
        # Nếu vẫn không lấy được sau retry
        #messagebox.showerror(Config.TITLE, "Không tìm thấy thông tin tài khoản trên Viettel Pay Pro")
        return False
    except Exception as e:
        logger.error(f"Lỗi kiểm tra username: {e}")
        #messagebox.showerror(Config.TITLE, "Không tìm thấy thông tin tài khoản trên Viettel Pay Pro")
        return False

def get_number_uses() -> Tuple[int, Dict[str, int]]:
    """Lấy số lần sử dụng các dịch vụ"""
    try:
        services = {
            Config.SERVICES['payment_internet']: 99999,
            Config.SERVICES['payment_card']: 99999,
            Config.SERVICES['lookup_card']: 99999,
            Config.SERVICES['lookup_ftth']: 99999,
            Config.SERVICES['payment_evn']: 99999
        }
        return 0, services
    except Exception as e:
        logger.error(f"Lỗi lấy số lần sử dụng: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi lấy số lần sử dụng: {e}")
        return 0, {}

def handle_choose_amount(am: str) -> str:
    """Xử lý lựa chọn số tiền"""
    try:
        amount_map = {
            "10.000đ": "0", "20.000đ": "1", "30.000đ": "2", "50.000đ": "3",
            "100.000đ": "4", "200.000đ": "5", "300.000đ": "6", "500.000đ": "7"
        }
        return amount_map.get(am, "0")
    except Exception as e:
        logger.error(f"Lỗi xử lý số tiền: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi xử lý số tiền: {e}")
        return "0"

def amount_by_cbil(cbil: str, element, lookup: bool = False) -> Tuple[bool, Any, Optional[str]]:
    """Lấy số tiền theo mã thuê bao"""
    try:
        amount = "Không tìm thấy mã thuê bao"
        payment_id = None
        
        html_content = element.get_attribute('outerHTML')
        soup = BeautifulSoup(html_content, 'html.parser')
        pay_content_groups = soup.find_all("div", class_="row pay-content mb-3")
        
        for group in pay_content_groups:
            p_tags = group.find_all("p")
            is_found = any(cbil in p_tag.text for p_tag in p_tags)
            
            if lookup and is_found:
                button_tag = group.find("button", {"id": re.compile(r'payMoneyForm:btnView\d*')})
                if button_tag:
                    payment_id = button_tag['id']
            
            if is_found:
                for p_tag in p_tags:
                    if "VND" in p_tag.text:
                        str_price = p_tag.text.split("VND")[0].strip()
                        amount = int(str_price.replace(",", ""))
                        if amount >= 5000:
                            return True, amount, payment_id
                        else:
                            return False, amount, payment_id
                            
        return False, amount, payment_id
        
    except Exception as e:
        logger.error(f"Lỗi lấy số tiền: {e}")
        return False, "Lỗi thanh toán", None

def get_error_alert_text() -> Optional[str]:
    """Trả về nội dung thông báo lỗi (role="alert") nếu có."""
    try:
        elements = driver.find_elements(By.CSS_SELECTOR, "li[role='alert'] span.ui-messages-error-summary")
        for el in elements:
            text_val = (el.text or "").strip()
            if text_val:
                return text_val
    except Exception:
        pass
    return None

def get_info_alert_text() -> Optional[str]:
    """Trả về nội dung thông báo info (role="alert") nếu có."""
    try:
        elements = driver.find_elements(By.CSS_SELECTOR, "li[role='alert'] span.ui-messages-info-summary")
        for el in elements:
            text_val = (el.text or "").strip()
            if text_val:
                return text_val
    except Exception:
        pass
    return None


def initialize_browser(username="default"):
    global driver
    driver = get_chrome_driver(username)
    if driver:
        driver.get(Config.DRIVER_LINK)
        return driver
    return None

def cleanup():
    global driver
    if driver:
        try:
            driver.quit()
        except Exception as e:
            logger.warning(f"Lỗi khi đóng driver: {e}")
        finally:
            driver = None

def is_logged_in(driver):
    # Bỏ qua kiểm tra trạng thái đăng nhập: luôn cho phép chạy
            return True

def login_process():
    """Đăng nhập tự động khi mở trình duyệt bằng tài khoản cấu hình."""
    try:
        inp_usr = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, "loginForm:userName")))
        inp_usr.clear() 
        inp_usr.send_keys(LOGIN_USERNAME)
        inp_pwd = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, "loginForm:password")))
        inp_pwd.clear() 
        inp_pwd.send_keys(LOGIN_PASSWORD)
        print("[LOGIN] Đã điền thông tin đăng nhập")
    except Exception as e:
        logger.warning(f"Lỗi đăng nhập: {e}")
        pass

def show_services_form():
    try:
        main_frm = tk.Frame(root)
        main_frm.pack(expand=True, side="top", padx=6, pady=6, fill="both")
        tklbl_choose = tk.Label(main_frm, text="Loại thanh toán:")
        tklbl_choose.pack(side="left")
        tkcbb_choose = ttk.Combobox(main_frm, values=[
            "Tra cứu FTTH",
            "Gạch điện EVN", 
            "Nạp tiền đa mạng",
            "Nạp tiền mạng Viettel",
            "Thanh toán TV - Internet",
            "Tra cứu nợ thuê bao trả sau"
        ], width="32", state="readonly")
        tkcbb_choose.pack(side="left", padx=6, expand=True, fill="x")
        tkcbb_choose.set("Tra cứu FTTH")
        def handle_choose_services(event, choose, main_frm):
            service = choose.get()
            clear_widgets(main_frm)
            if service == "Tra cứu FTTH":
                form_lookup_ftth()
            elif service == "Gạch điện EVN":
                form_debt_electric() 
            elif service == "Nạp tiền đa mạng":
                form_payment_phone()
            elif service == "Nạp tiền mạng Viettel":
                form_payment_viettel()
            elif service == "Thanh toán TV - Internet":
                form_payment_internet()
            elif service == "Tra cứu nợ thuê bao trả sau":
                form_lookup_card()
        tkcbb_choose.bind("<<ComboboxSelected>>", lambda event: handle_choose_services(event, tkcbb_choose, main_frm))
        handle_choose_services(None, tkcbb_choose, main_frm)
    except Exception as e:
        logger.error(f"Lỗi hiển thị form dịch vụ: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi hiển thị form dịch vụ: {e}")


root = tk.Tk()
root.title("HPK Tool - Viettel Pay Automation")
root.geometry("500x550") 
root.option_add("*Font", "Arial 10")
try:
    root.iconbitmap(Config.ICON_FILE)
except Exception as e:
    logger.warning(f"Không thể tải icon: {e}")
    pass

def clear_widgets(main_frm):
    for widget in root.winfo_children():
        if widget is not main_frm:
            widget.destroy()


def main():
    global driver
    # Khởi tạo trình duyệt và đăng nhập tự động bằng tài khoản mặc định
    try:
        driver = initialize_browser(LOGIN_USERNAME or "default")
        try:
            login_process()
        except Exception:
            pass
    except Exception as e:
        logger.error(f"Lỗi khởi tạo trình duyệt: {e}")
    # Hiển thị form dịch vụ luôn, không cần đọc cấu hình
    show_services_form()

if __name__ == "__main__":
    try:
        main()
        root.protocol("WM_DELETE_WINDOW", lambda: [cleanup(), root.destroy()])
        root.mainloop()
    except Exception as e:
        logger.error(f"Lỗi chính: {e}")
        #messagebox.showerror("Lỗi", f"Lỗi khởi động ứng dụng: {e}")
    finally:
        cleanup()
