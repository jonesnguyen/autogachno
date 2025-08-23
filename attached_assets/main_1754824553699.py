import os
import re
import time
import json 
import tkinter as tk
from tkinter import ttk, #messagebox
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import pytz
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from cryptography.fernet import Fernet
from pymongo.mongo_client import MongoClient
from bs4 import BeautifulSoup
import sys
import logging
import requests
import threading
from typing import List, Tuple, Optional, Dict, Any
from dataclasses import dataclass

# Cấu hình logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('app.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

@dataclass
class Config:
    """Cấu hình ứng dụng"""
    DRIVER_LINK: str = "https://kpp.bankplus.vn"
    FOLDER_RESULT: str = "ket_qua"
    TITLE: str = "Thông báo"
    CONFIG_FILE: str = "config.json"
    ICON_FILE: str = "viettelpay.ico"
    COPYRIGHT_KEY: bytes = b"h_ThisAAutoToolVjppro-CopyRight-ByCAOAC7690="
    STATUS_COMPLETE: str = "Đã xử lý"
    STATUS_INCOMPLETE: str = "Chưa xử lý"
    
    # MongoDB URI - nên đặt trong biến môi trường
    MONGODB_URI: str = os.getenv('MONGODB_URI', "mongodb+srv://huytq0104:huypro7690@hpk.jdkmqqs.mongodb.net/?retryWrites=true&w=majority&appName=HPK")
    
    # Service types
    SERVICES = {
        'payment_internet': 'payment_internet',
        'payment_card': 'deb_cart', 
        'lookup_card': 'lookup_cart',
        'lookup_ftth': 'lookup_ftth',
        'payment_evn': 'deb_evn'
    }
    
    # API Configuration
    API_BASE_URL: str = os.getenv('API_BASE_URL', "http://127.0.0.1:8080")
    API_TIMEOUT: int = 10

# Global variables
stop_flag = False
driver = None
dbfiles = None
times_exits = {}
api_server_thread = None

def get_exe_dir():
    exe_path = sys.argv[0]
    exe_dir = os.path.dirname(exe_path)
    return exe_dir

def start_api_server():
    """Khởi động API server trong background"""
    global api_server_thread
    try:
        # Import và khởi động API server
        from mock_api_server import start_api_server as start_server
        api_server_thread = start_server()
        logger.info("API Server started successfully")
        return True
    except Exception as e:
        logger.warning(f"Không thể khởi động API server: {e}")
        return False

def check_api_health():
    """Kiểm tra API server có hoạt động không"""
    try:
        response = requests.get(f"{Config.API_BASE_URL}/api/health", timeout=5)
        return response.status_code == 200
    except Exception:
        return False

def fetch_api_data(service_type: str) -> Optional[Dict]:
    """Gọi API để lấy dữ liệu mẫu"""
    try:
        url = f"{Config.API_BASE_URL}/api/data/{service_type}"
        response = requests.get(url, timeout=Config.API_TIMEOUT)
        
        if response.status_code == 200:
            data = response.json()
            if data.get("status") == "success":
                return data.get("data")
        return None
    except Exception as e:
        logger.error(f"Lỗi gọi API {service_type}: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi kết nối API: {e}")
        return None

def populate_text_widget(text_widget, data_list):
    """Đổ dữ liệu vào Text widget"""
    try:
        text_widget.config(state="normal")
        text_widget.delete("1.0", "end")
        if data_list:
            text_widget.insert("1.0", "\n".join(data_list))
        text_widget.config(state="normal")
    except Exception as e:
        logger.error(f"Lỗi đổ dữ liệu vào text widget: {e}")

def populate_entry_widget(entry_widget, value):
    """Đổ dữ liệu vào Entry widget"""
    try:
        entry_widget.delete(0, "end")
        if value:
            entry_widget.insert(0, str(value))
    except Exception as e:
        logger.error(f"Lỗi đổ dữ liệu vào entry widget: {e}")

def populate_combobox_widget(combobox_widget, value):
    """Đổ dữ liệu vào Combobox widget"""
    try:
        if value and value in combobox_widget['values']:
            combobox_widget.set(value)
    except Exception as e:
        logger.error(f"Lỗi đổ dữ liệu vào combobox widget: {e}")

def get_data_ftth(text_widget):
    """Lấy dữ liệu API cho Tra cứu FTTH"""
    try:
        data = fetch_api_data("tra_cuu_ftth")
        if data and "subscriber_codes" in data:
            populate_text_widget(text_widget, data["subscriber_codes"])
            #messagebox.showinfo(Config.TITLE, f"Đã tải {len(data['subscriber_codes'])} mã thuê bao FTTH")
        else:
            #messagebox.showwarning(Config.TITLE, "Không có dữ liệu từ API")
    except Exception as e:
        logger.error(f"Lỗi lấy dữ liệu FTTH: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi lấy dữ liệu FTTH: {e}")

def get_data_evn(text_widget, phone_widget, pin_widget):
    """Lấy dữ liệu API cho Gạch điện EVN"""
    try:
        data = fetch_api_data("gach_dien_evn")
        if data:
            if "bill_codes" in data:
                populate_text_widget(text_widget, data["bill_codes"])
            if "receiver_phone" in data:
                populate_entry_widget(phone_widget, data["receiver_phone"])
            if "pin" in data:
                populate_entry_widget(pin_widget, data["pin"])
            
            count = len(data.get("bill_codes", []))
            #messagebox.showinfo(Config.TITLE, f"Đã tải {count} mã hóa đơn điện EVN")
        else:
            #messagebox.showwarning(Config.TITLE, "Không có dữ liệu từ API")
    except Exception as e:
        logger.error(f"Lỗi lấy dữ liệu EVN: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi lấy dữ liệu EVN: {e}")

def get_data_multi_network(text_widget, pin_widget, form_widget, amount_widget):
    """Lấy dữ liệu API cho Nạp tiền đa mạng"""
    try:
        data = fetch_api_data("nap_tien_da_mang")
        if data:
            if "phone_numbers" in data:
                populate_text_widget(text_widget, data["phone_numbers"])
            if "pin" in data:
                populate_entry_widget(pin_widget, data["pin"])
            if "payment_type" in data:
                populate_combobox_widget(form_widget, data["payment_type"])
            if "amount" in data:
                populate_combobox_widget(amount_widget, data["amount"])
            
            count = len(data.get("phone_numbers", []))
            #messagebox.showinfo(Config.TITLE, f"Đã tải {count} số điện thoại đa mạng")
        else:
            #messagebox.showwarning(Config.TITLE, "Không có dữ liệu từ API")
    except Exception as e:
        logger.error(f"Lỗi lấy dữ liệu đa mạng: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi lấy dữ liệu đa mạng: {e}")

def get_data_viettel(text_widget, pin_widget, amount_widget):
    """Lấy dữ liệu API cho Nạp tiền mạng Viettel"""
    try:
        data = fetch_api_data("nap_tien_viettel")
        if data:
            if "phone_numbers" in data:
                populate_text_widget(text_widget, data["phone_numbers"])
            if "pin" in data:
                populate_entry_widget(pin_widget, data["pin"])
            if "amount" in data:
                populate_entry_widget(amount_widget, data["amount"])
            
            count = len(data.get("phone_numbers", []))
            #messagebox.showinfo(Config.TITLE, f"Đã tải {count} số điện thoại Viettel")
        else:
            #messagebox.showwarning(Config.TITLE, "Không có dữ liệu từ API")
    except Exception as e:
        logger.error(f"Lỗi lấy dữ liệu Viettel: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi lấy dữ liệu Viettel: {e}")

def get_data_tv_internet(text_widget, pin_widget):
    """Lấy dữ liệu API cho Thanh toán TV - Internet"""
    try:
        data = fetch_api_data("thanh_toan_tv_internet")
        if data:
            if "subscriber_codes" in data:
                populate_text_widget(text_widget, data["subscriber_codes"])
            if "pin" in data:
                populate_entry_widget(pin_widget, data["pin"])
            
            count = len(data.get("subscriber_codes", []))
            #messagebox.showinfo(Config.TITLE, f"Đã tải {count} mã thuê bao TV-Internet")
        else:
            #messagebox.showwarning(Config.TITLE, "Không có dữ liệu từ API")
    except Exception as e:
        logger.error(f"Lỗi lấy dữ liệu TV-Internet: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi lấy dữ liệu TV-Internet: {e}")

def get_data_postpaid(text_widget):
    """Lấy dữ liệu API cho Tra cứu nợ thuê bao trả sau"""
    try:
        data = fetch_api_data("tra_cuu_no_tra_sau")
        if data and "phone_numbers" in data:
            populate_text_widget(text_widget, data["phone_numbers"])
            #messagebox.showinfo(Config.TITLE, f"Đã tải {len(data['phone_numbers'])} số điện thoại trả sau")
        else:
            #messagebox.showwarning(Config.TITLE, "Không có dữ liệu từ API")
    except Exception as e:
        logger.error(f"Lỗi lấy dữ liệu trả sau: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi lấy dữ liệu trả sau: {e}")

def export_excel(data: List[Tuple[str, Any, str]], name_dir: str) -> bool:
    """Xuất dữ liệu ra file Excel"""
    today = datetime.now().strftime("%H%M-%d-%m-%Y")
    try:
        export_dir = os.path.join(os.getcwd(), f"{Config.FOLDER_RESULT}\\{name_dir}")
        os.makedirs(export_dir, exist_ok=True)
        
        file_name = f"{today}.xlsx"
        file_path = os.path.join(export_dir, file_name)

        wb = Workbook()
        ws = wb.active
        
        # Định dạng header
        headers = ['STT', 'Số thuê bao', 'Số tiền', 'Ghi chú']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # Ghi dữ liệu
        for idx, (phone, amount, note) in enumerate(data, start=2):
            ws.cell(row=idx, column=1, value=idx - 1)
            ws.cell(row=idx, column=2, value=phone)
            ws.cell(row=idx, column=3, value=amount)
            ws.cell(row=idx, column=4, value=note)
            
        wb.save(file_path)
        
        # Hỏi người dùng có muốn mở file không
        if messagebox.askyesno(Config.TITLE, f"Dữ liệu được lưu tại: \n{Config.FOLDER_RESULT}/{name_dir}/{today}"):
            try:
                os.startfile(file_path)
            except Exception as e:
                logger.warning(f"Không thể mở file Excel: {e}")
                
        return True
        
    except Exception as e:
        logger.error(f"Lỗi xuất Excel: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi xuất Excel: {e}")
        return False

def encode_json() -> Optional[Dict[str, Any]]:
    """Giải mã cấu hình từ file"""
    global dbfiles
    try:
        if not os.path.exists(Config.CONFIG_FILE):
            with open(Config.CONFIG_FILE, "w", encoding='utf-8') as json_file:
                json.dump({"files": ""}, json_file, indent=4, ensure_ascii=False)
        
        with open(Config.CONFIG_FILE, "r", encoding='utf-8') as json_file: 
            data = json.load(json_file)
        
        if not data.get("files"):
            return None
            
        cipher = Fernet(Config.COPYRIGHT_KEY)
        decrypted_data = bytes.fromhex(data["files"])
        result = cipher.decrypt(decrypted_data).decode("utf-8")
        dbfiles = json.loads(result)
        return dbfiles
    except Exception as e:
        logger.error(f"Lỗi đọc file cấu hình: {e}")
        #messagebox.showerror(Config.TITLE, f"Không thể đọc file cấu hình: {e}")
        return None

def valid_data(data: List[str]) -> bool:
    """Kiểm tra dữ liệu đầu vào"""
    try:
        if not check_username():
            return False
        for item in data:
            if not item or not item.strip():
                #messagebox.showwarning(Config.TITLE, "Vui lòng nhập đầy đủ thông tin")
                return False
        return True
    except Exception as e:
        logger.error(f"Lỗi kiểm tra dữ liệu: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi kiểm tra dữ liệu: {e}")
        return False

def delete_ctmed(cmted: tk.Text):
    """Xóa nội dung text widget"""
    cmted.config(state="normal")
    cmted.delete("1.0", "end")
    cmted.config(state="disabled")

def insert_ctmed(cmted: tk.Text, cbil: str):
    """Thêm text vào widget"""
    cmted.config(state="normal")
    cmted.insert("1.0", f"{cbil}\n")
    cmted.config(state="disabled")

def stop_tool():
    """Dừng chương trình"""
    global stop_flag
    stop_flag = True
    #messagebox.showinfo(Config.TITLE, "Đã dừng chương trình")

def update_stop_flag():
    """Reset stop flag"""
    global stop_flag
    stop_flag = False

def toggle_input_amount(select, label, combobox):
    selected_value = select.get()
    if selected_value == "Gạch nợ trả sau":
        combobox.pack_forget()
        label.pack_forget()
    else:
        combobox.pack(side="right")
        label.pack(side="right")
    root.update()

def handle_choose_select(choose: str) -> int:
    """Xử lý lựa chọn loại thanh toán"""
    try:
        choose = choose.strip()
        if choose == "Nạp trả trước":
            return 1
        else:
            return 2
    except Exception as e:
        logger.error(f"Lỗi xử lý loại thanh toán: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi xử lý loại thanh toán: {e}")
        return 1

def get_chrome_driver(username: str = "default") -> Optional[webdriver.Chrome]:
    """Tạo Chrome driver"""
    try:
        profile_dir = os.path.join(os.getcwd(), "chrome_profile", username)
        os.makedirs(profile_dir, exist_ok=True)
        
        options = Options()
        options.add_argument(f"--user-data-dir={profile_dir}")
        options.add_argument("--start-maximized")
        options.add_argument("--disable-notifications")
        options.add_argument("--disable-popup-blocking")
        options.add_experimental_option("excludeSwitches", ["enable-logging"])
        
        # Thêm các tùy chọn bảo mật
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        
        driver = webdriver.Chrome(options=options)
        return driver
    except Exception as e:
        logger.error(f"Lỗi khởi tạo Chrome Driver: {e}")
        return None

def check_username() -> bool:
    """Kiểm tra username đăng nhập"""
    try:
        dl_info = WebDriverWait(driver, 4).until(EC.presence_of_element_located((By.CLASS_NAME, "dl-info-detail")))
        first_div = dl_info.find_element(By.XPATH, "./div[1]")
        span_ele = first_div.find_element(By.XPATH, ".//span")
        username = span_ele.text
        if dbfiles.get("username").strip() != username:
            #messagebox.showerror(Config.TITLE, "Vui lòng sử dụng đúng tài khoản đã đăng ký")
            return False
        else:
            return True
    except Exception as e:
        logger.error(f"Lỗi kiểm tra username: {e}")
        #messagebox.showerror(Config.TITLE, "Không tìm thấy tên tài khoản trên Viettel Pay Pro")
        return False

def get_number_uses() -> Tuple[int, Dict[str, int]]:
    """Lấy số lần sử dụng các dịch vụ"""
    try:
        services = {
            Config.SERVICES['payment_internet']: 99999,
            Config.SERVICES['payment_card']: 99999,
            Config.SERVICES['lookup_card']: 99999,
            Config.SERVICES['lookup_ftth']: 99999,
            Config.SERVICES['payment_evn']: 99999
        }
        return 0, services
    except Exception as e:
        logger.error(f"Lỗi lấy số lần sử dụng: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi lấy số lần sử dụng: {e}")
        return 0, {}

def handle_choose_amount(am: str) -> str:
    """Xử lý lựa chọn số tiền"""
    try:
        amount_map = {
            "10.000đ": "0", "20.000đ": "1", "30.000đ": "2", "50.000đ": "3",
            "100.000đ": "4", "200.000đ": "5", "300.000đ": "6", "500.000đ": "7"
        }
        return amount_map.get(am, "0")
    except Exception as e:
        logger.error(f"Lỗi xử lý số tiền: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi xử lý số tiền: {e}")
        return "0"

def amount_by_cbil(cbil: str, element, lookup: bool = False) -> Tuple[bool, Any, Optional[str]]:
    """Lấy số tiền theo mã thuê bao"""
    try:
        amount = "Không tìm thấy mã thuê bao"
        payment_id = None
        
        html_content = element.get_attribute('outerHTML')
        soup = BeautifulSoup(html_content, 'html.parser')
        pay_content_groups = soup.find_all("div", class_="row pay-content mb-3")
        
        for group in pay_content_groups:
            p_tags = group.find_all("p")
            is_found = any(cbil in p_tag.text for p_tag in p_tags)
            
            if lookup and is_found:
                button_tag = group.find("button", {"id": re.compile(r'payMoneyForm:btnView\d*')})
                if button_tag:
                    payment_id = button_tag['id']
            
            if is_found:
                for p_tag in p_tags:
                    if "VND" in p_tag.text:
                        str_price = p_tag.text.split("VND")[0].strip()
                        amount = int(str_price.replace(",", ""))
                        if amount >= 5000:
                            return True, amount, payment_id
                        else:
                            return False, amount, payment_id
                            
        return False, amount, payment_id
        
    except Exception as e:
        logger.error(f"Lỗi lấy số tiền: {e}")
        return False, "Lỗi thanh toán", None

def payment_internet(tkinp_ctm, tkinp_ctmed, tkinp_pin):
    try:
        delete_ctmed(tkinp_ctmed)
        update_stop_flag()
        pin = tkinp_pin.get()
        cbils = tkinp_ctm.get("1.0", "end-1c").splitlines()
        if not valid_data([cbils, pin]):
            return False
        data = []
        for cbil in cbils:
            root.update()
            time.sleep(0.5)
            cbil = cbil.strip()
            if not stop_flag and cbil.strip() != "":
                try:
                    customer = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:contractCode")))
                    customer.clear()
                    customer.send_keys(cbil)
                    time.sleep(0.5)
                    payment_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:btnPay0")))
                    payment_button.click()
                    time.sleep(1)
                    WebDriverWait(driver, 16).until(EC.invisibility_of_element_located((By.ID, "payMoneyForm:j_idt6_modal")))
                    element41 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:j_idt41")))
                    is_amount, amount, payment_id = amount_by_cbil(cbil, element41, True)
                    if not is_amount:
                        data.append([cbil, amount, Config.STATUS_COMPLETE])
                        insert_ctmed(tkinp_ctmed, f"{cbil} - {amount}")
                        tkinp_ctm.delete("1.0", "1.end+1c")
                        continue
                    else:
                        payment_btn1 = WebDriverWait(driver, 16).until(EC.presence_of_element_located((By.ID, payment_id)))
                        payment_btn1.click()
                        pin_id = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:pinId")))
                        pin_id.clear()
                        pin_id.send_keys(pin)
                        pay_btn = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:btnPay")))
                        pay_btn.click()
                        try:
                            cfm_modal = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.ID, "payMoneyForm:dlgConfirm_modal")))
                            driver.execute_script("arguments[0].style.zIndex = '-99';", cfm_modal)
                        except:
                            pass
                        confirm_btn = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:yesId0")))
                        confirm_btn.click()
                        data.append([cbil, amount, Config.STATUS_COMPLETE])
                        insert_ctmed(tkinp_ctmed, f"{cbil} - {amount}")
                        tkinp_ctm.delete("1.0", "1.end+1c")
                except Exception as e:
                    data.append([cbil, 0, Config.STATUS_INCOMPLETE])
                    insert_ctmed(tkinp_ctmed, f"{cbil} - Lỗi")
                    tkinp_ctm.delete("1.0", "1.end+1c")
        time.sleep(2)
        if len(data) > 0:
            name_dir = "Thanh toán TV - Internet"
            export_excel(data, name_dir)
    except Exception as e:
        logger.error(f"Lỗi thanh toán internet: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi thanh toán internet: {e}")

def form_payment_internet():
    cus_frm = tk.Frame(root)
    cus_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    ctm_frm = tk.Frame(cus_frm)
    ctm_frm.pack(expand=True, side="left")
    ctmed_frm = tk.Frame(cus_frm)
    ctmed_frm.pack(expand=True, side="right")
    phnum_frm = tk.Frame(root)
    phnum_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    btn_frm = tk.Frame(root)
    btn_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    tklbl_ctm = tk.Label(ctm_frm, text="Mã thuê bao")
    tklbl_ctm.pack(side="top")
    tkinp_ctm = tk.Text(ctm_frm, height=12, width=24)
    tkinp_ctm.pack(side="left", pady=8)
    tklbl_ctm = tk.Label(ctmed_frm, text="Đã xử lý")
    tklbl_ctm.pack(side="top")
    tkinp_ctmed = tk.Text(ctmed_frm, width=32, height=12, bg="#ccc", state="disabled")
    tkinp_ctmed.pack(side="left", pady=8)
    tklbl_pin = tk.Label(phnum_frm, text="Mã pin:")
    tklbl_pin.pack(side="left")
    tkinp_pin = ttk.Entry(phnum_frm, width=22)
    tkinp_pin.pack(side="left", padx=4)
    style = ttk.Style()
    style.configure("Red.TButton", foreground="red")
    style.configure("Blue.TButton", foreground="blue")
    style.configure("Green.TButton", foreground="green")
    tkbtn_get_data = ttk.Button(btn_frm, text="Get dữ liệu", command=lambda: get_data_tv_internet(tkinp_ctm, tkinp_pin))
    tkbtn_get_data.pack(side='left', padx=5, pady=5)
    tkbtn_get_data.configure(style="Green.TButton")
    tkbtn_payment = ttk.Button(btn_frm, text="Bắt đầu", command=lambda: payment_internet(tkinp_ctm, tkinp_ctmed, tkinp_pin))
    tkbtn_payment.pack(side='left', padx=5, pady=5)
    tkbtn_payment.configure(style="Blue.TButton") 
    tkbtn_destroy = ttk.Button(btn_frm, text="Dừng lại", command=stop_tool)
    tkbtn_destroy.pack(side='right', padx=5, pady=5)
    tkbtn_destroy.configure(style="Red.TButton") 

def lookup_ftth(tkinp_ctm, tkinp_ctmed):
    try:
        delete_ctmed(tkinp_ctmed)
        update_stop_flag()
        cbils = tkinp_ctm.get("1.0", "end-1c").splitlines()
        if not valid_data([cbils]):
            return False
        data = []
        for cbil in cbils:
            root.update()
            time.sleep(1)
            cbil = cbil.strip()
            if not stop_flag and cbil.strip() != "":
                try:
                    customer = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:contractCode")))
                    customer.clear()
                    customer.send_keys(cbil)
                    payment_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:btnPay0")))
                    payment_button.click()
                    time.sleep(1)
                    WebDriverWait(driver, 16).until(EC.invisibility_of_element_located((By.ID, "payMoneyForm:j_idt6_modal")))
                    element41 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:j_idt41")))
                    is_amount, amount, payment_id = amount_by_cbil(cbil, element41, False)
                    data.append([cbil, amount, Config.STATUS_COMPLETE])
                    insert_ctmed(tkinp_ctmed, f"{cbil} - {amount}")
                    tkinp_ctm.delete("1.0", "1.end+1c")
                    continue
                except Exception as e:
                    data.append([cbil, 0, Config.STATUS_INCOMPLETE])
                    insert_ctmed(tkinp_ctmed, f"{cbil} - Lỗi")
                    tkinp_ctm.delete("1.0", "1.end+1c")
                    continue
        time.sleep(2)
        if len(data) > 0:
            name_dir = "Tra cứu FTTH"
            export_excel(data, name_dir)
    except Exception as e:
        logger.error(f"Lỗi tra cứu FTTH: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi tra cứu FTTH: {e}")

def form_lookup_ftth():
    cus_frm = tk.Frame(root)
    cus_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    ctm_frm = tk.Frame(cus_frm)
    ctm_frm.pack(expand=True, side="left")
    ctmed_frm = tk.Frame(cus_frm)
    ctmed_frm.pack(expand=True, side="right")
    btn_frm = tk.Frame(root)
    btn_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    tklbl_ctm = tk.Label(ctm_frm, text="Số thuê bao")
    tklbl_ctm.pack(side="top")
    tkinp_ctm = tk.Text(ctm_frm, height=16, width=24)
    tkinp_ctm.pack(side="left", pady=8)
    tklbl_ctm = tk.Label(ctmed_frm, text="Đã xử lý")
    tklbl_ctm.pack(side="top")
    tkinp_ctmed = tk.Text(ctmed_frm, height=16, width=32, bg="#ccc", state="disabled")
    tkinp_ctmed.pack(side="left", pady=8)
    style = ttk.Style()
    style.configure("Red.TButton", foreground="red")
    style.configure("Blue.TButton", foreground="blue")
    style.configure("Green.TButton", foreground="green")
    tkbtn_get_data = ttk.Button(btn_frm, text="Get dữ liệu", command=lambda: get_data_ftth(tkinp_ctm))
    tkbtn_get_data.pack(side='left', padx=5, pady=5)
    tkbtn_get_data.configure(style="Green.TButton")
    tkbtn_payment = ttk.Button(btn_frm, text="Bắt đầu", command=lambda: lookup_ftth(tkinp_ctm, tkinp_ctmed))
    tkbtn_payment.pack(side='left', padx=5, pady=5)
    tkbtn_payment.configure(style="Blue.TButton") 
    tkbtn_destroy = ttk.Button(btn_frm, text="Dừng lại", command=stop_tool)
    tkbtn_destroy.pack(side='right', padx=5, pady=5)
    tkbtn_destroy.configure(style="Red.TButton") 

def payment_phone(tkinp_ctm, tkinp_ctmed, tkinp_pin, tkcbb_form, tkcbb_amount):
    try:
        delete_ctmed(tkinp_ctmed)
        update_stop_flag()
        cbils = tkinp_ctm.get("1.0", "end-1c").splitlines()
        pin = tkinp_pin.get()
        cbb_type = tkcbb_form.get()
        type_sub = handle_choose_select(cbb_type)
        if type_sub == 1:
            amount = tkcbb_amount.get()
            isnext = valid_data([cbils, pin, amount])
            if isnext:
                rsl_amount = handle_choose_amount(amount)
        else:
            isnext = valid_data([cbils, pin])
        if not isnext:
            return False
        data = []
        for cbil in cbils:
            cbil = cbil.strip()
            root.update()
            time.sleep(1)
            if not stop_flag and cbil.strip() != "":
                driver.refresh()
                time.sleep(2)
                try:
                    phonenum = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:phoneNumberId")))
                    phonenum.clear()
                    phonenum.send_keys(cbil)
                    phonenum.send_keys(Keys.TAB)
                    time.sleep(1)
                except:
                    time.sleep(2)
                    phonenum = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:phoneNumberId")))
                    phonenum.clear()
                    phonenum.send_keys(cbil)
                    phonenum.send_keys(Keys.TAB)
                time.sleep(0.5)
                if type_sub == 1:
                    try:
                        try:
                            cfm_modalTT = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, "indexForm:dlgConfirmTT_modal")))
                            driver.execute_script("arguments[0].style.zIndex = '-99';", cfm_modalTT)
                            time.sleep(1)
                        except:
                            pass
                        spl_lbl = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:supplier")))
                        spl_lbl.click()
                        spl_0 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:supplier_0")))
                        spl_0.click()
                        cfm_pay = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:yesTTId")))
                        cfm_pay.click()
                    except:
                        pass
                    script = f"""
                    var element = document.querySelector('input[id="indexForm:subAmountId:{rsl_amount}"]').closest('div');
                    if (!element.classList.contains('ui-state-active')) {{
                        element.click();
                    }}
                    """
                    driver.execute_script(script)
                else:
                    try:
                        try:
                            cfm_modalTT = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.ID, "indexForm:dlgConfirmTT_modal")))
                            driver.execute_script("arguments[0].style.zIndex = '-99';", cfm_modalTT)
                        except:
                            pass
                        try:
                            time.sleep(0.5)
                            spl_lbl = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:supplier")))
                            spl_lbl.click()
                            time.sleep(0.5)
                            spl_1 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:supplier_1")))
                            spl_1.click()
                            time.sleep(0.5)
                            cfm_pay = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:yesTTId")))
                            cfm_pay.click()
                            time.sleep(0.5)
                        except:
                            pass
                        try:
                            btn_check = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:btnCheck")))
                            btn_check.click()
                        except:
                            pass
                        lbl_debt = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, "indexForm:debtId_input")))
                        debt_str = lbl_debt.get_attribute('value')
                        debt = int(debt_str.replace(".", "").replace(",", ""))
                        if debt >= 5000:
                            inp_amount = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:transAmountId_input")))
                            inp_amount.clear()
                            inp_amount.send_keys(debt)
                        else:
                            data.append([cbil, debt, stt_complete])
                            tkinp_ctm.delete("1.0", "1.end+1c")
                            insert_ctmed(tkinp_ctmed, f"{cbil} - {debt}")
                            continue
                    except:
                        data.append([cbil, 0, stt_complete])
                        tkinp_ctm.delete("1.0", "1.end+1c")
                        insert_ctmed(tkinp_ctmed, f"{cbil} - không nợ cước")
                        continue
                try:
                    pin_id = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:pinId")))
                    pin_id.clear()
                    pin_id.send_keys(pin)
                    btn_pay = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:btnPay")))
                    btn_pay.click()
                    try:
                        cfm_modal = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, "indexForm:dlgConfirm_modal")))
                        driver.execute_script("arguments[0].style.zIndex = '-99';", cfm_modal)
                    except:
                        pass
                    time.sleep(0.5)
                    btn_confirm = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:yesIdCard")))
                    btn_confirm.click()
                    if type_sub == 1:
                        data.append([cbil, amount, Config.STATUS_COMPLETE])
                        tkinp_ctm.delete("1.0", "1.end+1c")
                        insert_ctmed(tkinp_ctmed, f"{cbil} - {amount}")
                    else:
                        data.append([cbil, debt, Config.STATUS_COMPLETE])
                        tkinp_ctm.delete("1.0", "1.end+1c")
                        insert_ctmed(tkinp_ctmed, f"{cbil} - {debt}")
                except:
                    data.append([cbil, 0, Config.STATUS_INCOMPLETE])
                    tkinp_ctm.delete("1.0", "1.end+1c")
                    insert_ctmed(tkinp_ctmed, f"{cbil} - {debt}")
        time.sleep(2)
        if len(data) > 0:
            name_dir = "Nạp tiền đa mạng"
            export_excel(data, name_dir)
    except Exception as e:
        logger.error(f"Lỗi thanh toán điện thoại: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi thanh toán điện thoại: {e}")

def form_payment_phone():
    cus_frm = tk.Frame(root)
    cus_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    ctm_frm = tk.Frame(cus_frm)
    ctm_frm.pack(expand=True, side="left")
    ctmed_frm = tk.Frame(cus_frm)
    ctmed_frm.pack(expand=True, side="right")
    form_frm = tk.Frame(root)
    form_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    pin_frm = tk.Frame(root)
    pin_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    btn_frm = tk.Frame(root)
    btn_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    tklbl_ctm = tk.Label(ctm_frm, text="Số điện thoại")
    tklbl_ctm.pack(side="top")
    tkinp_ctm = tk.Text(ctm_frm, height=12, width=24)
    tkinp_ctm.pack(side="left", pady=8)
    tklbl_ctm = tk.Label(ctmed_frm, text="Đã xử lý")
    tklbl_ctm.pack(side="top")
    tkinp_ctmed = tk.Text(ctmed_frm, width=32, height=12, bg="#ccc", state="disabled")
    tkinp_ctmed.pack(side="left", pady=8)
    tklbl_form = tk.Label(form_frm, text="Hình thức:")
    tklbl_form.pack(side="left")
    tkcbb_form = ttk.Combobox(form_frm, values=["Nạp trả trước", "Gạch nợ trả sau"], width="14", state="readonly")
    tkcbb_form.pack(side="left")
    tkcbb_form.set("Nạp trả trước")
    tkcbb_form.bind("<<ComboboxSelected>>", lambda event: toggle_input_amount(tkcbb_form, tklbl_amount, tkcbb_amount))
    tkcbb_amount = ttk.Combobox(form_frm, values=["10.000đ", "20.000đ", "30.000đ", "50.000đ", "100.000đ", "200.000đ", "300.000đ", "500.000đ"], width="10", state="readonly")
    tkcbb_amount.pack(side="right")
    tklbl_amount = tk.Label(form_frm, text="Số tiền nạp:")
    tklbl_amount.pack(side="right")
    tklbl_pin = tk.Label(pin_frm, text="Mã pin:")
    tklbl_pin.pack(side="left")
    tkinp_pin = ttk.Entry(pin_frm, width=12)
    tkinp_pin.pack(side="left", padx=4)
    style = ttk.Style()
    style.configure("Red.TButton", foreground="red")
    style.configure("Blue.TButton", foreground="blue")
    style.configure("Green.TButton", foreground="green")
    tkbtn_get_data = ttk.Button(btn_frm, text="Get dữ liệu", command=lambda: get_data_multi_network(tkinp_ctm, tkinp_pin, tkcbb_form, tkcbb_amount))
    tkbtn_get_data.pack(side='left', padx=5, pady=5)
    tkbtn_get_data.configure(style="Green.TButton")
    tkbtn_payment = ttk.Button(btn_frm, text="Bắt đầu", command=lambda: payment_phone(tkinp_ctm, tkinp_ctmed, tkinp_pin, tkcbb_form, tkcbb_amount))
    tkbtn_payment.pack(side='left', padx=5, pady=5)
    tkbtn_payment.configure(style="Blue.TButton") 
    tkbtn_destroy = ttk.Button(btn_frm, text="Dừng lại", command=stop_tool)
    tkbtn_destroy.pack(side='right', padx=5, pady=5)
    tkbtn_destroy.configure(style="Red.TButton") 

def lookup_card(tkinp_ctm, tkinp_ctmed):
    try:
        delete_ctmed(tkinp_ctmed)
        update_stop_flag()
        cbils = tkinp_ctm.get("1.0", "end-1c").splitlines()
        if not valid_data([cbils]):
            return False
        data = []
        for cbil in cbils:
            cbil = cbil.strip()
            root.update()
            time.sleep(1)
            if not stop_flag and cbil.strip() != "":
                driver.refresh()
                time.sleep(2)
                try:
                    phonenum = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:phoneNumberId")))
                    phonenum.clear()
                    phonenum.send_keys(cbil)
                    phonenum.send_keys(Keys.TAB)
                    time.sleep(1)
                except:
                    time.sleep(2)
                    phonenum = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:phoneNumberId")))
                    phonenum.clear()
                    phonenum.send_keys(cbil)
                    phonenum.send_keys(Keys.TAB)
                try:
                    try:
                        cfm_modalTT = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.ID, "indexForm:dlgConfirmTT_modal")))
                        driver.execute_script("arguments[0].style.zIndex = '-99';", cfm_modalTT)
                    except:
                        pass
                    try:
                        time.sleep(0.5)
                        spl_lbl = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:supplier")))
                        spl_lbl.click()
                        time.sleep(0.5)
                        spl_1 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:supplier_1")))
                        spl_1.click()
                        time.sleep(0.5)
                        cfm_pay = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:yesTTId")))
                        cfm_pay.click()
                        time.sleep(0.5)
                    except:
                        pass
                    try:
                        btn_check = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:btnCheck")))
                        btn_check.click()
                    except:
                        pass
                    lbl_debt = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, "indexForm:debtId_input")))
                    debt_str = lbl_debt.get_attribute('value')
                    debt = int(debt_str.replace(".", "").replace(",", ""))
                    data.append([cbil, debt, Config.STATUS_COMPLETE])
                    tkinp_ctm.delete("1.0", "1.end+1c")
                    insert_ctmed(tkinp_ctmed, f"{cbil} - {debt}")
                    continue
                except:
                    data.append([cbil, "Không tìm thấy nợ cước", Config.STATUS_INCOMPLETE])
                    tkinp_ctm.delete("1.0", "1.end+1c")
                    insert_ctmed(tkinp_ctmed, f"{cbil} - null")
                    continue
        time.sleep(2)
        if len(data) > 0:
            name_dir = "Tra cứu nợ trả sau"
            export_excel(data, name_dir)
    except Exception as e:
        logger.error(f"Lỗi tra cứu nợ trả sau: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi tra cứu nợ trả sau: {e}")

def form_lookup_card():
    cus_frm = tk.Frame(root)
    cus_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    ctm_frm = tk.Frame(cus_frm)
    ctm_frm.pack(expand=True, side="left")
    ctmed_frm = tk.Frame(cus_frm)
    ctmed_frm.pack(expand=True, side="right")
    button_frm = tk.Frame(root)
    button_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    tklbl_ctm = tk.Label(ctm_frm, text="SĐT tra cứu")
    tklbl_ctm.pack(side="top")
    tkinp_ctm = tk.Text(ctm_frm, height=16, width=24)
    tkinp_ctm.pack(side="left", pady=8)
    tklbl_ctm = tk.Label(ctmed_frm, text="Đã xử lý")
    tklbl_ctm.pack(side="top")
    tkinp_ctmed = tk.Text(ctmed_frm, height=16, width=32, bg="#ccc", state="disabled")
    tkinp_ctmed.pack(side="left", pady=8)
    style = ttk.Style()
    style.configure("Red.TButton", foreground="red")
    style.configure("Blue.TButton", foreground="blue")
    style.configure("Green.TButton", foreground="green")
    tkbtn_get_data = ttk.Button(button_frm, text="Get dữ liệu", command=lambda: get_data_postpaid(tkinp_ctm))
    tkbtn_get_data.pack(side='left', padx=5, pady=5)
    tkbtn_get_data.configure(style="Green.TButton")
    tkbtn_payment = ttk.Button(button_frm, text="Bắt đầu", command=lambda: lookup_card(tkinp_ctm, tkinp_ctmed))
    tkbtn_payment.pack(side='left', padx=5, pady=5)
    tkbtn_payment.configure(style="Blue.TButton") 
    tkbtn_destroy = ttk.Button(button_frm, text="Dừng lại", command=stop_tool)
    tkbtn_destroy.pack(side='right', padx=5, pady=5)
    tkbtn_destroy.configure(style="Red.TButton") 

def payment_viettel(tkinp_ctm, tkinp_ctmed, tkinp_pin, tkinp_amount):
    try:
        delete_ctmed(tkinp_ctmed)
        update_stop_flag()
        pin = tkinp_pin.get()
        amount = tkinp_amount.get()
        cbils = tkinp_ctm.get("1.0", "end-1c").splitlines()
        if not valid_data([cbils, pin, amount]):
            return False
        data = []
        for cbil in cbils:
            root.update()
            time.sleep(0.5)
            cbil = cbil.strip()
            if not stop_flag and cbil.strip() != "":
                time.sleep(0.5)
                try:
                    phonenum = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:phoneNumberId")))
                    phonenum.clear()
                    phonenum.send_keys(cbil)
                    phonenum.send_keys(Keys.TAB)
                    time.sleep(0.5)
                except:
                    time.sleep(2)
                    phonenum = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:phoneNumberId")))
                    phonenum.clear()
                    phonenum.send_keys(cbil)
                    phonenum.send_keys(Keys.TAB)
                try:
                    cfm_modalTT = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.ID, "indexForm:dlgConfirmTT_modal")))
                    driver.execute_script("arguments[0].style.zIndex = '-99';", cfm_modalTT)
                    time.sleep(0.5)
                    spl_lbl = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.ID, "indexForm:supplier")))
                    spl_lbl.click()
                    time.sleep(0.5)
                    spl_0 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:supplier_0")))
                    spl_0.click()
                    time.sleep(0.5)
                    cfm_pay = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:yesTTId")))
                    cfm_pay.click()
                    time.sleep(0.5)
                except:
                    pass
                try:
                    inp_amount = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:transAmountId_input")))
                    inp_amount.clear()
                    inp_amount.send_keys(amount)
                    time.sleep(0.5)
                    pin_id = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:pinId")))
                    pin_id.clear()
                    pin_id.send_keys(pin)
                    time.sleep(0.5)
                    btn_pay = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:btnPay")))
                    btn_pay.click()
                    time.sleep(0.5)
                    try:
                        cfm_modal = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, "indexForm:dlgConfirm_modal")))
                        driver.execute_script("arguments[0].style.zIndex = '-99';", cfm_modal)
                    except:
                        pass
                    btn_confirm = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:yesIdCard")))
                    btn_confirm.click()
                    data.append([cbil, amount, Config.STATUS_COMPLETE])
                    insert_ctmed(tkinp_ctmed, f"{cbil} - {amount}")
                    tkinp_ctm.delete("1.0", "1.end+1c")
                except:
                    data.append([cbil, 0, Config.STATUS_INCOMPLETE])
                    insert_ctmed(tkinp_ctmed, f"{cbil} - null")
                    tkinp_ctm.delete("1.0", "1.end+1c")
                    continue
        time.sleep(2)
        if len(data) > 0:
            name_dir = "Nạp tiền mạng Viettel"
            export_excel(data, name_dir)
    except Exception as e:
        logger.error(f"Lỗi thanh toán Viettel: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi thanh toán Viettel: {e}")

def form_payment_viettel():
    cus_frm = tk.Frame(root)
    cus_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    ctm_frm = tk.Frame(cus_frm)
    ctm_frm.pack(expand=True, side="left")
    ctmed_frm = tk.Frame(cus_frm)
    ctmed_frm.pack(expand=True, side="right")
    form_frm = tk.Frame(root)
    form_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    pin_frm = tk.Frame(root)
    pin_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    button_frm = tk.Frame(root)
    button_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    tklbl_ctm = tk.Label(ctm_frm, text="Số điện thoại")
    tklbl_ctm.pack(side="top")
    tkinp_ctm = tk.Text(ctm_frm, height=12, width=24)
    tkinp_ctm.pack(side="left", pady=8)
    tklbl_ctm = tk.Label(ctmed_frm, text="Đã xử lý")
    tklbl_ctm.pack(side="top")
    tkinp_ctmed = tk.Text(ctmed_frm, width=32, height=12, bg="#ccc", state="disabled")
    tkinp_ctmed.pack(side="left", pady=8)
    tkinp_amount = tk.Entry(pin_frm, width=12)
    tkinp_amount.pack(side="right", padx=4)
    tklbl_amount = tk.Label(pin_frm, text="Số tiền nạp:")
    tklbl_amount.pack(side="right")
    tklbl_pin = tk.Label(pin_frm, text="Mã pin:")
    tklbl_pin.pack(side="left")
    tkinp_pin = ttk.Entry(pin_frm, width=12)
    tkinp_pin.pack(side="left", padx=4)
    style = ttk.Style()
    style.configure("Red.TButton", foreground="red")
    style.configure("Blue.TButton", foreground="blue")
    style.configure("Green.TButton", foreground="green")
    tkbtn_get_data = ttk.Button(button_frm, text="Get dữ liệu", command=lambda: get_data_viettel(tkinp_ctm, tkinp_pin, tkinp_amount))
    tkbtn_get_data.pack(side='left', padx=5, pady=5)
    tkbtn_get_data.configure(style="Green.TButton")
    tkbtn_payment = ttk.Button(button_frm, text="Bắt đầu", command=lambda: payment_viettel(tkinp_ctm, tkinp_ctmed, tkinp_pin, tkinp_amount))
    tkbtn_payment.pack(side='left', padx=5, pady=5)
    tkbtn_payment.configure(style="Blue.TButton") 
    tkbtn_destroy = ttk.Button(button_frm, text="Dừng lại", command=stop_tool)
    tkbtn_destroy.pack(side='right', padx=5, pady=5)
    tkbtn_destroy.configure(style="Red.TButton") 

def debt_electric(tkinp_ctm, tkinp_ctmed, tkinp_phone, tkinp_pin):
    try:
        delete_ctmed(tkinp_ctmed)
        update_stop_flag()
        pin = tkinp_pin.get()
        phone = tkinp_phone.get()
        cbils = tkinp_ctm.get("1.0", "end-1c").splitlines()
        if not valid_data([cbils, pin, phone]):
            return False
        data = []
        for cbil in cbils:
            cbil = cbil.strip()
            root.update()
            time.sleep(1)
            if not stop_flag and cbil.strip() != "":
                try:
                    customer = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:billCodeId")))
                    customer.clear()
                    customer.send_keys(cbil)
                    time.sleep(0.5)
                    phonenumber = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:benMsisdnId")))
                    phonenumber.clear()
                    phonenumber.send_keys(phone)
                    time.sleep(0.5)
                    pinid = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:pinId")))
                    pinid.clear()
                    pinid.send_keys(pin)
                    time.sleep(0.5)
                    payment = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:btnPay")))
                    payment.click()
                    time.sleep(0.5)
                    try:
                        cfm_modal = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, "payMoneyForm:dlgConfirm_modal")))
                        driver.execute_script("arguments[0].style.zIndex = '-99';", cfm_modal)
                    except:
                        pass
                    lblamount = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:j_idt49")))
                    try:
                        text_of_amount = lblamount.text
                        amount_str = text_of_amount.replace('VND', '').replace('.', '')
                        amount = int(amount_str)
                    except:
                        amount = lblamount.text
                    time.sleep(0.5)
                    confirm = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:yesIdEVN")))
                    confirm.click()
                    data.append([cbil, amount, Config.STATUS_COMPLETE])
                    tkinp_ctm.delete("1.0", "1.end+1c")
                    insert_ctmed(tkinp_ctmed, f"{cbil} - {amount}")
                    continue
                except Exception as e:
                    data.append([cbil, amount, Config.STATUS_INCOMPLETE])
                    tkinp_ctm.delete("1.0", "1.end+1c")
                    insert_ctmed(tkinp_ctmed, f"{cbil} - null")
                    continue
        time.sleep(2)
        if len(data) > 0:
            name_dir = "Thanh toán điện EVN"
            export_excel(data, name_dir)
    except Exception as e:
        logger.error(f"Lỗi thanh toán điện EVN: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi thanh toán điện EVN: {e}")

def form_debt_electric():
    cus_frm = tk.Frame(root)
    cus_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    ctm_frm = tk.Frame(cus_frm)
    ctm_frm.pack(expand=True, side="left")
    ctmed_frm = tk.Frame(cus_frm)
    ctmed_frm.pack(expand=True, side="right")
    phnum_frm = tk.Frame(root)
    phnum_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    btn_frm = tk.Frame(root)
    btn_frm.pack(expand=True, side="top", padx=14, pady=8, fill="both")
    tklbl_ctm = tk.Label(ctm_frm, text="Mã thuê bao")
    tklbl_ctm.pack(side="top")
    tkinp_ctm = tk.Text(ctm_frm, height=18, width=24)
    tkinp_ctm.pack(side="left", pady=8)
    tklbl_ctm = tk.Label(ctmed_frm, text="Đã xử lý")
    tklbl_ctm.pack(side="top")
    tkinp_ctmed = tk.Text(ctmed_frm, width=32, height=18, bg="#ccc", state="disabled")
    tkinp_ctmed.pack(side="left", pady=8)
    tklbl_phone = tk.Label(phnum_frm, text="SĐT người nhận:")
    tklbl_phone.pack(side="left")
    tkinp_phone = ttk.Entry(phnum_frm, width=16)
    tkinp_phone.pack(side="left", padx=4)
    tkinp_pin = ttk.Entry(phnum_frm, width=12)
    tkinp_pin.pack(side="right", padx=4)
    tklbl_pin = tk.Label(phnum_frm, text="Mã pin:")
    tklbl_pin.pack(side="right")
    style = ttk.Style()
    style.configure("Red.TButton", foreground="red")
    style.configure("Blue.TButton", foreground="blue")
    style.configure("Green.TButton", foreground="green")
    tkbtn_get_data = ttk.Button(btn_frm, text="Get dữ liệu", command=lambda: get_data_evn(tkinp_ctm, tkinp_phone, tkinp_pin))
    tkbtn_get_data.pack(side='left', padx=5, pady=5)
    tkbtn_get_data.configure(style="Green.TButton")
    tkbtn_payment = ttk.Button(btn_frm, text="Bắt đầu", command=lambda: debt_electric(tkinp_ctm, tkinp_ctmed, tkinp_phone, tkinp_pin))
    tkbtn_payment.pack(side='left', padx=5, pady=5)
    tkbtn_payment.configure(style="Blue.TButton") 
    tkbtn_destroy = ttk.Button(btn_frm, text="Dừng lại", command=stop_tool)
    tkbtn_destroy.pack(side='right', padx=5, pady=5)
    tkbtn_destroy.configure(style="Red.TButton") 

root = tk.Tk()
root.title("HPK Tool - Viettel Pay Automation")
root.geometry("500x550") 
root.option_add("*Font", "Arial 10")
try:
    root.iconbitmap(Config.ICON_FILE)
except Exception as e:
    logger.warning(f"Không thể tải icon: {e}")
    pass

def clear_widgets(main_frm):
    for widget in root.winfo_children():
        if widget is not main_frm:
            widget.destroy()

def read_config():
    try:
        if not os.path.exists(Config.CONFIG_FILE):
            with open(Config.CONFIG_FILE, "w", encoding='utf-8') as json_file:
                json.dump({"files": ""}, json_file, indent=4, ensure_ascii=False)
        with open(Config.CONFIG_FILE, "r", encoding='utf-8') as json_file: 
            data = json.load(json_file)
            if not "files" in data:
                data["files"] = ""
        with open(Config.CONFIG_FILE, "w", encoding='utf-8') as file:
            json.dump(data, file, indent=4, ensure_ascii=False)
        files = data.get("files")
        return files
    except Exception as e:
        logger.error(f"Lỗi đọc cấu hình: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi đọc cấu hình: {e}")
        return ""

def connect_database():
    try:
        client = MongoClient(Config.MONGODB_URI)
        db = client["Tools"]
        collection = db["Users"]
        return collection
    except Exception as e:
        logger.error(f"Lỗi kết nối database: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi kết nối database: {e}")
        return None

def set_file_config(files):
    try:
        with open(Config.CONFIG_FILE, "r", encoding='utf-8') as file:
            data = json.load(file)
            data["files"] = files
        with open(Config.CONFIG_FILE, "w", encoding='utf-8') as file:
            json.dump(data, file, indent=4, ensure_ascii=False)
    except Exception as e:
        logger.error(f"Lỗi lưu cấu hình: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi lưu cấu hình: {e}")

def handle_key_active(key):
    try:
        key = (key.get()).strip()
        data = collection.find_one({"key": key})
        if data and not data.get("active"):
            collection.find_one_and_update({"key": key}, {"$set": {"active": True}})
            set_file_config(data.get("files"))
            #messagebox.showinfo(Config.TITLE, "Đã kích hoạt thành công")
            root.destroy()
            root.quit()
            main()
        elif data and data.get("active"):
            #messagebox.showwarning(Config.TITLE, "Mã kích hoạt đã được sử dụng")
        else:
            #messagebox.showerror(Config.TITLE, "Mã kích hoạt không tồn tại")
    except Exception as e:
        logger.error(f"Lỗi kích hoạt key: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi kích hoạt: {e}")

def check_exits_key(key):
    if not key:
        show_activate_form()
        return False
    return True

def show_activate_form():
    active_frm = tk.Frame(root)
    active_frm.pack(expand=True, side="top", padx=6, pady=6, fill="both")
    tklbl_active = tk.Label(active_frm, text="Key:")
    tklbl_active.pack(side="left", padx=4)
    tkinp_active = tk.Entry(active_frm, width=46)
    tkinp_active.pack(side="left", padx=4)
    tkbtn_active = ttk.Button(active_frm, text="Kích hoạt", command=lambda: handle_key_active(tkinp_active))
    tkbtn_active.pack(side='left', padx=4)

def initialize_browser(username="default"):
    global driver
    driver = get_chrome_driver(username)
    if driver:
        driver.get(Config.DRIVER_LINK)
        return driver
    return None

def cleanup():
    global driver
    if driver:
        try:
            driver.quit()
        except Exception as e:
            logger.warning(f"Lỗi khi đóng driver: {e}")
        finally:
            driver = None

def is_logged_in(driver):
    try:
        WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, "loginForm:userName")))
        return False
    except:
        try:
            WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.CLASS_NAME, "dl-info-detail")))
            return True
        except:
            return False

def login_process(dbfiles):
    try:
        usr = dbfiles.get("username", "")
        pwd = dbfiles.get("password", "")
        inp_usr = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, "loginForm:userName")))
        inp_usr.clear() 
        inp_usr.send_keys(usr)
        inp_pwd = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, "loginForm:password")))
        inp_pwd.clear() 
        inp_pwd.send_keys(pwd)
    except Exception as e:
        logger.warning(f"Lỗi đăng nhập: {e}")
        pass

def show_services_form():
    try:
        main_frm = tk.Frame(root)
        main_frm.pack(expand=True, side="top", padx=6, pady=6, fill="both")
        tklbl_choose = tk.Label(main_frm, text="Loại thanh toán:")
        tklbl_choose.pack(side="left")
        tkcbb_choose = ttk.Combobox(main_frm, values=[
            "Tra cứu FTTH",
            "Gạch điện EVN", 
            "Nạp tiền đa mạng",
            "Nạp tiền mạng Viettel",
            "Thanh toán TV - Internet",
            "Tra cứu nợ thuê bao trả sau"
        ], width="32", state="readonly")
        tkcbb_choose.pack(side="left", padx=6, expand=True, fill="x")
        tkcbb_choose.set("Tra cứu FTTH")
        def handle_choose_services(event, choose, main_frm):
            service = choose.get()
            clear_widgets(main_frm)
            if service == "Tra cứu FTTH":
                form_lookup_ftth()
            elif service == "Gạch điện EVN":
                form_debt_electric() 
            elif service == "Nạp tiền đa mạng":
                form_payment_phone()
            elif service == "Nạp tiền mạng Viettel":
                form_payment_viettel()
            elif service == "Thanh toán TV - Internet":
                form_payment_internet()
            elif service == "Tra cứu nợ thuê bao trả sau":
                form_lookup_card()
        tkcbb_choose.bind("<<ComboboxSelected>>", lambda event: handle_choose_services(event, tkcbb_choose, main_frm))
        handle_choose_services(None, tkcbb_choose, main_frm)
    except Exception as e:
        logger.error(f"Lỗi hiển thị form dịch vụ: {e}")
        #messagebox.showerror(Config.TITLE, f"Lỗi hiển thị form dịch vụ: {e}")

def main():
    global dbfiles, times_exits
    
    # Kiểm tra API có sẵn, nếu không có thì khởi động mock server nội bộ
    if check_api_health():
        logger.info("API server sẵn sàng, sử dụng nguồn dữ liệu bên ngoài")
    else:
        logger.info("Đang khởi động API server mock...")
        api_started = start_api_server()
        if api_started:
            logger.info("API server mock đã khởi động thành công")
        else:
            logger.warning("Không thể khởi động API server mock, chức năng Get dữ liệu sẽ không hoạt động")
    
    files_data = read_config()
    if check_exits_key(files_data):
        try:
            dbfiles = encode_json()
            if not dbfiles:
                #messagebox.showerror(Config.TITLE, "Không thể giải mã file cấu hình")
                show_activate_form()
            else:
                username = dbfiles.get("username", "default")
                driver = initialize_browser(username)
                if not driver:
                    #messagebox.showerror(Config.TITLE, "Không thể khởi tạo trình duyệt, vui lòng thao tác thủ công")
                    show_services_form()
                else:
                    login_process(dbfiles)
                    try:
                        times_curr = get_number_uses()
                        times_exits = times_curr[1]
                        show_services_form()
                    except Exception as e:
                        logger.error(f"Lỗi lấy thông tin sử dụng: {e}")
                        #messagebox.showerror(Config.TITLE, f"Lỗi lấy thông tin sử dụng: {e}")
                        show_services_form()
        except Exception as e:
            logger.error(f"Lỗi khởi tạo: {e}")
            #messagebox.showerror(Config.TITLE, f"Lỗi khởi tạo: {e}")
            show_services_form()
    else:
        global collection
        collection = connect_database()
        show_activate_form()

if __name__ == "__main__":
    try:
        main()
        root.protocol("WM_DELETE_WINDOW", lambda: [cleanup(), root.destroy()])
        root.mainloop()
    except Exception as e:
        logger.error(f"Lỗi chính: {e}")
        #messagebox.showerror("Lỗi", f"Lỗi khởi động ứng dụng: {e}")
    finally:
        cleanup()